import base64
import hashlib
import io
import json
import logging
import os
import secrets
import shutil
import string
import sys
import time  # Thêm import time ở đầu file
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd
from fastapi import Request  # Thêm import này nếu chưa có
from fastapi import (Depends, FastAPI, HTTPException, Query, WebSocket,
                     WebSocketDisconnect)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, validator
from sqlalchemy import (Boolean, Column, DateTime, Float, ForeignKey, Integer,
                        String, Text, case, create_engine, func, or_)
from sqlalchemy.orm import (Session, declarative_base, joinedload,
                            relationship, sessionmaker)

# Suppress warnings
warnings.filterwarnings('ignore')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== DATABASE SETUP ====================
SQLALCHEMY_DATABASE_URL = "sqlite:///./classroom_ai.db"

engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    pool_pre_ping=True,
    pool_recycle=3600,
)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# ==================== DATABASE MODELS (CHỈ 2 BẢNG) ====================

class User(Base):
    """Bảng tài khoản đăng nhập với 2 phân quyền: admin và user"""
    __tablename__ = "users"
    
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True, nullable=False)
    email = Column(String, unique=True, index=True)
    hashed_password = Column(String, nullable=False)
    full_name = Column(String)
    is_active = Column(Boolean, default=True)
    is_admin = Column(Boolean, default=False)  # True = admin, False = user
    created_at = Column(DateTime, default=datetime.utcnow)

class StudentData(Base):
    """Bảng duy nhất cho dữ liệu học sinh: điểm danh, cảm xúc, hành vi, độ tập trung"""
    __tablename__ = "student_data"
    
    id = Column(Integer, primary_key=True, index=True)
    student_id = Column(String, index=True, nullable=False)
    student_name = Column(String, index=True)
    
    # Dữ liệu điểm danh
    date = Column(String, index=True)  # Ngày (YYYY-MM-DD)
    check_in_time = Column(DateTime, nullable=True)
    check_out_time = Column(DateTime, nullable=True)
    attendance_status = Column(String)  # present, absent, late, excused
    attendance_notes = Column(Text, nullable=True)
    
    # Dữ liệu cảm xúc
    emotion = Column(String, nullable=True)  # happy, sad, angry, neutral, surprised, disgusted, fearful
    emotion_confidence = Column(Float, nullable=True)  # 0-1
    
    # Dữ liệu hành vi và độ tập trung
    behavior_type = Column(String, nullable=True)  # engagement, participation, discipline, focus
    behavior_score = Column(Float, nullable=True)  # 0-100
    behavior_details = Column(Text, nullable=True)
    
    # Metadata
    class_name = Column(String, nullable=True)
    session_id = Column(String, nullable=True)  # For grouping data in same session
    recorded_by = Column(String, nullable=True)  # Người ghi nhận (có thể là teacher_id)
    recorded_at = Column(DateTime, default=datetime.utcnow)
    
    # Thêm các trường mới cho độ tập trung
    focus_score = Column(Float, nullable=True)  # 0-100 điểm độ tập trung
    concentration_level = Column(String, nullable=True)  # high, medium, low
    focus_duration = Column(Float, nullable=True)  # Thời gian tập trung (phút)
        
# ==================== DATABASE MODELS - THÊM BẢNG MỚI ====================

class ClassStudent(Base):
    """Bảng danh sách học sinh cố định của lớp học"""
    __tablename__ = "class_students"
    
    id = Column(Integer, primary_key=True, index=True)
    student_id = Column(String, unique=True, index=True, nullable=False)
    student_name = Column(String, index=True, nullable=False)
    student_code = Column(String, index=True)  # Mã học sinh
    class_name = Column(String, index=True, nullable=False)  # Tên lớp
    gender = Column(String, nullable=True)  # Giới tính
    date_of_birth = Column(String, nullable=True)  # Ngày sinh
    address = Column(String, nullable=True)  # Địa chỉ
    phone = Column(String, nullable=True)  # Số điện thoại
    email = Column(String, nullable=True)  # Email
    parent_name = Column(String, nullable=True)  # Tên phụ huynh
    parent_phone = Column(String, nullable=True)  # SĐT phụ huynh
    
    is_active = Column(Boolean, default=True)  # Còn học hay không
    enrollment_date = Column(String, nullable=True)  # Ngày nhập học
    
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
class StudentFeedback(Base):
    """Bảng lưu feedback từ học sinh"""
    __tablename__ = "student_feedback"
    
    id = Column(Integer, primary_key=True, index=True)
    student_id = Column(String, index=True, nullable=False)
    student_name = Column(String, index=True, nullable=False)
    
    # Feedback content
    feedback_text = Column(Text, nullable=True)  # Direct text feedback
    feedback_type = Column(String, nullable=False)  # text, voice
    
    # For voice feedback
    audio_path = Column(String, nullable=True)  # Path to audio file
    audio_duration = Column(Float, nullable=True)  # Duration in seconds
    transcribed_text = Column(Text, nullable=True)  # Transcribed text from STT
    confidence = Column(Float, nullable=True)  # STT confidence score
    
    # Metadata
    emotion = Column(String, nullable=True)
    rating = Column(Integer, nullable=True)  # 1-5 stars
    class_name = Column(String, nullable=True)
    session_id = Column(String, nullable=True)
    recorded_by = Column(String, nullable=True)  # "student", "teacher", "ai"
    
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
# Tạo tất cả bảng
Base.metadata.create_all(bind=engine)

# ==================== FEEDBACK MODELS ====================
class TextFeedbackCreate(BaseModel):
    student_id: str
    student_name: str
    feedback_text: str
    feedback_type: str = "text"  # text, voice
    emotion: Optional[str] = None
    rating: Optional[int] = None  # 1-5
    class_name: Optional[str] = None
    session_id: Optional[str] = None

# Sửa model VoiceFeedbackCreate
class VoiceFeedbackCreate(BaseModel):
    student_id: str
    student_name: str
    audio_base64: str  # Base64 encoded audio
    audio_format: Optional[str] = "wav"  # wav, webm, mp3
    feedback_type: Optional[str] = "voice"
    class_name: Optional[str] = None
    session_id: Optional[str] = None
    
    @validator('audio_base64')
    def validate_audio_base64(cls, v):
        if not v or len(v) < 100:  # Minimum length for base64 audio
            raise ValueError('Audio base64 data is too short')
        return v

class FeedbackResponse(BaseModel):
    id: int
    student_id: str
    student_name: str
    feedback_text: Optional[str]
    feedback_type: str
    transcribed_text: Optional[str]
    emotion: Optional[str]
    rating: Optional[int]
    audio_path: Optional[str]
    audio_duration: Optional[float]
    confidence: Optional[float]  # STT confidence
    class_name: Optional[str]
    session_id: Optional[str]
    created_at: datetime
    
# ==================== PYDANTIC MODELS ====================
class ResetDatabaseRequest(BaseModel):
    confirm: bool
    create_sample_data: bool = True
    
class ClassStudentCreate(BaseModel):
    student_id: str
    student_name: str
    class_name: str
    student_code: Optional[str] = None
    gender: Optional[str] = None
    date_of_birth: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    is_active: Optional[bool] = True
    enrollment_date: Optional[str] = None

class ClassStudentUpdate(BaseModel):
    student_name: Optional[str] = None
    class_name: Optional[str] = None
    student_code: Optional[str] = None
    gender: Optional[str] = None
    date_of_birth: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    is_active: Optional[bool] = None
    enrollment_date: Optional[str] = None

class ClassStudentResponse(BaseModel):
    id: int
    student_id: str
    student_name: str
    student_code: Optional[str]
    class_name: str
    gender: Optional[str]
    date_of_birth: Optional[str]
    address: Optional[str]
    phone: Optional[str]
    email: Optional[str]
    parent_name: Optional[str]
    parent_phone: Optional[str]
    is_active: bool
    enrollment_date: Optional[str]
    created_at: datetime
    updated_at: Optional[datetime]
    
class StudentCreate(BaseModel):
    student_name: str
    class_name: str
    is_active: bool = True

class StudentUpdate(BaseModel):
    student_name: Optional[str] = None
    class_name: Optional[str] = None
    is_active: Optional[bool] = None

class UserCreate(BaseModel):
    username: str
    email: str
    password: str
    full_name: str
    is_admin: bool = False

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: int
    username: str
    email: str
    full_name: str
    is_active: bool
    is_admin: bool
    created_at: datetime

class Token(BaseModel):
    access_token: str
    token_type: str
    expires_in: int
    user: UserResponse

class StudentDataCreate(BaseModel):
    student_id: str
    student_name: str
    date: str
    attendance_status: str
    emotion: Optional[str] = None
    emotion_confidence: Optional[float] = None
    behavior_type: Optional[str] = None
    behavior_score: Optional[float] = None
    behavior_details: Optional[str] = None
    class_name: Optional[str] = None
    session_id: Optional[str] = None
    recorded_by: Optional[str] = None
    check_in_time: Optional[datetime] = None
    check_out_time: Optional[datetime] = None
    attendance_notes: Optional[str] = None
    focus_score: Optional[float] = None
    concentration_level: Optional[str] = None
    focus_duration: Optional[float] = None

class StudentDataUpdate(BaseModel):
    student_id: Optional[str] = None
    student_name: Optional[str] = None
    date: Optional[str] = None
    attendance_status: Optional[str] = None
    emotion: Optional[str] = None
    emotion_confidence: Optional[float] = None
    behavior_type: Optional[str] = None
    behavior_score: Optional[float] = None
    behavior_details: Optional[str] = None
    class_name: Optional[str] = None
    session_id: Optional[str] = None
    recorded_by: Optional[str] = None
    check_in_time: Optional[datetime] = None
    check_out_time: Optional[datetime] = None
    attendance_notes: Optional[str] = None
    focus_score: Optional[float] = None
    concentration_level: Optional[str] = None
    focus_duration: Optional[float] = None

class AttendanceCreate(BaseModel):
    student_id: str
    student_name: str
    date: str
    attendance_status: str
    check_in_time: Optional[datetime] = None
    attendance_notes: Optional[str] = None
    class_name: Optional[str] = None
    recorded_by: Optional[str] = None

class EmotionCreate(BaseModel):
    student_id: str
    student_name: str
    emotion: str
    emotion_confidence: float
    date: Optional[str] = None
    session_id: Optional[str] = None
    recorded_by: Optional[str] = None

class BehaviorCreate(BaseModel):
    student_id: str
    student_name: str
    behavior_type: str
    behavior_score: float
    behavior_details: Optional[str] = None
    date: Optional[str] = None
    session_id: Optional[str] = None
    recorded_by: Optional[str] = None

class FocusCreate(BaseModel):
    """Model tạo bản ghi độ tập trung với cả emotion và behavior"""
    student_id: str
    student_name: str
    focus_score: float
    concentration_level: str  # high, medium, low
    focus_duration: Optional[float] = None
    
    # Thêm các field mới
    emotion: Optional[str] = None
    emotion_confidence: Optional[float] = None
    behavior_type: Optional[str] = None
    behavior_score: Optional[float] = None
    behavior_details: Optional[str] = None
    
    # Metadata
    date: Optional[str] = None
    session_id: Optional[str] = None
    recorded_by: Optional[str] = None
    class_name: Optional[str] = None
    attendance_status: Optional[str] = None  # Thêm trường điểm danh
    check_in_time: Optional[datetime] = None

# ==================== FIX: ADD MORE FLEXIBLE MODELS FOR AI ====================
class AIStudentDataCreate(BaseModel):
    """Model linh hoạt hơn cho dữ liệu từ AI system"""
    student_id: Optional[str] = None
    student_name: Optional[str] = None
    student_code: Optional[str] = None
    name: Optional[str] = None  # Alias cho student_name
    date: Optional[str] = None
    attendance_status: Optional[str] = "present"
    status: Optional[str] = None  # Alias cho attendance_status
    check_in_time: Optional[datetime] = None
    check_out_time: Optional[datetime] = None
    attendance_notes: Optional[str] = None
    notes: Optional[str] = None  # Alias cho attendance_notes
    emotion: Optional[str] = None
    emotion_confidence: Optional[float] = None
    confidence: Optional[float] = None  # Alias cho emotion_confidence
    behavior_type: Optional[str] = None
    behavior_score: Optional[float] = None
    score: Optional[float] = None  # Alias cho behavior_score
    behavior_details: Optional[str] = None
    details: Optional[str] = None  # Alias cho behavior_details
    class_name: Optional[str] = None
    class_id: Optional[int] = None
    session_id: Optional[str] = None
    recorded_by: Optional[str] = None
    teacher_id: Optional[int] = None
    focus_score: Optional[float] = None
    concentration_level: Optional[str] = None
    focus_duration: Optional[float] = None
    engagement: Optional[float] = None  # Có thể map sang focus_score
    behavior: Optional[str] = None  # Có thể map sang behavior_details

# ==================== FASTAPI APP ====================
app = FastAPI(
    title="Classroom Management System API",
    description="Hệ thống quản lý lớp học với điểm danh và theo dõi hành vi, cảm xúc, độ tập trung",
    version="1.0.0",
    docs_url="/api/docs",
    openapi_url="/api/openapi.json",
    redoc_url="/api/redoc"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== HELPER FUNCTIONS ====================
def is_unknown_student(student_name: str, student_id: str = "") -> bool:
    """Kiểm tra xem student có phải là unknown không"""
    unknown_keywords = [
        'unknown', 'Unknown', 'UNKNOWN',
        'unknow', 'Unknow', 'UNKNOW',
        'không rõ', 'Không rõ', 'KHÔNG RÕ',
        'chưa biết', 'Chưa biết', 'CHƯA BIẾT',
        '', None
    ]
    
    # Kiểm tra student_name
    if not student_name or student_name.strip() == "":
        return True
    
    student_name_lower = student_name.lower()
    
    # Kiểm tra các từ khóa unknown
    for keyword in unknown_keywords:
        if keyword and keyword.lower() in student_name_lower:
            return True
    
    # Kiểm tra pattern cụ thể
    unknown_patterns = [
        'unknown student',
        'unknow student', 
        'unknown face',
        'unknow face',
        'face unknown',
        'student unknown'
    ]
    
    for pattern in unknown_patterns:
        if pattern in student_name_lower:
            return True
    
    # Kiểm tra student_id nếu có
    if student_id:
        student_id_lower = student_id.lower()
        if 'unknown' in student_id_lower or 'unknow' in student_id_lower:
            return True
    
    return False

# ==================== DATABASE UTILS ====================
def get_db():
    db = SessionLocal()
    try:
        yield db
    except Exception as e:
        logger.error(f"Database connection error: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Database connection error")
    finally:
        db.close()

def get_user_by_username(db: Session, username: str):
    return db.query(User).filter(User.username == username).first()

def get_user_by_email(db: Session, email: str):
    return db.query(User).filter(User.email == email).first()

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def is_hashed_password(stored_password: Optional[str]) -> bool:
    if not stored_password or len(stored_password) != 64:
        return False
    return all(char in string.hexdigits for char in stored_password)

def verify_password(plain_password: str, stored_password: Optional[str]) -> bool:
    if not stored_password:
        return False
    if is_hashed_password(stored_password):
        return hash_password(plain_password) == stored_password
    return plain_password == stored_password

def authenticate_user(db: Session, username: str, password: str):
    user = get_user_by_username(db, username)
    if not user:
        return None
    if not verify_password(password, user.hashed_password):
        return None
    if not is_hashed_password(user.hashed_password):
        user.hashed_password = hash_password(password)
        db.commit()
        db.refresh(user)
    return user

# ==================== AUTHENTICATION & AUTHORIZATION DEPENDENCIES ====================

# Mock token storage (trong thực tế dùng JWT)
active_tokens = {}

def create_access_token(user: User):
    # Đơn giản: tạo token random cho demo
    token = secrets.token_hex(32)
    
    # Lưu token vào active_tokens (trong thực tế dùng Redis hoặc database)
    active_tokens[token] = {
        "user_id": user.id,
        "username": user.username,
        "is_admin": user.is_admin,
        "created_at": datetime.now().isoformat()
    }
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "expires_in": 3600 * 24,  # 24 hours
        "user": {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "full_name": user.full_name,
            "is_active": user.is_active,
            "is_admin": user.is_admin,
            "created_at": user.created_at.isoformat()
        }
    }

def validate_token_and_get_user(token: str, db: Session):
    """Validate token và trả về user - function có thể gọi trực tiếp"""
    if not token:
        raise HTTPException(status_code=401, detail="Token không tồn tại")
    
    # Trong thực tế, bạn cần decode JWT token
    # Ở đây dùng mock token system
    if token not in active_tokens:
        raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc đã hết hạn")
    
    user_id = active_tokens[token]["user_id"]
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="Người dùng không tồn tại hoặc đã bị khóa")
    
    return user

async def get_current_user_from_token(token: str = Query(None, alias="token"), db: Session = Depends(get_db)):
    """Lấy thông tin user từ token - dependency cho FastAPI"""
    return validate_token_and_get_user(token, db)

async def require_admin(user: User = Depends(get_current_user_from_token)):
    """Dependency yêu cầu quyền admin"""
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Không có quyền truy cập. Chỉ dành cho Admin.")
    return user

async def require_user(user: User = Depends(get_current_user_from_token)):
    """Dependency yêu cầu đăng nhập (cả admin và user thường)"""
    return user

# ==================== WEBSOCKET MANAGER ====================
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def send_personal_message(self, message: str, websocket: WebSocket):
        await websocket.send_text(message)

    async def broadcast(self, message: Any):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

manager_ws = ConnectionManager()

# ==================== FEEDBACK ENDPOINTS ====================
async def ensure_wav_format(audio_path: str) -> str:
    """
    Đảm bảo audio file là WAV format chuẩn cho SpeechRecognition
    """
    try:
        import struct
        import wave

        # Kiểm tra nếu đã là WAV chuẩn
        try:
            with wave.open(audio_path, 'rb') as wav_file:
                # Check if it's proper WAV
                if wav_file.getnchannels() in [1, 2] and wav_file.getsampwidth() in [1, 2]:
                    logger.info(f"✅ File is proper WAV: {audio_path}")
                    return audio_path
        except:
            pass  # Not a proper WAV, need to convert
        
        # Convert sang WAV chuẩn
        logger.info(f"🔄 Converting to proper WAV format: {audio_path}")
        
        # Phương án 1: Dùng pydub
        try:
            from pydub import AudioSegment

            # Load audio (pydub tự detect format)
            audio = AudioSegment.from_file(audio_path)
            
            # Convert to proper format for SpeechRecognition
            # - Mono channel
            # - 16kHz sample rate  
            # - 16-bit depth
            audio = audio.set_channels(1)
            audio = audio.set_frame_rate(16000)
            audio = audio.set_sample_width(2)  # 16-bit
            
            # Save as WAV
            wav_path = audio_path.rsplit('.', 1)[0] + '_converted.wav'
            audio.export(wav_path, format="wav")
            
            logger.info(f"✅ Converted to WAV: {wav_path}")
            return wav_path
            
        except ImportError:
            # Phương án 2: Dùng ffmpeg command line
            import subprocess
            
            wav_path = audio_path.rsplit('.', 1)[0] + '_converted.wav'
            
            cmd = [
                'ffmpeg', '-i', audio_path,
                '-ac', '1',           # Mono
                '-ar', '16000',       # 16kHz
                '-acodec', 'pcm_s16le',  # 16-bit PCM
                '-f', 'wav',          # WAV format
                '-y',                 # Overwrite
                wav_path
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode == 0:
                logger.info(f"✅ FFmpeg conversion successful: {wav_path}")
                return wav_path
            else:
                logger.error(f"FFmpeg failed: {result.stderr}")
                raise Exception(f"Audio conversion failed: {result.stderr}")
                
    except Exception as e:
        logger.error(f"Audio conversion error: {e}")
        # Thử phương án fallback
        return await convert_audio_fallback(audio_path)

async def convert_audio_fallback(audio_path: str) -> str:
    """
    Phương án fallback cho audio conversion
    """
    try:
        # Thử đọc raw data và viết lại dưới dạng WAV
        import array
        import wave

        # Đọc toàn bộ file như binary
        with open(audio_path, 'rb') as f:
            raw_data = f.read()
        
        # Tạo WAV mới với giả định đây là 16-bit mono 44.1kHz
        wav_path = audio_path.rsplit('.', 1)[0] + '_raw.wav'
        
        with wave.open(wav_path, 'w') as wav_file:
            wav_file.setnchannels(1)          # Mono
            wav_file.setsampwidth(2)         # 16-bit
            wav_file.setframerate(16000)     # 16kHz
            
            # Giả định data là 16-bit PCM
            # Nếu data quá ngắn, padding với zeros
            if len(raw_data) < 32000:  # 1 second of 16kHz 16-bit mono
                # Create silent audio
                import numpy as np
                silence = np.zeros(16000, dtype=np.int16)
                wav_file.writeframes(silence.tobytes())
            else:
                # Use first N bytes as audio data
                # Trim to multiple of 2 (16-bit samples)
                data_len = min(len(raw_data), 32000)  # Max 2 seconds
                data_len = data_len - (data_len % 2)  # Make even
                wav_file.writeframes(raw_data[:data_len])
        
        logger.info(f"⚠️ Created fallback WAV: {wav_path}")
        return wav_path
        
    except Exception as e:
        logger.error(f"Fallback conversion also failed: {e}")
        raise Exception(f"Cannot convert audio: {str(e)}")

async def transcribe_with_speech_recognition_fixed(audio_path: str):
    """
    SpeechRecognition với xử lý audio format
    """
    try:
        import speech_recognition as sr
        
        logger.info("🔄 Loading SpeechRecognition...")
        
        # Khởi tạo recognizer
        recognizer = sr.Recognizer()
        
        # Đảm bảo file là WAV chuẩn
        logger.info(f"🔧 Ensuring WAV format: {audio_path}")
        wav_path = await ensure_wav_format(audio_path)
        
        if not os.path.exists(wav_path):
            raise Exception(f"Converted WAV file not found: {wav_path}")
        
        # Kiểm tra kích thước file sau conversion
        file_size = os.path.getsize(wav_path)
        logger.info(f"📁 Converted WAV size: {file_size} bytes")
        
        if file_size < 1000:  # File quá nhỏ
            logger.warning("Converted file too small, may be silent")
            raise Exception("Audio file is too small or silent")
        
        # Đọc audio file
        logger.info("🎵 Reading audio file...")
        with sr.AudioFile(wav_path) as source:
            # Điều chỉnh cho nhiễu nền
            logger.info("🎚️ Adjusting for ambient noise...")
            recognizer.adjust_for_ambient_noise(source, duration=0.5)
            
            # Ghi âm
            logger.info("⏺️ Recording audio...")
            audio = recognizer.record(source)
            
            logger.info("🔄 Sending to Google Speech Recognition...")
            
            # Nhận diện với Google Web Speech API
            # Thêm timeout và retry
            try:
                text = recognizer.recognize_google(audio, language="vi-VN")
            except sr.RequestError as e:
                # Thử lại với English nếu Vietnamese fail
                logger.warning(f"Vietnamese recognition failed, trying English: {e}")
                text = recognizer.recognize_google(audio, language="en-US")
            
            # Confidence mặc định
            confidence = 0.7
            
            logger.info(f"✅ Google Speech Recognition success: {len(text)} chars")
            
            # Clean up converted file
            try:
                if wav_path != audio_path and os.path.exists(wav_path):
                    os.remove(wav_path)
            except:
                pass
            
            return text, confidence
            
    except ImportError:
        raise Exception(
            "SpeechRecognition not installed. "
            "Install with: pip install SpeechRecognition pydub"
        )
    except sr.UnknownValueError:
        # Audio không có tiếng nói rõ ràng
        raise Exception("Không thể nhận diện tiếng nói trong audio")
    except sr.RequestError as e:
        raise Exception(f"Lỗi kết nối Google API: {str(e)}")
    except Exception as e:
        logger.error(f"SpeechRecognition error: {e}", exc_info=True)
        raise
    
async def transcribe_audio_simple(audio_path: str):
    """
    STT đơn giản - với xử lý audio format và MULTIPLE RETRY
    """
    try:
        logger.info(f"🎤 Starting simple STT for: {audio_path}")
        
        # Kiểm tra file
        if not os.path.exists(audio_path):
            raise Exception(f"Audio file not found: {audio_path}")
        
        file_size = os.path.getsize(audio_path)
        logger.info(f"📁 Audio file size: {file_size} bytes")
        
        if file_size < 1000:  # File quá nhỏ
            logger.warning("⚠️ Audio file too small, may be silent")
            return "[Audio too short for speech recognition]", 0.0
        
        # =========== PHƯƠNG PHÁP 1: Google Speech Recognition ===========
        try:
            logger.info("🔄 Trying Google Speech Recognition...")
            text, confidence = await transcribe_with_speech_recognition_fixed(audio_path)
            
            # Kiểm tra kết quả
            if text and len(text.strip()) > 3:
                logger.info(f"✅ Google Speech Recognition success: {text[:50]}...")
                return text, confidence
        except Exception as e:
            logger.warning(f"⚠️ Google Speech Recognition failed: {e}")
        
        # =========== PHƯƠNG PHÁP 2: Whisper (nếu có) ===========
        try:
            logger.info("🔄 Trying Whisper...")
            text, confidence = await transcribe_with_whisper_simple(audio_path)
            
            if text and len(text.strip()) > 3:
                logger.info(f"✅ Whisper success: {text[:50]}...")
                return text, confidence
        except Exception as e:
            logger.warning(f"⚠️ Whisper failed: {e}")
        
        # =========== PHƯƠNG PHÁP 3: Đọc raw file ===========
        try:
            # Kiểm tra xem có phải file silent không
            text = await detect_silent_audio(audio_path)
            if text:
                return text, 0.5
        except Exception as e:
            logger.warning(f"⚠️ Silent detection failed: {e}")
        
        # Nếu tất cả đều fail
        logger.warning("❌ All STT methods failed")
        return "[Voice feedback - không thể nhận diện tiếng nói]", 0.0
        
    except Exception as e:
        logger.error(f"❌ Simple STT error: {e}")
        return f"[STT error: {str(e)}]", 0.0

async def transcribe_with_speech_recognition_simple(audio_path: str):
    """
    Sử dụng SpeechRecognition với Google Web API - NHẸ, KHÔNG CẦN TORCH
    """
    try:
        import speech_recognition as sr
        
        logger.info("🔄 Loading SpeechRecognition...")
        
        # Khởi tạo recognizer
        recognizer = sr.Recognizer()
        
        # Kiểm tra file format, convert sang WAV nếu cần
        wav_path = audio_path
        if not audio_path.lower().endswith('.wav'):
            wav_path = await convert_to_wav_simple(audio_path)
        
        # Đọc audio file
        with sr.AudioFile(wav_path) as source:
            # Adjust for ambient noise
            recognizer.adjust_for_ambient_noise(source, duration=0.5)
            audio = recognizer.record(source)
            
            logger.info("🔄 Sending to Google Speech Recognition...")
            
            # Nhận diện với Google Web Speech API (MIỄN PHÍ)
            text = recognizer.recognize_google(audio, language="vi-VN")
            
            # Confidence mặc định
            confidence = 0.7
            
            logger.info(f"✅ Google Speech Recognition success: {len(text)} chars")
            return text, confidence
            
    except ImportError:
        # Hướng dẫn cài đặt
        raise Exception(
            "SpeechRecognition not installed. "
            "Install with: pip install SpeechRecognition pydub"
        )
    except sr.UnknownValueError:
        raise Exception("Google Speech Recognition could not understand the audio")
    except sr.RequestError as e:
        raise Exception(f"Could not request results from Google: {e}")
    except Exception as e:
        logger.error(f"SpeechRecognition error: {e}")
        raise

async def convert_to_wav_simple(input_path: str):
    """
    Convert audio sang WAV format đơn giản
    """
    try:
        # Phương án 1: Dùng pydub
        try:
            from pydub import AudioSegment
            
            audio = AudioSegment.from_file(input_path)
            wav_path = input_path.rsplit('.', 1)[0] + '.wav'
            
            # Convert to mono, 16kHz
            audio = audio.set_channels(1)
            audio = audio.set_frame_rate(16000)
            audio.export(wav_path, format="wav")
            
            return wav_path
            
        except ImportError:
            # Phương án 2: Dùng ffmpeg command line
            import subprocess
            
            wav_path = input_path.rsplit('.', 1)[0] + '.wav'
            
            cmd = [
                'ffmpeg', '-i', input_path,
                '-ac', '1',
                '-ar', '16000',
                '-acodec', 'pcm_s16le',
                '-y',  # Overwrite
                wav_path
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode == 0:
                return wav_path
            else:
                raise Exception(f"FFmpeg failed: {result.stderr}")
                
    except Exception as e:
        logger.error(f"Audio conversion error: {e}")
        # Nếu không convert được, thử dùng file gốc
        return input_path

def get_audio_duration_simple(filepath: str) -> float:
    """Get audio duration đơn giản"""
    try:
        # Dùng pydub
        from pydub import AudioSegment
        audio = AudioSegment.from_file(filepath)
        return len(audio) / 1000.0
    except:
        try:
            # Dùng wave module
            import wave
            with wave.open(filepath, 'rb') as wav_file:
                frames = wav_file.getnframes()
                rate = wav_file.getframerate()
                return frames / float(rate)
        except:
            return 0.0
        
async def transcribe_audio_with_whisper(audio_path: str):
    """
    Chuyển đổi audio thành text bằng Whisper
    """
    try:
        logger.info(f"🎤 Starting Whisper STT for: {audio_path}")
        
        # Kiểm tra file tồn tại
        if not os.path.exists(audio_path):
            raise Exception(f"Audio file not found: {audio_path}")
        
        # Kiểm tra kích thước file
        file_size = os.path.getsize(audio_path)
        if file_size < 100:  # Quá nhỏ
            raise Exception(f"Audio file too small: {file_size} bytes")
        
        logger.info(f"📁 Audio file size: {file_size} bytes")
        
        # Phương án 1: Sử dụng OpenAI Whisper package
        try:
            return await transcribe_with_openai_whisper(audio_path)
        except Exception as e1:
            logger.warning(f"OpenAI Whisper failed: {e1}")
            
            # Phương án 2: Sử dụng Whisper command line
            try:
                return await transcribe_with_whisper_cli(audio_path)
            except Exception as e2:
                logger.warning(f"Whisper CLI failed: {e2}")
                
                # Phương án 3: Sử dụng SpeechRecognition làm fallback
                try:
                    return await transcribe_with_speech_recognition(audio_path)
                except Exception as e3:
                    logger.error(f"All STT methods failed: {e3}")
                    raise Exception(f"All STT methods failed: {e1}, {e2}, {e3}")
                    
    except Exception as e:
        logger.error(f"❌ STT error: {e}")
        raise

async def transcribe_with_openai_whisper(audio_path: str):
    """
    Sử dụng OpenAI Whisper Python package
    """
    try:
        import whisper
        
        logger.info("🔄 Loading Whisper model...")
        
        # Load model nhẹ nhất (tiny) để xử lý nhanh
        # Các model có sẵn: tiny, base, small, medium, large
        model = whisper.load_model("tiny")
        
        logger.info("✅ Whisper model loaded")
        
        # Transcribe audio
        result = model.transcribe(
            audio_path,
            language="vi",  # Ngôn ngữ tiếng Việt
            fp16=False,     # Sử dụng CPU
            verbose=False   # Không hiển thị chi tiết
        )
        
        text = result["text"].strip()
        
        # Lấy confidence từ segments
        confidence = 0.8  # Mặc định
        if result.get("segments"):
            # Tính confidence trung bình từ các segments
            confidences = [seg.get("confidence", 0.5) for seg in result["segments"]]
            if confidences:
                confidence = sum(confidences) / len(confidences)
        
        logger.info(f"✅ Whisper transcription: {len(text)} chars, confidence: {confidence:.2f}")
        
        return text, confidence
        
    except ImportError:
        logger.warning("Whisper package not installed")
        # Hướng dẫn cài đặt
        raise Exception(
            "Whisper not installed. Install with: pip install openai-whisper\n"
            "Also install ffmpeg: sudo apt install ffmpeg (Ubuntu) or brew install ffmpeg (Mac)"
        )
    except Exception as e:
        logger.error(f"Whisper transcription error: {e}")
        raise

async def transcribe_with_whisper_cli(audio_path: str):
    """
    Sử dụng Whisper command line interface
    """
    try:
        import json
        import subprocess
        import tempfile
        
        logger.info("🔄 Trying Whisper CLI...")
        
        # Kiểm tra whisper command có tồn tại không
        try:
            subprocess.run(["whisper", "--help"], capture_output=True, check=True)
        except:
            # Cố gắng cài đặt whisper
            logger.info("Installing Whisper CLI...")
            subprocess.run([sys.executable, "-m", "pip", "install", "openai-whisper"], 
                          capture_output=True)
        
        # Tạo file output tạm thời
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as tmp_file:
            output_file = tmp_file.name
        
        # Chạy whisper command line
        cmd = [
            "whisper",
            audio_path,
            "--language", "vi",
            "--output_format", "json",
            "--output_dir", os.path.dirname(output_file),
            "--model", "tiny",
            "--fp16", "False"
        ]
        
        logger.info(f"Running command: {' '.join(cmd)}")
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode == 0:
            # Đọc kết quả từ file JSON
            json_file = output_file.replace(".json", "") + ".json"
            if os.path.exists(json_file):
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                text = data.get("text", "").strip()
                
                # Tính confidence
                confidence = 0.7
                if "segments" in data and data["segments"]:
                    confidences = [seg.get("confidence", 0.5) for seg in data["segments"]]
                    if confidences:
                        confidence = sum(confidences) / len(confidences)
                
                # Xóa file tạm
                try:
                    os.remove(json_file)
                    os.remove(output_file)
                except:
                    pass
                
                if text:
                    logger.info(f"✅ Whisper CLI success: {len(text)} chars")
                    return text, confidence
                
        logger.error(f"Whisper CLI failed: {result.stderr}")
        raise Exception(f"Whisper CLI error: {result.stderr}")
        
    except Exception as e:
        logger.error(f"Whisper CLI error: {e}")
        raise

async def transcribe_with_speech_recognition(audio_path: str):
    """
    Fallback: Sử dụng SpeechRecognition với Google API
    """
    try:
        import speech_recognition as sr
        
        logger.info("🔄 Trying SpeechRecognition...")
        
        # Khởi tạo recognizer
        recognizer = sr.Recognizer()
        
        # Convert audio sang định dạng phù hợp nếu cần
        wav_path = audio_path
        if not audio_path.lower().endswith('.wav'):
            # Convert sang WAV
            wav_path = audio_path.rsplit('.', 1)[0] + '.wav'
            await convert_audio_to_wav(audio_path, wav_path)
        
        # Đọc audio file
        with sr.AudioFile(wav_path) as source:
            # Điều chỉnh cho nhiễu nền
            recognizer.adjust_for_ambient_noise(source, duration=0.5)
            audio = recognizer.record(source)
            
            # Nhận diện với Google Web Speech API
            text = recognizer.recognize_google(audio, language="vi-VN")
            
            logger.info(f"✅ SpeechRecognition success: {len(text)} chars")
            return text, 0.7  # Confidence mặc định
            
    except ImportError:
        raise Exception("SpeechRecognition not installed. Install with: pip install SpeechRecognition")
    except sr.UnknownValueError:
        raise Exception("Could not understand audio")
    except sr.RequestError as e:
        raise Exception(f"Google API error: {str(e)}")
    except Exception as e:
        logger.error(f"SpeechRecognition error: {e}")
        raise

async def convert_audio_to_wav(input_path: str, output_path: str):
    """
    Convert audio sang định dạng WAV 16kHz mono
    """
    try:
        # Phương án 1: Sử dụng pydub
        try:
            from pydub import AudioSegment
            
            audio = AudioSegment.from_file(input_path)
            audio = audio.set_channels(1)  # Mono
            audio = audio.set_frame_rate(16000)  # 16kHz
            audio.export(output_path, format="wav")
            return output_path
            
        except ImportError:
            # Phương án 2: Sử dụng ffmpeg command line
            import subprocess
            
            cmd = [
                'ffmpeg', '-i', input_path,
                '-ac', '1',  # Mono
                '-ar', '16000',  # 16kHz
                '-acodec', 'pcm_s16le',  # 16-bit PCM
                '-y',  # Overwrite output
                output_path
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode == 0:
                return output_path
            else:
                raise Exception(f"FFmpeg failed: {result.stderr}")
                
    except Exception as e:
        logger.error(f"Audio conversion error: {e}")
        raise Exception(f"Audio conversion failed: {str(e)}")
    
# Tạo thư mục lưu audio files
FEEDBACK_AUDIO_DIR = "feedback_audio"
os.makedirs(FEEDBACK_AUDIO_DIR, exist_ok=True)

def save_audio_file(audio_base64: str, filename: str) -> str:
    """Save base64 audio to file"""
    try:
        # Remove data URL prefix if present
        if ',' in audio_base64:
            audio_base64 = audio_base64.split(',')[1]
        
        audio_data = base64.b64decode(audio_base64)
        filepath = os.path.join(FEEDBACK_AUDIO_DIR, filename)
        
        with open(filepath, 'wb') as f:
            f.write(audio_data)
        
        return filepath
    except Exception as e:
        logger.error(f"Error saving audio file: {e}")
        raise HTTPException(status_code=500, detail=f"Error saving audio: {str(e)}")

def get_audio_duration(filepath: str) -> float:
    """Get audio duration in seconds"""
    try:
        import wave
        with wave.open(filepath, 'r') as audio_file:
            frames = audio_file.getnframes()
            rate = audio_file.getframerate()
            return frames / float(rate)
    except:
        try:
            from pydub import AudioSegment
            audio = AudioSegment.from_file(filepath)
            return len(audio) / 1000.0  # Convert ms to seconds
        except:
            return 0.0

@app.post("/api/feedback/text")
async def create_text_feedback(
    feedback_data: TextFeedbackCreate,
    db: Session = Depends(get_db)
):
    """Nhận feedback dạng text từ học sinh"""
    try:
        feedback = StudentFeedback(
            student_id=feedback_data.student_id,
            student_name=feedback_data.student_name,
            feedback_text=feedback_data.feedback_text,
            feedback_type="text",
            emotion=feedback_data.emotion,
            rating=feedback_data.rating,
            class_name=feedback_data.class_name or "AI Class",
            session_id=feedback_data.session_id or f"FB_{int(time.time())}",
            recorded_by="student",  # Assuming feedback is from student
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.add(feedback)
        db.commit()
        db.refresh(feedback)
        
        logger.info(f"✅ Text feedback saved: {feedback_data.student_name} - {len(feedback_data.feedback_text)} chars")
        
        return {
            "status": "success",
            "message": "Feedback đã được lưu thành công",
            "feedback_id": feedback.id,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Error saving text feedback: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.post("/api/feedback/voice")
async def create_voice_feedback(
    feedback_data: VoiceFeedbackCreate,
    db: Session = Depends(get_db)
):
    """Nhận feedback dạng voice và chuyển đổi sang text"""
    try:
        logger.info(f"🎤 Receiving voice feedback from {feedback_data.student_name}")
        
        # 1. Save audio file
        timestamp = int(time.time())
        filename = f"feedback_{feedback_data.student_id}_{timestamp}.{feedback_data.audio_format}"
        filepath = os.path.join(FEEDBACK_AUDIO_DIR, filename)
        
        # Decode base64
        try:
            if ',' in feedback_data.audio_base64:
                audio_base64 = feedback_data.audio_base64.split(',')[1]
            else:
                audio_base64 = feedback_data.audio_base64
                
            audio_data = base64.b64decode(audio_base64)
            
            with open(filepath, 'wb') as f:
                f.write(audio_data)
                
            logger.info(f"✅ Audio saved: {filepath} ({len(audio_data)} bytes)")
            
        except Exception as e:
            logger.error(f"❌ Error saving audio: {e}")
            raise HTTPException(status_code=400, detail=f"Lỗi lưu audio: {str(e)}")
        
        # 2. Get audio duration
        audio_duration = 0.0
        try:
            audio_duration = get_audio_duration_simple(filepath)
            logger.info(f"⏱️ Audio duration: {audio_duration:.2f} seconds")
        except Exception as e:
            logger.warning(f"⚠️ Could not get audio duration: {e}")
        
        # 3. Transcribe audio to text (SIMPLE VERSION)
        transcribed_text = ""
        confidence = 0.0
        stt_method = "none"
        
        try:
            logger.info("🔄 Starting STT processing...")
            transcribed_text, confidence = await transcribe_audio_simple(filepath)
            stt_method = "speech_recognition"
            logger.info(f"✅ STT successful: {len(transcribed_text)} characters")
        except Exception as stt_error:
            logger.error(f"❌ STT failed: {stt_error}")
            transcribed_text = f"[Voice feedback - STT failed: {str(stt_error)}]"
            confidence = 0.0
            stt_method = "failed"
            
            # Cung cấp hướng dẫn cài đặt
            install_guide = """
            Để sử dụng Speech-to-Text, cần cài đặt:
            1. pip install SpeechRecognition
            2. pip install pydub
            3. Cài đặt ffmpeg trên hệ thống:
               - Windows: Download từ https://ffmpeg.org/download.html
               - Ubuntu: sudo apt install ffmpeg
               - Mac: brew install ffmpeg
            """
            logger.info(f"📋 Installation guide: {install_guide}")
        
        # 4. Save to database
        feedback = StudentFeedback(
            student_id=feedback_data.student_id,
            student_name=feedback_data.student_name,
            feedback_text=transcribed_text,
            feedback_type="voice",
            audio_path=filepath,
            audio_duration=audio_duration,
            transcribed_text=transcribed_text,
            confidence=confidence,
            class_name=feedback_data.class_name or "AI Class",
            session_id=feedback_data.session_id or f"FB_VOICE_{timestamp}",
            recorded_by="student",
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.add(feedback)
        db.commit()
        db.refresh(feedback)
        
        logger.info(f"✅ Voice feedback saved: ID={feedback.id}, method={stt_method}, "
                   f"duration={audio_duration:.1f}s, confidence={confidence:.2f}")
        
        # 5. Return response
        response_data = {
            "status": "success" if stt_method != "failed" else "partial_success",
            "message": "Voice feedback đã được lưu và chuyển đổi thành công" if stt_method != "failed" 
                      else "Voice feedback đã được lưu nhưng STT thất bại",
            "feedback_id": feedback.id,
            "transcribed_text": transcribed_text,
            "confidence": confidence,
            "audio_duration": audio_duration,
            "stt_method": stt_method,
            "audio_file": filename,
            "timestamp": datetime.now().isoformat()
        }
        
        if stt_method == "failed":
            response_data["warning"] = "Cần cài đặt SpeechRecognition"
            response_data["install_commands"] = [
                "pip install SpeechRecognition",
                "pip install pydub",
                "# Cài ffmpeg trên hệ thống"
            ]
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"❌ Error in create_voice_feedback: {e}", exc_info=True)
        raise HTTPException(
            status_code=500, 
            detail=f"Lỗi server: {str(e)}"
        )

@app.post("/api/feedback/process-stt/{feedback_id}")
async def process_stt_for_feedback(
    feedback_id: int,
    db: Session = Depends(get_db)
):
    """Xử lý STT cho feedback đã lưu"""
    try:
        feedback = db.query(StudentFeedback).filter(StudentFeedback.id == feedback_id).first()
        if not feedback:
            raise HTTPException(status_code=404, detail="Không tìm thấy feedback")
        
        if feedback.feedback_type != "voice":
            raise HTTPException(status_code=400, detail="Không phải voice feedback")
        
        if not feedback.audio_path or not os.path.exists(feedback.audio_path):
            raise HTTPException(status_code=400, detail="Không tìm thấy file audio")
        
        logger.info(f"🔄 Processing STT for feedback {feedback_id}: {feedback.audio_path}")
        
        # Transcribe audio
        transcribed_text = ""
        confidence = 0.0
        
        try:
            transcribed_text, confidence = await transcribe_audio_with_whisper(feedback.audio_path)
            stt_method = "whisper"
            logger.info(f"✅ STT successful for feedback {feedback_id}: {len(transcribed_text)} chars")
        except Exception as e:
            logger.error(f"❌ STT failed for feedback {feedback_id}: {e}")
            transcribed_text = f"[STT failed: {str(e)}]"
            confidence = 0.0
            stt_method = "failed"
        
        # Update feedback
        feedback.transcribed_text = transcribed_text
        feedback.feedback_text = transcribed_text
        feedback.confidence = confidence
        feedback.updated_at = datetime.now()
        
        db.commit()
        
        return {
            "status": "success" if stt_method != "failed" else "partial_success",
            "message": "STT xử lý thành công" if stt_method != "failed" else "STT thất bại",
            "feedback_id": feedback.id,
            "transcribed_text": transcribed_text,
            "confidence": confidence,
            "stt_method": stt_method,
            "audio_file": os.path.basename(feedback.audio_path)
        }
        
    except Exception as e:
        logger.error(f"Error processing STT: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý STT: {str(e)}")

@app.get("/api/feedback/unprocessed")
async def get_unprocessed_feedback(
    db: Session = Depends(get_db)
):
    """Lấy danh sách feedback chưa được xử lý STT"""
    try:
        # Tìm các voice feedback chưa có transcribed_text hoặc confidence thấp
        unprocessed = db.query(StudentFeedback).filter(
            StudentFeedback.feedback_type == "voice",
            or_(
                StudentFeedback.transcribed_text.is_(None),
                StudentFeedback.transcribed_text == "",
                StudentFeedback.transcribed_text.like("[Voice feedback%"),
                StudentFeedback.confidence < 0.3
            )
        ).order_by(StudentFeedback.created_at.desc()).limit(20).all()
        
        return {
            "status": "success",
            "count": len(unprocessed),
            "feedbacks": [
                {
                    "id": fb.id,
                    "student_name": fb.student_name,
                    "created_at": fb.created_at.isoformat() if fb.created_at else None,
                    "audio_file": os.path.basename(fb.audio_path) if fb.audio_path else None,
                    "has_audio": os.path.exists(fb.audio_path) if fb.audio_path else False
                }
                for fb in unprocessed
            ]
        }
        
    except Exception as e:
        logger.error(f"Error getting unprocessed feedback: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")
    
@app.post("/api/feedback/process-voice/{feedback_id}")
async def process_voice_feedback(
    feedback_id: int,
    db: Session = Depends(get_db)
):
    """Xử lý STT cho voice feedback đã lưu"""
    try:
        feedback = db.query(StudentFeedback).filter(StudentFeedback.id == feedback_id).first()
        if not feedback:
            raise HTTPException(status_code=404, detail="Không tìm thấy feedback")
        
        if feedback.feedback_type != "voice" or not feedback.audio_path:
            raise HTTPException(status_code=400, detail="Không phải voice feedback hoặc không có audio")
        
        logger.info(f"🔄 Processing STT for feedback {feedback_id}: {feedback.audio_path}")
        
        # Try multiple STT methods
        transcribed_text = ""
        confidence = 0.0
        audio_duration = 0.0
        
        # Get audio duration
        try:
            audio_duration = get_audio_duration(feedback.audio_path)
            feedback.audio_duration = audio_duration
        except:
            pass
        
        # Try STT methods
        methods = [
            ("whisper_simple", transcribe_with_whisper_simple),
            ("speech_recognition", transcribe_with_speech_recognition),
        ]
        
        for method_name, method_func in methods:
            try:
                logger.info(f"  Trying {method_name}...")
                transcribed_text, confidence = await method_func(feedback.audio_path)
                if transcribed_text and len(transcribed_text) > 3:
                    logger.info(f"  ✅ {method_name} success: {len(transcribed_text)} chars")
                    break
            except Exception as e:
                logger.warning(f"  ⚠️ {method_name} failed: {e}")
                continue
        
        # Update feedback record
        if transcribed_text:
            feedback.transcribed_text = transcribed_text
            feedback.feedback_text = f"[Voice] {transcribed_text}"
            feedback.confidence = confidence
        else:
            feedback.feedback_text = "[Voice feedback - không thể chuyển đổi thành text]"
            feedback.transcribed_text = ""
        
        feedback.updated_at = datetime.now()
        db.commit()
        
        return {
            "status": "success",
            "message": "Xử lý STT thành công",
            "feedback_id": feedback.id,
            "transcribed_text": transcribed_text,
            "confidence": confidence,
            "audio_duration": audio_duration,
            "method_used": method_name if transcribed_text else "none"
        }
        
    except Exception as e:
        logger.error(f"Error processing voice feedback: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý: {str(e)}")

# ==================== SIMPLE STT METHODS ====================

async def transcribe_with_whisper_simple(audio_path: str):
    """Đơn giản nhất: sử dụng whisper command line nếu có"""
    try:
        # Kiểm tra xem whisper command có tồn tại không
        import subprocess

        # Try to use whisper CLI (pip install openai-whisper)
        result = subprocess.run(
            ["whisper", audio_path, "--language", "vi", "--output_format", "txt"],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0:
            # Read the output file
            txt_file = audio_path.rsplit('.', 1)[0] + '.txt'
            if os.path.exists(txt_file):
                with open(txt_file, 'r', encoding='utf-8') as f:
                    text = f.read().strip()
                return text, 0.8
        
        raise Exception("Whisper CLI failed")
        
    except Exception as e:
        logger.warning(f"Whisper simple failed: {e}")
        raise

async def transcribe_with_speech_recognition(audio_path: str):
    """Sử dụng SpeechRecognition với Google Web API"""
    try:
        import speech_recognition as sr

        # Convert audio to WAV format if needed
        wav_path = audio_path
        if not audio_path.endswith('.wav'):
            import subprocess
            wav_path = audio_path.rsplit('.', 1)[0] + '.wav'
            subprocess.run([
                "ffmpeg", "-i", audio_path, "-ar", "16000", "-ac", "1", 
                "-c:a", "pcm_s16le", wav_path
            ], capture_output=True)
        
        recognizer = sr.Recognizer()
        with sr.AudioFile(wav_path) as source:
            audio = recognizer.record(source)
            text = recognizer.recognize_google(audio, language="vi-VN")
            return text, 0.7
            
    except ImportError:
        logger.warning("SpeechRecognition not installed")
        raise Exception("SpeechRecognition not available")
    except Exception as e:
        logger.warning(f"SpeechRecognition failed: {e}")
        raise

# ==================== WHISPER STT IMPLEMENTATION ====================
import asyncio
import subprocess
import tempfile


async def transcribe_with_whisper(audio_path: str):
    """
    Transcribe audio using Whisper (local) - HOÀN TOÀN MIỄN PHÍ
    Có 2 phương án: Whisper.cpp (nhanh) hoặc openai-whisper (chính thức)
    """
    try:
        # Phương án 1: Whisper.cpp (nhanh, nhẹ, không cần Python package)
        return await transcribe_with_whisper_cpp(audio_path)
    except Exception as e1:
        logger.warning(f"Whisper.cpp failed: {e1}")
        try:
            # Phương án 2: openai-whisper Python package
            return await transcribe_with_openai_whisper(audio_path)
        except Exception as e2:
            logger.warning(f"OpenAI Whisper failed: {e2}")
            # Phương án 3: Google Speech Recognition fallback (miễn phí, cần internet)
            return await transcribe_with_google_speech(audio_path)

async def transcribe_with_whisper_cpp(audio_path: str):
    """
    Sử dụng Whisper.cpp - cực kỳ nhanh và nhẹ
    Cần download whisper.cpp từ: https://github.com/ggerganov/whisper.cpp
    """
    try:
        # Kiểm tra whisper.cpp có tồn tại không
        whisper_cpp_path = "./whisper.cpp/main"  # Hoặc đường dẫn đầy đủ
        model_path = "./whisper.cpp/models/ggml-base.bin"  # Base model
        
        if not os.path.exists(whisper_cpp_path):
            # Nếu không có whisper.cpp, dùng Python package
            raise Exception("Whisper.cpp not found, using Python package")
        
        # Convert audio to WAV 16kHz nếu cần
        temp_wav = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
        temp_wav.close()
        
        # Convert audio to proper format
        import subprocess
        subprocess.run([
            "ffmpeg", "-i", audio_path, "-ar", "16000", "-ac", "1", "-c:a", "pcm_s16le", temp_wav.name
        ], capture_output=True)
        
        # Run whisper.cpp
        result = subprocess.run([
            whisper_cpp_path,
            "-m", model_path,
            "-f", temp_wav.name,
            "-oj",  # Output JSON
            "-nt",  # No translation
            "-l", "vi"  # Language Vietnamese
        ], capture_output=True, text=True)
        
        # Clean up temp file
        os.unlink(temp_wav.name)
        
        if result.returncode == 0:
            # Parse JSON output
            output_json = json.loads(result.stdout)
            if output_json and "transcription" in output_json:
                transcription = output_json["transcription"]
                # Tính confidence đơn giản dựa trên độ dài
                confidence = min(0.9, len(transcription) / 100.0)
                return transcription, confidence
        
        raise Exception("Whisper.cpp transcription failed")
        
    except Exception as e:
        logger.error(f"Whisper.cpp error: {e}")
        raise

async def transcribe_with_openai_whisper(audio_path: str):
    """
    Sử dụng openai-whisper Python package
    Cài đặt: pip install openai-whisper
    """
    try:
        import whisper

        # Load model (tiny, base, small, medium, large)
        # Tiny model là nhẹ nhất, phù hợp cho real-time
        model = whisper.load_model("tiny")
        
        # Transcribe
        result = model.transcribe(
            audio_path,
            language="vi",  # Vietnamese
            fp16=False,  # CPU mode
            verbose=False
        )
        
        text = result["text"].strip()
        # Tính confidence đơn giản (có thể lấy từ result["segments"])
        confidence = 0.8  # Mặc định
        
        return text, confidence
        
    except ImportError:
        logger.warning("openai-whisper not installed. Installing...")
        # Có thể tự động cài đặt
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openai-whisper"])
        # Thử lại
        return await transcribe_with_openai_whisper(audio_path)
    except Exception as e:
        logger.error(f"OpenAI Whisper error: {e}")
        raise

async def transcribe_with_google_speech(audio_path: str):
    """
    Fallback: Google Speech Recognition (miễn phí, cần internet)
    Cài đặt: pip install SpeechRecognition
    """
    try:
        import speech_recognition as sr
        
        recognizer = sr.Recognizer()
        
        # Load audio file
        with sr.AudioFile(audio_path) as source:
            audio = recognizer.record(source)
            
            # Try Google Web Speech API (miễn phí)
            text = recognizer.recognize_google(audio, language="vi-VN")
            confidence = 0.7  # Google không trả về confidence
            
            return text, confidence
            
    except ImportError:
        logger.warning("SpeechRecognition not installed")
        raise Exception("SpeechRecognition package not available")
    except sr.UnknownValueError:
        raise Exception("Google Speech Recognition could not understand audio")
    except sr.RequestError as e:
        raise Exception(f"Could not request results from Google: {e}")
    except Exception as e:
        logger.error(f"Google Speech error: {e}")
        raise

@app.get("/api/feedback")
async def get_feedback_list(
    student_id: Optional[str] = None,
    feedback_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy danh sách feedback"""
    try:
        query = db.query(StudentFeedback)
        
        if student_id:
            query = query.filter(StudentData.student_id == student_id)
        
        if feedback_type:
            query = query.filter(StudentFeedback.feedback_type == feedback_type)
        
        if start_date:
            query = query.filter(StudentFeedback.created_at >= start_date)
        
        if end_date:
            query = query.filter(StudentFeedback.created_at <= end_date)
        
        # Total count
        total_count = query.count()
        
        # Pagination
        offset = (page - 1) * limit
        query = query.order_by(StudentFeedback.created_at.desc())
        feedbacks = query.offset(offset).limit(limit).all()
        
        return {
            "status": "success",
            "count": len(feedbacks),
            "total": total_count,
            "page": page,
            "limit": limit,
            "feedbacks": [
                {
                    "id": fb.id,
                    "student_id": fb.student_id,
                    "student_name": fb.student_name,
                    "feedback_text": fb.feedback_text,
                    "transcribed_text": fb.transcribed_text,
                    "feedback_type": fb.feedback_type,
                    "audio_duration": fb.audio_duration,
                    "confidence": fb.confidence,
                    "emotion": fb.emotion,
                    "rating": fb.rating,
                    "class_name": fb.class_name,
                    "session_id": fb.session_id,
                    "created_at": fb.created_at.isoformat() if fb.created_at else None
                }
                for fb in feedbacks
            ]
        }
        
    except Exception as e:
        logger.error(f"Error getting feedback list: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.get("/api/feedback/stats")
async def get_feedback_stats(
    days: int = Query(7, ge=1, le=30),
    db: Session = Depends(get_db)
):
    """Thống kê feedback"""
    try:
        start_date = datetime.now() - timedelta(days=days)
        
        # Total feedbacks
        total = db.query(StudentFeedback).filter(
            StudentFeedback.created_at >= start_date
        ).count()
        
        # By type
        text_count = db.query(StudentFeedback).filter(
            StudentFeedback.created_at >= start_date,
            StudentFeedback.feedback_type == "text"
        ).count()
        
        voice_count = db.query(StudentFeedback).filter(
            StudentFeedback.created_at >= start_date,
            StudentFeedback.feedback_type == "voice"
        ).count()
        
        # By student
        top_students = db.query(
            StudentFeedback.student_name,
            func.count(StudentFeedback.id).label('count')
        ).filter(
            StudentFeedback.created_at >= start_date
        ).group_by(
            StudentFeedback.student_name
        ).order_by(func.count(StudentFeedback.id).desc()).limit(5).all()
        
        # Recent feedbacks
        recent = db.query(StudentFeedback).filter(
            StudentFeedback.created_at >= start_date
        ).order_by(StudentFeedback.created_at.desc()).limit(5).all()
        
        return {
            "status": "success",
            "period_days": days,
            "stats": {
                "total_feedbacks": total,
                "text_feedbacks": text_count,
                "voice_feedbacks": voice_count,
                "feedback_rate": round(total / days, 1) if days > 0 else 0
            },
            "top_students": [
                {"student_name": name, "count": count}
                for name, count in top_students
            ],
            "recent_feedbacks": [
                {
                    "student_name": fb.student_name,
                    "type": fb.feedback_type,
                    "preview": (fb.feedback_text or fb.transcribed_text or "")[:50] + "...",
                    "created_at": fb.created_at.isoformat() if fb.created_at else None
                }
                for fb in recent
            ]
        }
        
    except Exception as e:
        logger.error(f"Error getting feedback stats: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

# ==================== STUDENT MANAGEMENT ENDPOINTS ====================
@app.get("/api/students/list")
async def get_students_list(
    class_name: Optional[str] = None,
    is_active: Optional[bool] = True,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy danh sách học sinh từ bảng students"""
    try:
        query = db.query(Student)
        
        if class_name:
            query = query.filter(Student.class_name.ilike(f"%{class_name}%"))
        
        if is_active is not None:
            query = query.filter(Student.is_active == is_active)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        query = query.order_by(Student.student_name.asc())
        students = query.offset(offset).limit(limit).all()
        
        return {
            "status": "success",
            "count": len(students),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "students": [
                {
                    "id": student.id,
                    "student_id": student.student_id,
                    "student_name": student.student_name,
                    "class_name": student.class_name,
                    "is_active": student.is_active,
                    "created_at": student.created_at.isoformat() if student.created_at else None,
                    "updated_at": student.updated_at.isoformat() if student.updated_at else None
                }
                for student in students
            ]
        }
        
    except Exception as e:
        logger.error(f"Error in get_students_list: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.post("/api/students")
async def create_student(
    student_data: StudentCreate,
    db: Session = Depends(get_db)
):
    """Tạo học sinh mới"""
    try:
        # Kiểm tra student_id đã tồn tại chưa
        existing_student = db.query(Student).filter(
            Student.student_id == student_data.student_id
        ).first()
        
        if existing_student:
            raise HTTPException(status_code=400, detail="Mã học sinh đã tồn tại")
        
        # Tạo học sinh mới
        student = Student(
            student_id=student_data.student_id,
            student_name=student_data.student_name,
            class_name=student_data.class_name,
            is_active=student_data.is_active,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.add(student)
        db.commit()
        db.refresh(student)
        
        return {
            "status": "success",
            "message": "Tạo học sinh thành công",
            "student": {
                "id": student.id,
                "student_id": student.student_id,
                "student_name": student.student_name,
                "class_name": student.class_name,
                "is_active": student.is_active
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create student error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.put("/api/students/{student_id}")
async def update_student(
    student_id: str,
    student_data: StudentUpdate,
    db: Session = Depends(get_db)
):
    """Cập nhật thông tin học sinh"""
    try:
        student = db.query(Student).filter(Student.student_id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Không tìm thấy học sinh")
        
        # Cập nhật các trường có giá trị
        update_data = student_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(student, field, value)
        
        student.updated_at = datetime.now()
        db.commit()
        db.refresh(student)
        
        return {
            "status": "success",
            "message": "Cập nhật học sinh thành công",
            "student": {
                "id": student.id,
                "student_id": student.student_id,
                "student_name": student.student_name,
                "class_name": student.class_name,
                "is_active": student.is_active
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update student error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.get("/api/attendance/summary")
async def get_attendance_summary(
    date: Optional[str] = None,
    class_name: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Lấy thống kê điểm danh:
    - Tổng số học sinh
    - Số học sinh có mặt
    - Số học sinh vắng
    - Danh sách học sinh vắng
    """
    try:
        today = date or datetime.now().strftime("%Y-%m-%d")
        
        print(f"📊 Đang tính thống kê điểm danh ngày: {today}")
        
        # Lấy danh sách tất cả học sinh từ StudentData (cách mới)
        print("🔄 Lấy danh sách học sinh từ StudentData...")
        
        # Lấy danh sách học sinh duy nhất từ StudentData
        student_query = db.query(
            StudentData.student_id,
            StudentData.student_name,
            StudentData.class_name,
            func.max(StudentData.recorded_at).label('last_recorded')
        ).filter(
            StudentData.student_id.isnot(None),
            StudentData.student_name.isnot(None)
        ).group_by(
            StudentData.student_id,
            StudentData.student_name,
            StudentData.class_name
        )
        
        if class_name:
            student_query = student_query.filter(StudentData.class_name.ilike(f"%{class_name}%"))
        
        all_students = student_query.all()
        
        if not all_students:
            return {
                "status": "success",
                "date": today,
                "message": "Không có học sinh nào trong hệ thống",
                "summary": {
                    "total_students": 0,
                    "present_count": 0,
                    "absent_count": 0,
                    "attendance_rate": 0
                },
                "absent_students": []
            }
        
        total_students = len(all_students)
        print(f"📊 Tổng số học sinh từ StudentData: {total_students}")
        
        # Lấy danh sách học sinh đã điểm danh hôm nay
        attendance_records = db.query(StudentData).filter(
            StudentData.date == today,
            StudentData.attendance_status.isnot(None)
        ).all()
        
        # Tạo dict để tra cứu nhanh
        student_attendance = {}
        for record in attendance_records:
            if record.student_id:
                student_attendance[record.student_id] = {
                    "status": record.attendance_status,
                    "check_in_time": record.check_in_time,
                    "recorded_at": record.recorded_at
                }
        
        # Phân loại học sinh
        present_students = []
        absent_students = []
        
        for student in all_students:
            attendance_info = student_attendance.get(student.student_id)
            
            if attendance_info and attendance_info["status"] in ["present", "late"]:
                # Đã điểm danh hôm nay
                present_students.append({
                    "student_id": student.student_id,
                    "student_name": student.student_name,
                    "class_name": student.class_name or "Không xác định",
                    "check_in_time": attendance_info["check_in_time"].isoformat() if attendance_info["check_in_time"] else None,
                    "status": attendance_info["status"]
                })
            else:
                # Chưa điểm danh hôm nay
                absent_students.append({
                    "student_id": student.student_id,
                    "student_name": student.student_name,
                    "class_name": student.class_name or "Không xác định",
                    "check_in_time": None,
                    "status": "absent",
                    "reason": "Chưa điểm danh"
                })
        
        present_count = len(present_students)
        absent_count = len(absent_students)
        
        print(f"📊 Có mặt: {present_count}, Vắng: {absent_count}")
        
        # Tính tỷ lệ điểm danh
        attendance_rate = round((present_count / total_students) * 100, 1) if total_students > 0 else 0
        
        # Lấy thời gian cập nhật gần nhất
        latest_update = db.query(func.max(StudentData.recorded_at)).filter(
            StudentData.date == today
        ).scalar()
        
        return {
            "status": "success",
            "date": today,
            "last_updated": latest_update.isoformat() if latest_update else None,
            "summary": {
                "total_students": total_students,
                "present_count": present_count,
                "absent_count": absent_count,
                "attendance_rate": attendance_rate,
                "class_name": class_name or "Tất cả các lớp"
            },
            "present_students": present_students,  # Thêm danh sách học sinh có mặt
            "absent_students": absent_students,
            "data_source": "student_data_table"
        }
        
    except Exception as e:
        logger.error(f"❌ Error in get_attendance_summary: {e}")
        import traceback
        traceback.print_exc()
        
        # Fallback data từ demo
        today = date or datetime.now().strftime("%Y-%m-%d")
        return {
            "status": "success",
            "date": today,
            "message": "Using fallback data",
            "summary": {
                "total_students": 10,
                "present_count": 7,
                "absent_count": 3,
                "attendance_rate": 70.0,
                "class_name": class_name or "Lớp 10A1"
            },
            "present_students": [
                {"student_id": "SV001", "student_name": "Nguyễn Văn A", "class_name": "Lớp 10A1", "check_in_time": "07:30", "status": "present"},
                {"student_id": "SV002", "student_name": "Trần Thị B", "class_name": "Lớp 10A1", "check_in_time": "07:35", "status": "present"},
                {"student_id": "SV004", "student_name": "Phạm Thị D", "class_name": "Lớp 10A1", "check_in_time": "07:40", "status": "present"},
                {"student_id": "SV005", "student_name": "Hoàng Văn E", "class_name": "Lớp 10A1", "check_in_time": "07:42", "status": "present"},
                {"student_id": "SV007", "student_name": "Nguyễn Văn G", "class_name": "Lớp 10A1", "check_in_time": "07:55", "status": "present"},
                {"student_id": "SV008", "student_name": "Trần Thị H", "class_name": "Lớp 10A1", "check_in_time": "08:00", "status": "late"},
                {"student_id": "SV010", "student_name": "Lê Thị K", "class_name": "Lớp 10A1", "check_in_time": "07:38", "status": "present"}
            ],
            "absent_students": [
                {"student_id": "SV003", "student_name": "Lê Văn C", "class_name": "Lớp 10A1", "check_in_time": None, "status": "absent", "reason": "Chưa điểm danh"},
                {"student_id": "SV006", "student_name": "Vũ Thị F", "class_name": "Lớp 10A1", "check_in_time": None, "status": "absent", "reason": "Chưa điểm danh"},
                {"student_id": "SV009", "student_name": "Đỗ Văn I", "class_name": "Lớp 10A1", "check_in_time": None, "status": "absent", "reason": "Chưa điểm danh"}
            ],
            "data_source": "fallback"
        }

@app.get("/api/attendance/daily")
async def get_daily_attendance(
    date: Optional[str] = None,
    class_name: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Lấy chi tiết điểm danh hàng ngày:
    - Danh sách tất cả học sinh với trạng thái điểm danh
    - Thời gian điểm danh
    - Cảm xúc và độ tập trung
    """
    try:
        today = date or datetime.now().strftime("%Y-%m-%d")
        
        # Lấy danh sách học sinh từ bảng students
        student_query = db.query(Student).filter(Student.is_active == True)
        
        if class_name:
            student_query = student_query.filter(Student.class_name.ilike(f"%{class_name}%"))
        
        all_students = student_query.all()
        
        # Lấy dữ liệu điểm danh hôm nay
        attendance_query = db.query(StudentData).filter(
            StudentData.date == today,
            StudentData.attendance_status.isnot(None)
        )
        
        if class_name:
            attendance_query = attendance_query.filter(
                StudentData.class_name.ilike(f"%{class_name}%")
            )
        
        attendance_records = attendance_query.order_by(StudentData.recorded_at.desc()).all()
        
        # Tạo dict để tra cứu nhanh
        attendance_dict = {}
        for record in attendance_records:
            if record.student_id:
                attendance_dict[record.student_id] = {
                    "status": record.attendance_status,
                    "check_in_time": record.check_in_time.isoformat() if record.check_in_time else None,
                    "check_out_time": record.check_out_time.isoformat() if record.check_out_time else None,
                    "emotion": record.emotion,
                    "emotion_confidence": record.emotion_confidence,
                    "focus_score": record.focus_score,
                    "concentration_level": record.concentration_level,
                    "recorded_at": record.recorded_at.isoformat() if record.recorded_at else None
                }
        
        # Tạo danh sách chi tiết
        attendance_details = []
        for student in all_students:
            attendance_info = attendance_dict.get(student.student_id)
            
            if attendance_info:
                # Đã điểm danh
                status_info = {
                    "status": "present",
                    "attendance_status": attendance_info["status"],
                    "check_in_time": attendance_info["check_in_time"],
                    "emotion": attendance_info["emotion"],
                    "focus_score": attendance_info["focus_score"]
                }
            else:
                # Chưa điểm danh
                status_info = {
                    "status": "absent",
                    "attendance_status": "absent",
                    "check_in_time": None,
                    "emotion": None,
                    "focus_score": None
                }
            
            attendance_details.append({
                "student_id": student.student_id,
                "student_name": student.student_name,
                "class_name": student.class_name,
                **status_info
            })
        
        # Sắp xếp: có mặt trước, vắng sau
        attendance_details.sort(key=lambda x: (x["status"] == "absent", x["student_name"]))
        
        # Thống kê
        present_count = len([s for s in attendance_details if s["status"] == "present"])
        absent_count = len([s for s in attendance_details if s["status"] == "absent"])
        total_count = len(attendance_details)
        attendance_rate = round((present_count / total_count) * 100, 1) if total_count > 0 else 0
        
        # Lấy học sinh vắng
        absent_students = [s for s in attendance_details if s["status"] == "absent"]
        
        return {
            "status": "success",
            "date": today,
            "summary": {
                "total_students": total_count,
                "present_count": present_count,
                "absent_count": absent_count,
                "attendance_rate": attendance_rate,
                "class_name": class_name or "Tất cả các lớp"
            },
            "absent_students": [
                {
                    "student_id": s["student_id"],
                    "student_name": s["student_name"],
                    "class_name": s["class_name"]
                }
                for s in absent_students
            ],
            "attendance_details": attendance_details
        }
        
    except Exception as e:
        logger.error(f"Error in get_daily_attendance: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")
    
# ==================== CLASS STUDENTS ENDPOINTS ====================

@app.get("/api/class/students")
async def get_class_students(
    class_name: Optional[str] = None,
    is_active: Optional[bool] = True,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy danh sách học sinh cố định của lớp học"""
    try:
        query = db.query(ClassStudent)
        
        if class_name:
            query = query.filter(ClassStudent.class_name.ilike(f"%{class_name}%"))
        
        if is_active is not None:
            query = query.filter(ClassStudent.is_active == is_active)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        query = query.order_by(ClassStudent.student_name.asc())
        students = query.offset(offset).limit(limit).all()
        
        return {
            "status": "success",
            "count": len(students),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "students": [
                {
                    "id": student.id,
                    "student_id": student.student_id,
                    "student_name": student.student_name,
                    "student_code": student.student_code,
                    "class_name": student.class_name,
                    "gender": student.gender,
                    "date_of_birth": student.date_of_birth,
                    "address": student.address,
                    "phone": student.phone,
                    "email": student.email,
                    "parent_name": student.parent_name,
                    "parent_phone": student.parent_phone,
                    "is_active": student.is_active,
                    "enrollment_date": student.enrollment_date,
                    "created_at": student.created_at.isoformat() if student.created_at else None,
                    "updated_at": student.updated_at.isoformat() if student.updated_at else None
                }
                for student in students
            ]
        }
        
    except Exception as e:
        logger.error(f"Error in get_class_students: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.post("/api/class/students")
async def create_class_student(
    student_data: ClassStudentCreate,
    db: Session = Depends(get_db)
):
    """Thêm học sinh vào danh sách lớp cố định"""
    try:
        # Kiểm tra student_id đã tồn tại chưa
        existing_student = db.query(ClassStudent).filter(
            ClassStudent.student_id == student_data.student_id
        ).first()
        
        if existing_student:
            raise HTTPException(status_code=400, detail="Mã học sinh đã tồn tại trong lớp")
        
        # Tạo học sinh mới
        student = ClassStudent(
            student_id=student_data.student_id,
            student_name=student_data.student_name,
            class_name=student_data.class_name,
            student_code=student_data.student_code,
            gender=student_data.gender,
            date_of_birth=student_data.date_of_birth,
            address=student_data.address,
            phone=student_data.phone,
            email=student_data.email,
            parent_name=student_data.parent_name,
            parent_phone=student_data.parent_phone,
            is_active=student_data.is_active,
            enrollment_date=student_data.enrollment_date,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.add(student)
        db.commit()
        db.refresh(student)
        
        return {
            "status": "success",
            "message": "Thêm học sinh vào lớp thành công",
            "student": {
                "id": student.id,
                "student_id": student.student_id,
                "student_name": student.student_name,
                "class_name": student.class_name
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create class student error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")
        
# ==================== AUTHENTICATION ENDPOINTS ====================
@app.post("/api/auth/login", response_model=Dict[str, Any])
async def login(user_data: UserLogin, db: Session = Depends(get_db)):
    """Đăng nhập"""
    try:
        user = authenticate_user(db, user_data.username, user_data.password)
        if not user:
            raise HTTPException(status_code=401, detail="Tên đăng nhập hoặc mật khẩu không đúng")
        
        if not user.is_active:
            raise HTTPException(status_code=403, detail="Tài khoản đã bị khóa")
        
        token_data = create_access_token(user)
        
        return {
            "status": "success",
            "message": "Đăng nhập thành công",
            **token_data
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.get("/api/auth/check")
async def check_auth(
    token: str = Query(..., description="Access token"),
    db: Session = Depends(get_db)
):
    """Kiểm tra token và trả về thông tin quyền hạn"""
    try:
        user = validate_token_and_get_user(token, db)
        
        return {
            "status": "success",
            "authenticated": True,
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "full_name": user.full_name,
                "is_active": user.is_active,
                "is_admin": user.is_admin,
                "created_at": user.created_at.isoformat() if user.created_at else None
            },
            "permissions": {
                "can_view_admin_dashboard": user.is_admin,
                "can_manage_students": user.is_admin,
                "can_view_reports": user.is_admin,
                "can_view_basic_info": True  # Ai cũng xem được thông tin cơ bản
            }
        }
        
    except HTTPException as e:
        return {
            "status": "error",
            "authenticated": False,
            "message": e.detail
        }
    except Exception as e:
        logger.error(f"Check auth error: {e}")
        return {
            "status": "error",
            "authenticated": False,
            "message": "Lỗi server"
        }

@app.post("/api/auth/register")
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    """Đăng ký tài khoản"""
    try:
        # Kiểm tra username đã tồn tại
        existing_user = get_user_by_username(db, user_data.username)
        if existing_user:
            raise HTTPException(status_code=400, detail="Username đã tồn tại")
        
        # Kiểm tra email đã tồn tại
        existing_email = get_user_by_email(db, user_data.email)
        if existing_email:
            raise HTTPException(status_code=400, detail="Email đã tồn tại")
        
        # Tạo user mới
        hashed_password = hash_password(user_data.password)
        user = User(
            username=user_data.username,
            email=user_data.email,
            hashed_password=hashed_password,
            full_name=user_data.full_name,
            is_active=True,
            is_admin=user_data.is_admin,
            created_at=datetime.now()
        )
        
        db.add(user)
        db.commit()
        db.refresh(user)
        
        return {
            "status": "success",
            "message": "Đăng ký thành công",
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "full_name": user.full_name,
                "is_admin": user.is_admin
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Registration error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.post("/api/auth/demo-login")
async def demo_login(db: Session = Depends(get_db)):
    """Đăng nhập demo (tạo tài khoản demo nếu chưa có)"""
    try:
        user = get_user_by_username(db, "demo")
        if not user:
            hashed_password = hash_password("demo123")
            user = User(
                username="demo",
                email="demo@classroom.com",
                hashed_password=hashed_password,
                full_name="Demo Teacher",
                is_active=True,
                is_admin=False,
                created_at=datetime.now()
            )
            db.add(user)
            db.commit()
            db.refresh(user)
        
        token_data = create_access_token(user)
        
        return {
            "status": "success",
            "message": "Đăng nhập demo thành công",
            **token_data
        }
        
    except Exception as e:
        logger.error(f"Demo login error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.get("/api/user/dashboard")
async def user_dashboard(
    token: str = Query(..., description="Access token"),
    db: Session = Depends(get_db)
):
    """Dashboard dành riêng cho user thường (giáo viên)"""
    try:
        user = await get_current_user_from_token(token, db)
        
        if user.is_admin:
            # Nếu là admin truy cập trang user, trả về thông báo
            return {
                "status": "success",
                "user_type": "admin",
                "message": "Bạn đang truy cập trang user với quyền admin",
                "redirect_suggested": True,
                "suggested_url": "/api/dashboard/stats"
            }
        
        # Lấy thông tin cơ bản cho user thường
        today = datetime.now().strftime("%Y-%m-%d")
        
        # Lấy lớp học của giáo viên (giả sử có field teacher_class)
        teacher_class = "Lớp 10A1"  # Trong thực tế lấy từ database
        
        # Lấy điểm danh hôm nay của lớp
        today_attendance = db.query(StudentData).filter(
            StudentData.date == today,
            StudentData.attendance_status.isnot(None),
            StudentData.class_name.ilike(f"%{teacher_class}%")
        ).all()
        
        present_today = sum(1 for record in today_attendance if record.attendance_status == "present")
        absent_today = sum(1 for record in today_attendance if record.attendance_status == "absent")
        total_today = len(today_attendance)
        
        return {
            "status": "success",
            "user_type": "user",
            "message": "hello user",
            "user_info": {
                "full_name": user.full_name,
                "username": user.username,
                "class": teacher_class
            },
            "today_summary": {
                "total_students": total_today,
                "present": present_today,
                "absent": absent_today,
                "attendance_rate": round((present_today / max(total_today, 1)) * 100, 1)
            },
            "features": [
                "Xem điểm danh lớp",
                "Ghi nhận cảm xúc học sinh",
                "Theo dõi hành vi",
                "Xem báo cáo lớp"
            ]
        }
        
    except HTTPException as e:
        raise
    except Exception as e:
        logger.error(f"Error in user_dashboard: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

# ==================== STUDENT DATA ENDPOINTS ====================
@app.get("/api/students")
async def get_students(
    class_name: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    recent_minutes: int = Query(30, ge=1, le=1440, description="Lấy dữ liệu trong X phút gần nhất"),
    db: Session = Depends(get_db)
):
    """Lấy danh sách học sinh duy nhất từ dữ liệu - ƯU TIÊN DỮ LIỆU GẦN ĐÂY - BỎ QUA UNKNOWN"""
    try:
        # Tính thời gian bắt đầu (X phút trước)
        recent_threshold = datetime.now() - timedelta(minutes=recent_minutes)
        
        print(f"📊 Getting students data from last {recent_minutes} minutes...")
        
        # 🔴 THAY ĐỔI: Ưu tiên lấy dữ liệu gần đây
        subquery = db.query(
            StudentData.student_id,
            StudentData.student_name,
            StudentData.class_name,
            func.max(StudentData.recorded_at).label('last_recorded'),
            func.count(StudentData.id).label('recent_count')
        ).filter(
            StudentData.recorded_at >= recent_threshold  # 🔴 CHỈ lấy dữ liệu gần đây
        ).group_by(
            StudentData.student_id,
            StudentData.student_name,
            StudentData.class_name
        ).subquery()
        
        query = db.query(
            subquery.c.student_id,
            subquery.c.student_name,
            subquery.c.class_name,
            subquery.c.last_recorded,
            subquery.c.recent_count
        ).order_by(
            subquery.c.last_recorded.desc()  # 🔴 Sắp xếp theo thời gian mới nhất
        )
        
        # Filters
        if class_name:
            query = query.filter(subquery.c.class_name.ilike(f"%{class_name}%"))
        
        if search:
            query = query.filter(
                or_(
                    subquery.c.student_name.ilike(f"%{search}%"),
                    subquery.c.student_id.ilike(f"%{search}%")
                )
            )
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        students = query.offset(offset).limit(limit).all()
        
        # Nếu không có dữ liệu gần đây, lấy tất cả
        if total_count == 0:
            print("⚠️ No recent data, fetching all students...")
            
            subquery_all = db.query(
                StudentData.student_id,
                StudentData.student_name,
                StudentData.class_name,
                func.max(StudentData.recorded_at).label('last_recorded'),
                func.count(StudentData.id).label('recent_count')
            ).group_by(
                StudentData.student_id,
                StudentData.student_name,
                StudentData.class_name
            ).subquery()
            
            query_all = db.query(
                subquery_all.c.student_id,
                subquery_all.c.student_name,
                subquery_all.c.class_name,
                subquery_all.c.last_recorded,
                subquery_all.c.recent_count
            ).order_by(subquery_all.c.last_recorded.desc())
            
            if class_name:
                query_all = query_all.filter(subquery_all.c.class_name.ilike(f"%{class_name}%"))
            
            if search:
                query_all = query_all.filter(
                    or_(
                        subquery_all.c.student_name.ilike(f"%{search}%"),
                        subquery_all.c.student_id.ilike(f"%{search}%")
                    )
                )
            
            total_count = query_all.count()
            students = query_all.offset(offset).limit(limit).all()
        
        # Lấy thống kê cho mỗi học sinh
        student_list = []
        for student in students:
            # BỎ QUA UNKNOWN STUDENTS
            if is_unknown_student(student.student_name, student.student_id):
                logger.debug(f"⏭️ Filtering out unknown student from results: {student.student_name}")
                continue
            
            # Lấy thông tin điểm danh từ 30 phút gần nhất
            recent_threshold = datetime.now() - timedelta(minutes=30)
            
            attendance_stats = db.query(
                func.count(StudentData.id).label('total_records'),
                func.sum(case((StudentData.attendance_status == 'present', 1), else_=0)).label('present_count'),
                func.sum(case((StudentData.attendance_status == 'absent', 1), else_=0)).label('absent_count'),
                func.sum(case((StudentData.attendance_status == 'late', 1), else_=0)).label('late_count')
            ).filter(
                StudentData.student_id == student.student_id,
                StudentData.attendance_status.isnot(None),
                StudentData.recorded_at >= recent_threshold  # 🔴 Chỉ lấy gần đây
            ).first()
            
            # Lấy điểm tập trung trung bình từ 30 phút gần nhất
            avg_focus = db.query(func.avg(StudentData.focus_score)).filter(
                StudentData.student_id == student.student_id,
                StudentData.focus_score.isnot(None),
                StudentData.recorded_at >= recent_threshold  # 🔴 Chỉ lấy gần đây
            ).scalar() or 0
            
            # Lấy điểm hành vi trung bình từ 30 phút gần nhất
            avg_behavior = db.query(func.avg(StudentData.behavior_score)).filter(
                StudentData.student_id == student.student_id,
                StudentData.behavior_score.isnot(None),
                StudentData.recorded_at >= recent_threshold  # 🔴 Chỉ lấy gần đây
            ).scalar() or 0
            
            # Lấy emotion mới nhất
            latest_emotion = db.query(
                StudentData.emotion,
                StudentData.emotion_confidence,
                StudentData.recorded_at
            ).filter(
                StudentData.student_id == student.student_id,
                StudentData.emotion.isnot(None)
            ).order_by(StudentData.recorded_at.desc()).first()
            
            # Lấy behavior mới nhất
            latest_behavior = db.query(
                StudentData.behavior_type,
                StudentData.behavior_details,
                StudentData.recorded_at
            ).filter(
                StudentData.student_id == student.student_id,
                StudentData.behavior_type.isnot(None)
            ).order_by(StudentData.recorded_at.desc()).first()
            
            student_list.append({
                "student_id": student.student_id,
                "student_name": student.student_name,
                "class_name": student.class_name,
                "last_recorded": student.last_recorded.isoformat() if student.last_recorded else None,
                "recent_detection_count": student.recent_count or 0,
                "stats": {
                    "total_records": attendance_stats.total_records or 0,
                    "attendance": {
                        "present": attendance_stats.present_count or 0,
                        "absent": attendance_stats.absent_count or 0,
                        "late": attendance_stats.late_count or 0
                    },
                    "avg_focus": round(avg_focus, 1),
                    "avg_behavior": round(avg_behavior, 1)
                },
                "latest_emotion": {
                    "emotion": latest_emotion.emotion if latest_emotion else "unknown",
                    "confidence": latest_emotion.emotion_confidence if latest_emotion else 0,
                    "time": latest_emotion.recorded_at.isoformat() if latest_emotion and latest_emotion.recorded_at else None
                } if latest_emotion else None,
                "latest_behavior": {
                    "type": latest_behavior.behavior_type if latest_behavior else "unknown",
                    "details": latest_behavior.behavior_details if latest_behavior else "",
                    "time": latest_behavior.recorded_at.isoformat() if latest_behavior and latest_behavior.recorded_at else None
                } if latest_behavior else None
            })
        
        print(f"✅ Found {len(student_list)} students in recent data (after unknown filter)")
        
        return {
            "status": "success",
            "count": len(student_list),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "data_source": "recent" if recent_minutes > 0 else "all",
            "recent_minutes": recent_minutes,
            "students": student_list
        }
        
    except Exception as e:
        logger.error(f"❌ Error in get_students: {e}")
        import traceback
        traceback.print_exc()
        
        # Fallback data với students từ AI detection
        current_time = datetime.now().isoformat()
        return {
            "status": "success",
            "count": 3,
            "total": 3,
            "page": 1,
            "limit": 20,
            "total_pages": 1,
            "data_source": "fallback",
            "recent_minutes": 30,
            "students": [
                {
                    "student_id": "AI_STUDENT_001",
                    "student_name": "Nam",  # 👈 Tên từ AI detection
                    "class_name": "AI Class",
                    "last_recorded": current_time,
                    "recent_detection_count": 5,
                    "stats": {
                        "total_records": 5,
                        "attendance": {"present": 5, "absent": 0, "late": 0},
                        "avg_focus": 82.5,
                        "avg_behavior": 78.3
                    },
                    "latest_emotion": {
                        "emotion": "neutral",
                        "confidence": 0.72,
                        "time": current_time
                    },
                    "latest_behavior": {
                        "type": "engagement",
                        "details": "raising_hand",
                        "time": current_time
                    }
                },
                {
                    "student_id": "AI_STUDENT_002",
                    "student_name": "Student 1",
                    "class_name": "AI Class",
                    "last_recorded": current_time,
                    "recent_detection_count": 3,
                    "stats": {
                        "total_records": 3,
                        "attendance": {"present": 3, "absent": 0, "late": 0},
                        "avg_focus": 75.8,
                        "avg_behavior": 72.1
                    },
                    "latest_emotion": {
                        "emotion": "happy",
                        "confidence": 0.85,
                        "time": current_time
                    },
                    "latest_behavior": {
                        "type": "normal",
                        "details": "sitting",
                        "time": current_time
                    }
                },
                {
                    "student_id": "AI_STUDENT_003",
                    "student_name": "Student 2",
                    "class_name": "AI Class",
                    "last_recorded": current_time,
                    "recent_detection_count": 2,
                    "stats": {
                        "total_records": 2,
                        "attendance": {"present": 2, "absent": 0, "late": 0},
                        "avg_focus": 68.3,
                        "avg_behavior": 65.7
                    },
                    "latest_emotion": {
                        "emotion": "sad",
                        "confidence": 0.65,
                        "time": current_time
                    },
                    "latest_behavior": {
                        "type": "discipline",
                        "details": "normal",
                        "time": current_time
                    }
                }
            ]
        }

@app.get("/api/student-data")
async def get_student_data(
    student_id: Optional[str] = None,
    student_name: Optional[str] = None,
    date: Optional[str] = None,
    class_name: Optional[str] = None,
    attendance_status: Optional[str] = None,
    behavior_type: Optional[str] = None,
    emotion: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu học sinh với pagination và filter - BỎ QUA UNKNOWN"""
    try:
        query = db.query(StudentData)
        
        # Filters
        if student_id:
            query = query.filter(StudentData.student_id == student_id)
        
        if student_name:
            query = query.filter(StudentData.student_name.ilike(f"%{student_name}%"))
        
        if date:
            query = query.filter(StudentData.date == date)
        
        if class_name:
            query = query.filter(StudentData.class_name.ilike(f"%{class_name}%"))
        
        if attendance_status:
            query = query.filter(StudentData.attendance_status == attendance_status)
        
        if behavior_type:
            query = query.filter(StudentData.behavior_type == behavior_type)
        
        if emotion:
            query = query.filter(StudentData.emotion == emotion)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        query = query.order_by(StudentData.recorded_at.desc())
        student_data = query.offset(offset).limit(limit).all()
        
        # Process results - BỎ QUA UNKNOWN
        data_list = []
        for data in student_data:
            try:
                # BỎ QUA UNKNOWN STUDENTS
                if is_unknown_student(data.student_name, data.student_id):
                    continue
                    
                data_dict = {
                    "id": data.id,
                    "student_id": data.student_id,
                    "student_name": data.student_name,
                    "date": data.date,
                    "attendance_status": data.attendance_status,
                    "check_in_time": data.check_in_time.isoformat() if data.check_in_time else None,
                    "check_out_time": data.check_out_time.isoformat() if data.check_out_time else None,
                    "attendance_notes": data.attendance_notes,
                    "emotion": data.emotion,
                    "emotion_confidence": data.emotion_confidence,
                    "behavior_type": data.behavior_type,
                    "behavior_score": data.behavior_score,
                    "behavior_details": data.behavior_details,
                    "class_name": data.class_name,
                    "session_id": data.session_id,
                    "recorded_by": data.recorded_by,
                    "recorded_at": data.recorded_at.isoformat() if data.recorded_at else None,
                    "focus_score": data.focus_score,
                    "concentration_level": data.concentration_level,
                    "focus_duration": data.focus_duration
                }
                data_list.append(data_dict)
            except Exception as e:
                logger.error(f"Error processing student data {data.id}: {e}")
                continue
        
        return {
            "status": "success",
            "count": len(data_list),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "student_data": data_list
        }
        
    except Exception as e:
        logger.error(f"Error in get_student_data: {e}")
        return {
            "status": "success",
            "count": 0,
            "total": 0,
            "page": page,
            "limit": limit,
            "total_pages": 0,
            "student_data": []
        }

@app.post("/api/student-data")
async def create_student_data(
    student_data: StudentDataCreate,
    db: Session = Depends(get_db)
):
    """Tạo bản ghi dữ liệu học sinh mới - BỎ QUA UNKNOWN"""
    try:
        # KIỂM TRA NẾU LÀ UNKNOWN STUDENT
        if is_unknown_student(student_data.student_name, student_data.student_id):
            logger.info(f"⏭️ Skipping unknown student: {student_data.student_name}")
            return {
                "status": "success",
                "message": "Skipped unknown student",
                "unknown_filtered": True,
                "timestamp": datetime.now().isoformat()
            }
        
        # Tạo bản ghi mới
        data = StudentData(
            student_id=student_data.student_id,
            student_name=student_data.student_name,
            date=student_data.date,
            attendance_status=student_data.attendance_status,
            emotion=student_data.emotion,
            emotion_confidence=student_data.emotion_confidence,
            behavior_type=student_data.behavior_type,
            behavior_score=student_data.behavior_score,
            behavior_details=student_data.behavior_details,
            class_name=student_data.class_name,
            session_id=student_data.session_id,
            recorded_by=student_data.recorded_by,
            check_in_time=student_data.check_in_time,
            check_out_time=student_data.check_out_time,
            attendance_notes=student_data.attendance_notes,
            focus_score=student_data.focus_score,
            concentration_level=student_data.concentration_level,
            focus_duration=student_data.focus_duration,
            recorded_at=datetime.now()
        )
        
        db.add(data)
        db.commit()
        db.refresh(data)
        
        # Broadcast real-time update
        await manager_ws.broadcast({
            "type": "student_data_update",
            "timestamp": datetime.now().isoformat(),
            "data": {
                "student_id": data.student_id,
                "student_name": data.student_name,
                "attendance_status": data.attendance_status,
                "emotion": data.emotion
            }
        })
        
        return {
            "status": "success",
            "message": "Tạo dữ liệu học sinh thành công",
            "data": {
                "id": data.id,
                "student_id": data.student_id,
                "student_name": data.student_name,
                "date": data.date,
                "attendance_status": data.attendance_status
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Create student data error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.put("/api/student-data/{data_id}")
async def update_student_data(
    data_id: int,
    student_data: StudentDataUpdate,
    db: Session = Depends(get_db)
):
    """Cập nhật dữ liệu học sinh"""
    try:
        data = db.query(StudentData).filter(StudentData.id == data_id).first()
        if not data:
            raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu học sinh")
        
        # Cập nhật các trường có giá trị
        update_data = student_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(data, field, value)
        
        db.commit()
        db.refresh(data)
        
        return {
            "status": "success",
            "message": "Cập nhật dữ liệu học sinh thành công",
            "data": {
                "id": data.id,
                "student_id": data.student_id,
                "student_name": data.student_name
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update student data error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

@app.delete("/api/student-data/{data_id}")
async def delete_student_data(
    data_id: int,
    db: Session = Depends(get_db)
):
    """Xóa dữ liệu học sinh"""
    try:
        data = db.query(StudentData).filter(StudentData.id == data_id).first()
        if not data:
            raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu học sinh")
        
        student_id = data.student_id
        student_name = data.student_name
        
        db.delete(data)
        db.commit()
        
        return {
            "status": "success",
            "message": "Xóa dữ liệu học sinh thành công"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete student data error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

# ==================== ATTENDANCE ENDPOINTS ====================
@app.get("/api/attendance")
async def get_attendance(
    date: Optional[str] = None,
    student_id: Optional[str] = None,
    class_name: Optional[str] = None,
    attendance_status: Optional[str] = None,
    include_student_info: bool = Query(True, description="Bao gồm thông tin học sinh từ bảng students"),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy lịch sử điểm danh - CẬP NHẬT với student info - BỎ QUA UNKNOWN"""
    try:
        if not date:
            date = datetime.now().strftime("%Y-%m-%d")
        
        query = db.query(StudentData).filter(
            StudentData.attendance_status.isnot(None)
        )
        
        if date:
            query = query.filter(StudentData.date == date)
        
        if student_id:
            query = query.filter(StudentData.student_id == student_id)
        
        if class_name:
            query = query.filter(StudentData.class_name == class_name)
        
        if attendance_status:
            query = query.filter(StudentData.attendance_status == attendance_status)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        query = query.order_by(StudentData.check_in_time.desc())
        attendance_records = query.offset(offset).limit(limit).all()
        
        # Process results - BỎ QUA UNKNOWN
        attendance_list = []
        for record in attendance_records:
            try:
                # BỎ QUA UNKNOWN STUDENTS
                if is_unknown_student(record.student_name, record.student_id):
                    continue
                    
                # Lấy thông tin học sinh từ bảng students nếu có
                student_info = None
                if include_student_info and record.student_id:
                    student = db.query(Student).filter(
                        Student.student_id == record.student_id
                    ).first()
                    
                    if student:
                        student_info = {
                            "student_name": student.student_name,
                            "class_name": student.class_name,
                            "is_active": student.is_active
                        }
                
                attendance_data = {
                    "id": record.id,
                    "student_id": record.student_id,
                    "student_name": record.student_name,
                    "date": record.date,
                    "check_in_time": record.check_in_time.isoformat() if record.check_in_time else None,
                    "check_out_time": record.check_out_time.isoformat() if record.check_out_time else None,
                    "attendance_status": record.attendance_status,
                    "attendance_notes": record.attendance_notes,
                    "class_name": record.class_name,
                    "recorded_by": record.recorded_by,
                    "recorded_at": record.recorded_at.isoformat() if record.recorded_at else None,
                    "emotion": record.emotion,
                    "behavior_score": record.behavior_score,
                    "focus_score": record.focus_score,
                    "student_info": student_info  # Thêm thông tin học sinh
                }
                attendance_list.append(attendance_data)
            except Exception as e:
                logger.error(f"Error processing attendance record {record.id}: {e}")
                continue
        
        return {
            "status": "success",
            "count": len(attendance_list),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "date": date,
            "attendance": attendance_list
        }
        
    except Exception as e:
        logger.error(f"Error in get_attendance: {e}")

@app.post("/api/attendance")
async def create_attendance(
    attendance_data: AttendanceCreate,
    db: Session = Depends(get_db)
):
    """Tạo bản ghi điểm danh - CHỈ CHO PHÉP 1 LẦN/NGÀY - BỎ QUA UNKNOWN"""
    try:
        # KIỂM TRA NẾU LÀ UNKNOWN STUDENT
        if is_unknown_student(attendance_data.student_name, attendance_data.student_id):
            logger.info(f"⏭️ Skipping unknown student attendance: {attendance_data.student_name}")
            return {
                "status": "success",
                "message": "Skipped unknown student",
                "unknown_filtered": True,
                "timestamp": datetime.now().isoformat()
            }
        
        today = datetime.now().strftime("%Y-%m-%d")
        
        # KIỂM TRA: Học sinh đã điểm danh hôm nay chưa?
        existing_attendance = db.query(StudentData).filter(
            StudentData.student_id == attendance_data.student_id,
            StudentData.date == today,
            StudentData.attendance_status.isnot(None)
        ).first()
        
        if existing_attendance:
            raise HTTPException(
                status_code=400, 
                detail=f"Học sinh {attendance_data.student_name} đã điểm danh hôm nay lúc {existing_attendance.check_in_time.strftime('%H:%M')}"
            )
        
        # Kiểm tra học sinh có trong danh sách lớp không (nếu cần)
        class_student = db.query(ClassStudent).filter(
            ClassStudent.student_id == attendance_data.student_id,
            ClassStudent.is_active == True
        ).first()
        
        if not class_student:
            # Vẫn cho điểm danh nhưng cảnh báo
            logger.warning(f"Học sinh {attendance_data.student_id} không có trong danh sách lớp")
        
        # Tạo bản ghi điểm danh
        attendance = StudentData(
            student_id=attendance_data.student_id,
            student_name=attendance_data.student_name,
            date=today,
            attendance_status=attendance_data.attendance_status,
            check_in_time=attendance_data.check_in_time or datetime.now(),
            check_out_time=None,  # Chưa checkout
            attendance_notes=attendance_data.attendance_notes,
            class_name=attendance_data.class_name or (class_student.class_name if class_student else "Chưa xác định"),
            recorded_by=attendance_data.recorded_by or "system",
            recorded_at=datetime.now()
        )
        
        db.add(attendance)
        db.commit()
        db.refresh(attendance)
        
        # Broadcast real-time update
        await manager_ws.broadcast({
            "type": "attendance_update",
            "timestamp": datetime.now().isoformat(),
            "data": {
                "student_id": attendance.student_id,
                "student_name": attendance.student_name,
                "attendance_status": attendance.attendance_status,
                "check_in_time": attendance.check_in_time.isoformat(),
                "is_first_today": True  # Thêm flag mới
            }
        })
        
        return {
            "status": "success",
            "message": "Điểm danh thành công",
            "attendance": {
                "id": attendance.id,
                "student_id": attendance.student_id,
                "student_name": attendance.student_name,
                "date": attendance.date,
                "attendance_status": attendance.attendance_status,
                "check_in_time": attendance.check_in_time.isoformat() if attendance.check_in_time else None
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create attendance error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

# ==================== FIX: ADD AI-COMPATIBLE ATTENDANCE ENDPOINT ====================        
@app.put("/api/attendance/{attendance_id}/checkout")
async def check_out(
    attendance_id: int,
    db: Session = Depends(get_db)
):
    """Check out cho điểm danh"""
    try:
        attendance = db.query(StudentData).filter(StudentData.id == attendance_id).first()
        if not attendance:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi điểm danh")
        
        if attendance.check_out_time:
            raise HTTPException(status_code=400, detail="Học sinh đã check out")
        
        attendance.check_out_time = datetime.now()
        db.commit()
        
        # Broadcast real-time update
        await manager_ws.broadcast({
            "type": "checkout_update",
            "timestamp": datetime.now().isoformat(),
            "data": {
                "student_id": attendance.student_id,
                "student_name": attendance.student_name,
                "check_out_time": attendance.check_out_time.isoformat()
            }
        })
        
        return {
            "status": "success",
            "message": "Check out thành công",
            "check_out_time": attendance.check_out_time.isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Checkout error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

# ==================== EMOTION ENDPOINTS ====================
@app.get("/api/emotion")
async def get_emotion_data(
    date: Optional[str] = None,
    student_id: Optional[str] = None,
    emotion: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu cảm xúc - BỎ QUA UNKNOWN"""
    try:
        query = db.query(StudentData).filter(
            StudentData.emotion.isnot(None)
        )
        
        if date:
            query = query.filter(StudentData.date == date)
        
        if student_id:
            query = query.filter(StudentData.student_id == student_id)
        
        if emotion:
            query = query.filter(StudentData.emotion == emotion)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        query = query.order_by(StudentData.recorded_at.desc())
        emotion_records = query.offset(offset).limit(limit).all()
        
        # Filter out unknown students
        filtered_records = []
        for record in emotion_records:
            if not is_unknown_student(record.student_name, record.student_id):
                filtered_records.append(record)
        
        return {
            "status": "success",
            "count": len(filtered_records),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "emotion_data": [
                {
                    "id": record.id,
                    "student_id": record.student_id,
                    "student_name": record.student_name,
                    "emotion": record.emotion,
                    "emotion_confidence": record.emotion_confidence,
                    "date": record.date,
                    "session_id": record.session_id,
                    "recorded_by": record.recorded_by,
                    "recorded_at": record.recorded_at.isoformat() if record.recorded_at else None
                }
                for record in filtered_records
            ]
        }
        
    except Exception as e:
        logger.error(f"Error in get_emotion_data: {e}")
        today = datetime.now().strftime("%Y-%m-%d")
        return {
            "status": "success",
            "count": 3,
            "total": 3,
            "page": 1,
            "limit": 20,
            "total_pages": 1,
            "emotion_data": [
                {
                    "id": 1,
                    "student_id": "SV001",
                    "student_name": "Nguyễn Văn A",
                    "emotion": "happy",
                    "emotion_confidence": 0.85,
                    "date": today,
                    "session_id": "SESS001",
                    "recorded_by": "teacher1",
                    "recorded_at": datetime.now().isoformat()
                },
                {
                    "id": 2,
                    "student_id": "SV002",
                    "student_name": "Trần Thị B",
                    "emotion": "neutral",
                    "emotion_confidence": 0.72,
                    "date": today,
                    "session_id": "SESS001",
                    "recorded_by": "teacher1",
                    "recorded_at": datetime.now().isoformat()
                },
                {
                    "id": 3,
                    "student_id": "SV003",
                    "student_name": "Lê Văn C",
                    "emotion": "sad",
                    "emotion_confidence": 0.65,
                    "date": today,
                    "session_id": "SESS001",
                    "recorded_by": "teacher1",
                    "recorded_at": datetime.now().isoformat()
                }
            ]
        }

@app.post("/api/emotion")
async def create_emotion(
    emotion_data: EmotionCreate,
    db: Session = Depends(get_db)
):
    """Tạo bản ghi cảm xúc - BỎ QUA UNKNOWN"""
    try:
        # KIỂM TRA NẾU LÀ UNKNOWN STUDENT
        if is_unknown_student(emotion_data.student_name, emotion_data.student_id):
            logger.info(f"⏭️ Skipping unknown student emotion: {emotion_data.student_name}")
            return {
                "status": "success",
                "message": "Skipped unknown student",
                "unknown_filtered": True,
                "timestamp": datetime.now().isoformat()
            }
        
        date = emotion_data.date or datetime.now().strftime("%Y-%m-%d")
        
        # Tạo bản ghi cảm xúc
        emotion = StudentData(
            student_id=emotion_data.student_id,
            student_name=emotion_data.student_name,
            emotion=emotion_data.emotion,
            emotion_confidence=emotion_data.emotion_confidence,
            date=date,
            session_id=emotion_data.session_id,
            recorded_by=emotion_data.recorded_by,
            recorded_at=datetime.now()
        )
        
        db.add(emotion)
        db.commit()
        db.refresh(emotion)
        
        # Broadcast real-time update
        await manager_ws.broadcast({
            "type": "emotion_update",
            "timestamp": datetime.now().isoformat(),
            "data": {
                "student_id": emotion.student_id,
                "student_name": emotion.student_name,
                "emotion": emotion.emotion,
                "emotion_confidence": emotion.emotion_confidence
            }
        })
        
        return {
            "status": "success",
            "message": "Ghi nhận cảm xúc thành công",
            "emotion": {
                "id": emotion.id,
                "student_id": emotion.student_id,
                "student_name": emotion.student_name,
                "emotion": emotion.emotion,
                "emotion_confidence": emotion.emotion_confidence
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Create emotion error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

# ==================== BEHAVIOR ENDPOINTS ====================
@app.get("/api/behavior")
async def get_behavior_data(
    date: Optional[str] = None,
    student_id: Optional[str] = None,
    behavior_type: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu hành vi - BỎ QUA UNKNOWN"""
    try:
        query = db.query(StudentData).filter(
            StudentData.behavior_type.isnot(None)
        )
        
        if date:
            query = query.filter(StudentData.date == date)
        
        if student_id:
            query = query.filter(StudentData.student_id == student_id)
        
        if behavior_type:
            query = query.filter(StudentData.behavior_type == behavior_type)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        query = query.order_by(StudentData.recorded_at.desc())
        behavior_records = query.offset(offset).limit(limit).all()
        
        # Filter out unknown students
        filtered_records = []
        for record in behavior_records:
            if not is_unknown_student(record.student_name, record.student_id):
                filtered_records.append(record)
        
        return {
            "status": "success",
            "count": len(filtered_records),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "behavior_data": [
                {
                    "id": record.id,
                    "student_id": record.student_id,
                    "student_name": record.student_name,
                    "behavior_type": record.behavior_type,
                    "behavior_score": record.behavior_score,
                    "behavior_details": record.behavior_details,
                    "date": record.date,
                    "session_id": record.session_id,
                    "recorded_by": record.recorded_by,
                    "recorded_at": record.recorded_at.isoformat() if record.recorded_at else None
                }
                for record in filtered_records
            ]
        }
        
    except Exception as e:
        logger.error(f"Error in get_behavior_data: {e}")
        today = datetime.now().strftime("%Y-%m-%d")
        return {
            "status": "success",
            "count": 3,
            "total": 3,
            "page": 1,
            "limit": 20,
            "total_pages": 1,
            "behavior_data": [
                {
                    "id": 1,
                    "student_id": "SV001",
                    "student_name": "Nguyễn Văn A",
                    "behavior_type": "engagement",
                    "behavior_score": 85.5,
                    "behavior_details": "Tích cực tham gia thảo luận",
                    "date": today,
                    "session_id": "SESS001",
                    "recorded_by": "teacher1",
                    "recorded_at": datetime.now().isoformat()
                },
                {
                    "id": 2,
                    "student_id": "SV002",
                    "student_name": "Trần Thị B",
                    "behavior_type": "participation",
                    "behavior_score": 90.0,
                    "behavior_details": "Phát biểu xây dựng bài",
                    "date": today,
                    "session_id": "SESS001",
                    "recorded_by": "teacher1",
                    "recorded_at": datetime.now().isoformat()
                },
                {
                    "id": 3,
                    "student_id": "SV003",
                    "student_name": "Lê Văn C",
                    "behavior_type": "discipline",
                    "behavior_score": 95.0,
                    "behavior_details": "Giữ trật tự trong giờ học",
                    "date": today,
                    "session_id": "SESS001",
                    "recorded_by": "teacher1",
                    "recorded_at": datetime.now().isoformat()
                }
            ]
        }

@app.post("/api/behavior")
async def create_behavior(
    behavior_data: BehaviorCreate,
    db: Session = Depends(get_db)
):
    """Tạo bản ghi hành vi - BỎ QUA UNKNOWN"""
    try:
        # KIỂM TRA NẾU LÀ UNKNOWN STUDENT
        if is_unknown_student(behavior_data.student_name, behavior_data.student_id):
            logger.info(f"⏭️ Skipping unknown student behavior: {behavior_data.student_name}")
            return {
                "status": "success",
                "message": "Skipped unknown student",
                "unknown_filtered": True,
                "timestamp": datetime.now().isoformat()
            }
        
        date = behavior_data.date or datetime.now().strftime("%Y-%m-%d")
        
        # Tạo bản ghi hành vi
        behavior = StudentData(
            student_id=behavior_data.student_id,
            student_name=behavior_data.student_name,
            behavior_type=behavior_data.behavior_type,
            behavior_score=behavior_data.behavior_score,
            behavior_details=behavior_data.behavior_details,
            date=date,
            session_id=behavior_data.session_id,
            recorded_by=behavior_data.recorded_by,
            recorded_at=datetime.now()
        )
        
        db.add(behavior)
        db.commit()
        db.refresh(behavior)
        
        # Broadcast real-time update
        await manager_ws.broadcast({
            "type": "behavior_update",
            "timestamp": datetime.now().isoformat(),
            "data": {
                "student_id": behavior.student_id,
                "student_name": behavior.student_name,
                "behavior_type": behavior.behavior_type,
                "behavior_score": behavior.behavior_score
            }
        })
        
        return {
            "status": "success",
            "message": "Ghi nhận hành vi thành công",
            "behavior": {
                "id": behavior.id,
                "student_id": behavior.student_id,
                "student_name": behavior.student_name,
                "behavior_type": behavior.behavior_type,
                "behavior_score": behavior.behavior_score
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Create behavior error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")


# ==================== ENGAGEMENT ENDPOINTS ====================
class EngagementCreate(BaseModel):
    """Model cho engagement data từ AI system"""
    student_id: Optional[str] = None
    student_name: Optional[str] = None
    name: Optional[str] = None  # Alias
    engagement_score: float  # 0-100 điểm
    concentration_level: str  # very_high, high, medium, low, very_low
    emotion: Optional[str] = None
    behavior: Optional[str] = None
    emotion_confidence: Optional[float] = None
    date: Optional[str] = None
    session_id: Optional[str] = None
    recorded_by: Optional[str] = None
    class_name: Optional[str] = None

@app.post("/api/ai/engagement")
async def create_ai_engagement(
    engagement_data: EngagementCreate,
    db: Session = Depends(get_db)
):
    """Tạo bản ghi engagement từ AI system - BỎ QUA UNKNOWN"""
    try:
        # Map fields
        student_id = engagement_data.student_id or f"ENGAGE_{int(time.time()) % 10000:04d}"
        student_name = engagement_data.student_name or engagement_data.name or ""
        
        # KIỂM TRA NẾU LÀ UNKNOWN STUDENT
        if is_unknown_student(student_name, student_id):
            logger.info(f"⏭️ Skipping unknown student engagement: {student_name}")
            return {
                "status": "success",
                "message": "Skipped unknown student",
                "unknown_filtered": True,
                "timestamp": datetime.now().isoformat()
            }
        
        # Tạo bản ghi
        record = StudentData(
            student_id=student_id,
            student_name=student_name,
            date=engagement_data.date or datetime.now().strftime("%Y-%m-%d"),
            
            # Dữ liệu emotion
            emotion=engagement_data.emotion or "neutral",
            emotion_confidence=engagement_data.emotion_confidence or 0.5,
            
            # Dữ liệu behavior
            behavior_type="engagement",
            behavior_score=engagement_data.engagement_score,
            behavior_details=engagement_data.behavior or "AI detected behavior",
            
            # Dữ liệu focus/engagement
            focus_score=engagement_data.engagement_score,
            concentration_level=engagement_data.concentration_level,
            focus_duration=45.0,  # Mặc định 45 phút
            
            # Dữ liệu attendance (auto-mark present)
            attendance_status="present",
            check_in_time=datetime.now(),
            
            # Metadata
            class_name=engagement_data.class_name or "AI Classroom",
            session_id=engagement_data.session_id or f"ENGAGE_{int(time.time())}",
            recorded_by=engagement_data.recorded_by or "AI System",
            recorded_at=datetime.now()
        )
        
        db.add(record)
        db.commit()
        db.refresh(record)
        
        logger.info(f"✅ AI engagement recorded: {student_name} - {engagement_data.concentration_level} ({engagement_data.engagement_score})")
        
        # Broadcast real-time update
        await manager_ws.broadcast({
            "type": "engagement_update",
            "timestamp": datetime.now().isoformat(),
            "data": {
                "student_id": record.student_id,
                "student_name": record.student_name,
                "engagement_score": record.focus_score,
                "concentration_level": record.concentration_level,
                "emotion": record.emotion,
                "behavior": record.behavior_details
            }
        })
        
        return {
            "status": "success",
            "message": "AI engagement recorded successfully",
            "engagement": {
                "id": record.id,
                "student_id": record.student_id,
                "student_name": record.student_name,
                "engagement_score": record.focus_score,
                "concentration_level": record.concentration_level,
                "emotion": record.emotion
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Create AI engagement error: {e}")
        return JSONResponse(
            status_code=200,
            content={
                "status": "error",
                "message": f"Error recording engagement: {str(e)}",
                "timestamp": datetime.now().isoformat()
            }
        )

@app.get("/api/engagement/realtime")
async def get_realtime_engagement(
    recent_minutes: int = Query(5, ge=1, le=60, description="Lấy dữ liệu X phút gần nhất"),
    db: Session = Depends(get_db)
):
    """Lấy engagement data real-time từ AI system - BỎ QUA UNKNOWN"""
    try:
        # Tính thời gian bắt đầu
        time_threshold = datetime.now() - timedelta(minutes=recent_minutes)
        
        # Lấy dữ liệu gần đây
        engagement_records = db.query(StudentData).filter(
            StudentData.recorded_at >= time_threshold,
            StudentData.focus_score.isnot(None),
            or_(
                StudentData.recorded_by == "AI System",
                StudentData.recorded_by == "AI Recognition System",
                StudentData.class_name == "AI Classroom"
            )
        ).order_by(StudentData.recorded_at.desc()).all()
        
        # FILTER OUT UNKNOWN STUDENTS
        valid_records = []
        for record in engagement_records:
            if not is_unknown_student(record.student_name, record.student_id):
                valid_records.append(record)
        
        logger.info(f"🔄 Engagement data: {len(engagement_records)} total, {len(valid_records)} after unknown filter")
        
        # Nhóm theo student
        students_dict = {}
        for record in valid_records:
            student_key = f"{record.student_id}_{record.student_name}"
            
            if student_key not in students_dict:
                students_dict[student_key] = {
                    "student_id": record.student_id,
                    "student_name": record.student_name,
                    "latest_engagement": record.focus_score,
                    "latest_concentration": record.concentration_level,
                    "latest_emotion": record.emotion,
                    "latest_behavior": record.behavior_type,
                    "latest_time": record.recorded_at,
                    "engagement_history": [],
                    "emotion_history": []
                }
            
            # Thêm vào history
            students_dict[student_key]["engagement_history"].append({
                "score": record.focus_score,
                "concentration": record.concentration_level,
                "time": record.recorded_at.isoformat()
            })
            
            if record.emotion:
                students_dict[student_key]["emotion_history"].append({
                    "emotion": record.emotion,
                    "confidence": record.emotion_confidence,
                    "time": record.recorded_at.isoformat()
                })
        
        # Chuyển dict sang list
        students_list = list(students_dict.values())
        
        # Tính stats tổng
        total_students = len(students_list)
        if total_students > 0:
            avg_engagement = np.mean([s["latest_engagement"] for s in students_list])
            emotion_counts = {}
            for student in students_list:
                emotion = student["latest_emotion"]
                emotion_counts[emotion] = emotion_counts.get(emotion, 0) + 1
            
            top_emotion = max(emotion_counts.items(), key=lambda x: x[1])[0] if emotion_counts else "unknown"
        else:
            avg_engagement = 0
            top_emotion = "unknown"
        
        return {
            "status": "success",
            "timestamp": datetime.now().isoformat(),
            "data_source": "realtime",
            "recent_minutes": recent_minutes,
            "summary": {
                "total_students": total_students,
                "avg_engagement": round(avg_engagement, 1),
                "top_emotion": top_emotion,
                "data_freshness": "live" if recent_minutes <= 5 else "recent"
            },
            "students": students_list
        }
        
    except Exception as e:
        logger.error(f"Error in get_realtime_engagement: {e}")
        return {
            "status": "success",
            "timestamp": datetime.now().isoformat(),
            "data_source": "demo",
            "recent_minutes": recent_minutes,
            "summary": {
                "total_students": 3,
                "avg_engagement": 78.5,
                "top_emotion": "neutral",
                "data_freshness": "demo"
            },
            "students": [
                {
                    "student_id": "AI_001",
                    "student_name": "Student 1",
                    "latest_engagement": 85.2,
                    "latest_concentration": "high",
                    "latest_emotion": "happy",
                    "latest_behavior": "writing",
                    "latest_time": datetime.now().isoformat()
                },
                {
                    "student_id": "AI_002",
                    "student_name": "Student 2",
                    "latest_engagement": 72.8,
                    "latest_concentration": "medium",
                    "latest_emotion": "neutral",
                    "latest_behavior": "look_straight",
                    "latest_time": datetime.now().isoformat()
                },
                {
                    "student_id": "AI_003",
                    "student_name": "Student 3",
                    "latest_engagement": 65.5,
                    "latest_concentration": "low",
                    "latest_emotion": "sad",
                    "latest_behavior": "look_around",
                    "latest_time": datetime.now().isoformat()
                }
            ]
        }
        
@app.get("/api/focus")
async def get_focus_data(
    date: Optional[str] = None,
    student_id: Optional[str] = None,
    concentration_level: Optional[str] = None,
    include_emotion: bool = Query(True, description="Bao gồm dữ liệu cảm xúc"),
    include_behavior: bool = Query(True, description="Bao gồm dữ liệu hành vi"),
    recent_minutes: int = Query(30, ge=1, le=1440, description="Lấy dữ liệu trong X phút gần nhất"),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu độ tập trung KÈM cảm xúc và hành vi - BỎ QUA UNKNOWN"""
    try:
        # Tính thời gian bắt đầu cho dữ liệu gần đây
        recent_threshold = datetime.now() - timedelta(minutes=recent_minutes)
        
        query = db.query(StudentData).filter(
            StudentData.focus_score.isnot(None)
        )
        
        # Filter theo thời gian gần đây
        if recent_minutes > 0:
            query = query.filter(StudentData.recorded_at >= recent_threshold)
        
        if date:
            query = query.filter(StudentData.date == date)
        
        if student_id:
            query = query.filter(StudentData.student_id == student_id)
        
        if concentration_level:
            query = query.filter(StudentData.concentration_level == concentration_level)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        query = query.order_by(StudentData.recorded_at.desc())
        focus_records = query.offset(offset).limit(limit).all()
        
        # Chuẩn bị dữ liệu trả về - BỎ QUA UNKNOWN
        focus_data = []
        for record in focus_records:
            # BỎ QUA UNKNOWN STUDENTS
            if is_unknown_student(record.student_name, record.student_id):
                continue
            
            # Lấy dữ liệu cảm xúc nếu có và được yêu cầu
            emotion_info = None
            if include_emotion and (record.emotion or record.emotion_confidence):
                emotion_info = {
                    "emotion": record.emotion,
                    "emotion_confidence": record.emotion_confidence,
                    "emotion_recorded_at": record.recorded_at.isoformat() if record.recorded_at else None
                }
            
            # Lấy dữ liệu hành vi nếu có và được yêu cầu
            behavior_info = None
            if include_behavior and (record.behavior_type or record.behavior_score or record.behavior_details):
                behavior_info = {
                    "behavior_type": record.behavior_type,
                    "behavior_score": record.behavior_score,
                    "behavior_details": record.behavior_details,
                    "behavior_recorded_at": record.recorded_at.isoformat() if record.recorded_at else None
                }
            
            # Lấy dữ liệu attendance nếu có
            attendance_info = None
            if record.attendance_status:
                attendance_info = {
                    "attendance_status": record.attendance_status,
                    "check_in_time": record.check_in_time.isoformat() if record.check_in_time else None,
                    "attendance_notes": record.attendance_notes
                }
            
            # Tạo bản ghi đầy đủ
            focus_record = {
                "id": record.id,
                "student_id": record.student_id,
                "student_name": record.student_name,
                
                # Dữ liệu độ tập trung
                "focus_score": record.focus_score,
                "concentration_level": record.concentration_level,
                "focus_duration": record.focus_duration,
                
                # Dữ liệu cảm xúc
                "emotion": emotion_info,
                
                # Dữ liệu hành vi
                "behavior": behavior_info,
                
                # Dữ liệu điểm danh
                "attendance": attendance_info,
                
                # Metadata
                "date": record.date,
                "class_name": record.class_name,
                "session_id": record.session_id,
                "recorded_by": record.recorded_by,
                "recorded_at": record.recorded_at.isoformat() if record.recorded_at else None,
                
                # Thông tin thống kê thêm
                "comprehensive_score": None,  # Có thể tính sau
                "engagement_level": None  # Có thể tính dựa trên focus + behavior
            }
            
            # Tính điểm tổng hợp nếu có đủ dữ liệu
            if record.focus_score is not None:
                comprehensive_score = record.focus_score                
                focus_record["comprehensive_score"] = round(comprehensive_score, 1)
                
                # Xác định mức độ engagement
                if comprehensive_score >= 85:
                    focus_record["engagement_level"] = "very_high"
                elif comprehensive_score >= 70:
                    focus_record["engagement_level"] = "high"
                elif comprehensive_score >= 55:
                    focus_record["engagement_level"] = "medium"
                else:
                    focus_record["engagement_level"] = "low"
            
            focus_data.append(focus_record)
        
        # Tính thống kê tổng hợp
        stats = {
            "total_records": len(focus_data),
            "avg_focus_score": 0,
            "avg_behavior_score": 0,
            "emotion_distribution": {},
            "concentration_distribution": {}
        }
        
        if focus_data:
            # Tính điểm trung bình
            focus_scores = [r["focus_score"] for r in focus_data if r["focus_score"] is not None]
            behavior_scores = [r["behavior"]["behavior_score"] for r in focus_data if r["behavior"] and r["behavior"]["behavior_score"] is not None]
            
            if focus_scores:
                stats["avg_focus_score"] = round(np.mean(focus_scores), 1)
                stats["min_focus_score"] = round(min(focus_scores), 1)
                stats["max_focus_score"] = round(max(focus_scores), 1)
            
            if behavior_scores:
                stats["avg_behavior_score"] = round(np.mean(behavior_scores), 1)
                stats["min_behavior_score"] = round(min(behavior_scores), 1)
                stats["max_behavior_score"] = round(max(behavior_scores), 1)
            
            # Phân phối cảm xúc
            for record in focus_data:
                if record["emotion"] and record["emotion"]["emotion"]:
                    emotion = record["emotion"]["emotion"]
                    stats["emotion_distribution"][emotion] = stats["emotion_distribution"].get(emotion, 0) + 1
                
                if record["concentration_level"]:
                    concentration = record["concentration_level"]
                    stats["concentration_distribution"][concentration] = stats["concentration_distribution"].get(concentration, 0) + 1
            
            # Tìm cảm xúc phổ biến nhất
            if stats["emotion_distribution"]:
                top_emotion = max(stats["emotion_distribution"].items(), key=lambda x: x[1])
                stats["top_emotion"] = top_emotion[0]
                stats["top_emotion_count"] = top_emotion[1]
        
        return {
            "status": "success",
            "count": len(focus_data),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "recent_minutes": recent_minutes,
            "stats": stats,
            "focus_data": focus_data
        }
        
    except Exception as e:
        logger.error(f"❌ Error in get_focus_data: {e}")
        
        # Fallback với dữ liệu demo đầy đủ
        today = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().isoformat()
        
        demo_data = [
            {
                "id": 1,
                "student_id": "SV001",
                "student_name": "Nguyễn Văn A",
                "focus_score": 85.0,
                "concentration_level": "high",
                "focus_duration": 45.5,
                "emotion": {
                    "emotion": "happy",
                    "emotion_confidence": 0.85,
                    "emotion_recorded_at": current_time
                },
                "behavior": {
                    "behavior_type": "engagement",
                    "behavior_score": 88.0,
                    "behavior_details": "Tích cực tham gia thảo luận",
                    "behavior_recorded_at": current_time
                },
                "attendance": {
                    "attendance_status": "present",
                    "check_in_time": "07:30:00",
                    "attendance_notes": "Đúng giờ"
                },
                "date": today,
                "class_name": "Lớp 10A1",
                "session_id": "SESS001",
                "recorded_by": "teacher1",
                "recorded_at": current_time,
                "comprehensive_score": 86.1,
                "engagement_level": "very_high"
            },
            {
                "id": 2,
                "student_id": "SV002",
                "student_name": "Trần Thị B",
                "focus_score": 72.5,
                "concentration_level": "medium",
                "focus_duration": 38.0,
                "emotion": {
                    "emotion": "neutral",
                    "emotion_confidence": 0.72,
                    "emotion_recorded_at": current_time
                },
                "behavior": {
                    "behavior_type": "participation",
                    "behavior_score": 75.0,
                    "behavior_details": "Phát biểu xây dựng bài",
                    "behavior_recorded_at": current_time
                },
                "attendance": {
                    "attendance_status": "present",
                    "check_in_time": "07:35:00",
                    "attendance_notes": None
                },
                "date": today,
                "class_name": "Lớp 10A1",
                "session_id": "SESS001",
                "recorded_by": "teacher1",
                "recorded_at": current_time,
                "comprehensive_score": 73.3,
                "engagement_level": "high"
            },
            {
                "id": 3,
                "student_id": "SV003",
                "student_name": "Lê Văn C",
                "focus_score": 60.0,
                "concentration_level": "low",
                "focus_duration": 25.5,
                "emotion": {
                    "emotion": "sad",
                    "emotion_confidence": 0.65,
                    "emotion_recorded_at": current_time
                },
                "behavior": {
                    "behavior_type": "discipline",
                    "behavior_score": 55.0,
                    "behavior_details": "Thỉnh thoảng mất tập trung",
                    "behavior_recorded_at": current_time
                },
                "attendance": {
                    "attendance_status": "present",
                    "check_in_time": "08:00:00",
                    "attendance_notes": "Đến muộn"
                },
                "date": today,
                "class_name": "Lớp 10A1",
                "session_id": "SESS001",
                "recorded_by": "teacher1",
                "recorded_at": current_time,
                "comprehensive_score": 58.5,
                "engagement_level": "medium"
            }
        ]
        
        # Demo stats
        demo_stats = {
            "total_records": 3,
            "avg_focus_score": 72.5,
            "avg_behavior_score": 72.7,
            "min_focus_score": 60.0,
            "max_focus_score": 85.0,
            "min_behavior_score": 55.0,
            "max_behavior_score": 88.0,
            "emotion_distribution": {"happy": 1, "neutral": 1, "sad": 1},
            "concentration_distribution": {"high": 1, "medium": 1, "low": 1},
            "top_emotion": "happy",
            "top_emotion_count": 1
        }
        
        return {
            "status": "success",
            "count": 3,
            "total": 3,
            "page": page,
            "limit": limit,
            "total_pages": 1,
            "recent_minutes": recent_minutes,
            "stats": demo_stats,
            "focus_data": demo_data,
            "data_source": "fallback"
        }

@app.post("/api/focus")
async def create_focus(
    focus_data: FocusCreate,
    db: Session = Depends(get_db)
):
    """Tạo bản ghi độ tập trung KÈM cảm xúc và hành vi - BỎ QUA UNKNOWN"""
    try:
        # KIỂM TRA NẾU LÀ UNKNOWN STUDENT
        if is_unknown_student(focus_data.student_name, focus_data.student_id):
            logger.info(f"⏭️ Skipping unknown student focus: {focus_data.student_name}")
            return {
                "status": "success",
                "message": "Skipped unknown student",
                "unknown_filtered": True,
                "timestamp": datetime.now().isoformat()
            }
        
        date = focus_data.date or datetime.now().strftime("%Y-%m-%d")
        
        # Tạo bản ghi độ tập trung với đầy đủ thông tin
        focus = StudentData(
            student_id=focus_data.student_id,
            student_name=focus_data.student_name,
            
            # Dữ liệu độ tập trung
            focus_score=focus_data.focus_score,
            concentration_level=focus_data.concentration_level,
            focus_duration=focus_data.focus_duration,
            
            # Dữ liệu cảm xúc (nếu có)
            emotion=focus_data.emotion,
            emotion_confidence=focus_data.emotion_confidence,
            
            # Dữ liệu hành vi (nếu có)
            behavior_type=focus_data.behavior_type,
            behavior_score=focus_data.behavior_score,
            behavior_details=focus_data.behavior_details,
            
            # Dữ liệu điểm danh (nếu có)
            attendance_status=focus_data.attendance_status or "present",
            check_in_time=focus_data.check_in_time or datetime.now(),
            
            # Metadata
            date=date,
            class_name=focus_data.class_name or "Chưa xác định",
            session_id=focus_data.session_id or f"FOCUS_{int(time.time())}",
            recorded_by=focus_data.recorded_by or "system",
            recorded_at=datetime.now()
        )
        
        db.add(focus)
        db.commit()
        db.refresh(focus)
        
        # Tính điểm tổng hợp
        comprehensive_score = focus.focus_score
        
        # Xác định mức độ engagement
        engagement_level = "unknown"
        if comprehensive_score >= 85:
            engagement_level = "very_high"
        elif comprehensive_score >= 70:
            engagement_level = "high"
        elif comprehensive_score >= 55:
            engagement_level = "medium"
        else:
            engagement_level = "low"
        
        # Broadcast real-time update với đầy đủ thông tin
        await manager_ws.broadcast({
            "type": "focus_update",
            "timestamp": datetime.now().isoformat(),
            "data": {
                "student_id": focus.student_id,
                "student_name": focus.student_name,
                
                # Dữ liệu độ tập trung
                "focus_score": focus.focus_score,
                "concentration_level": focus.concentration_level,
                "focus_duration": focus.focus_duration,
                
                # Dữ liệu cảm xúc
                "emotion": focus.emotion,
                "emotion_confidence": focus.emotion_confidence,
                
                # Dữ liệu hành vi
                "behavior_type": focus.behavior_type,
                "behavior_score": focus.behavior_score,
                "behavior_details": focus.behavior_details,
                
                # Thông tin tính toán
                "comprehensive_score": round(comprehensive_score, 1),
                "engagement_level": engagement_level,
                
                # Metadata
                "class_name": focus.class_name,
                "session_id": focus.session_id
            }
        })
        
        logger.info(f"✅ Focus created: {focus.student_name} - {focus.concentration_level} ({focus.focus_score})")
        
        return {
            "status": "success",
            "message": "Ghi nhận độ tập trung thành công",
            "focus": {
                "id": focus.id,
                "student_id": focus.student_id,
                "student_name": focus.student_name,
                
                # Dữ liệu độ tập trung
                "focus_score": focus.focus_score,
                "concentration_level": focus.concentration_level,
                "focus_duration": focus.focus_duration,
                
                # Dữ liệu cảm xúc
                "emotion": focus.emotion,
                "emotion_confidence": focus.emotion_confidence,
                
                # Dữ liệu hành vi
                "behavior_type": focus.behavior_type,
                "behavior_score": focus.behavior_score,
                "behavior_details": focus.behavior_details,
                
                # Dữ liệu điểm danh
                "attendance_status": focus.attendance_status,
                "check_in_time": focus.check_in_time.isoformat() if focus.check_in_time else None,
                
                # Thông tin tính toán
                "comprehensive_score": round(comprehensive_score, 1),
                "engagement_level": engagement_level,
                
                # Metadata
                "date": focus.date,
                "class_name": focus.class_name,
                "session_id": focus.session_id,
                "recorded_by": focus.recorded_by,
                "recorded_at": focus.recorded_at.isoformat() if focus.recorded_at else None
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"❌ Create focus error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi server: {str(e)}")

# ==================== DASHBOARD ENDPOINTS ====================
@app.get("/api/class/dashboard-stats")
async def get_class_dashboard_stats(
    class_name: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Lấy thống kê dashboard với danh sách học sinh cố định - BỎ QUA UNKNOWN"""
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        
        # 1. Lấy tổng số học sinh cố định trong lớp
        query = db.query(ClassStudent).filter(ClassStudent.is_active == True)
        
        if class_name:
            query = query.filter(ClassStudent.class_name == class_name)
        
        total_students = query.count()
        
        if total_students == 0:
            return {
                "status": "success",
                "date": today,
                "summary": {
                    "total_students": 0,
                    "present_count": 0,
                    "absent_count": 0,
                    "attendance_rate": 0,
                    "class_name": class_name or "Chưa có lớp"
                },
                "realtime_stats": {
                    "detected_students": 0,
                    "detection_rate": 0,
                    "avg_focus_score": 0,
                    "avg_behavior_score": 0,
                    "top_emotion": "neutral"
                },
                "students": []
            }
        
        # 2. Lấy danh sách học sinh cố định
        class_students = query.all()
        
        # 3. Lấy điểm danh hôm nay
        attendance_records = db.query(StudentData).filter(
            StudentData.date == today,
            StudentData.attendance_status.isnot(None)
        )
        
        if class_name:
            attendance_records = attendance_records.filter(
                StudentData.class_name == class_name
            )
        
        attendance_data = attendance_records.all()
        
        # 4. Tính số học sinh có mặt (từ bảng student_data)
        present_student_ids = set()
        for record in attendance_data:
            if record.student_id and record.attendance_status in ["present", "late"]:
                present_student_ids.add(record.student_id)
        
        present_count = len(present_student_ids)
        absent_count = max(0, total_students - present_count)
        attendance_rate = round((present_count / total_students) * 100, 1) if total_students > 0 else 0
        
        # 5. Lấy dữ liệu real-time (phát hiện AI trong 30 phút gần nhất)
        recent_threshold = datetime.now() - timedelta(minutes=30)
        
        detected_students = db.query(StudentData).filter(
            StudentData.recorded_at >= recent_threshold,
            StudentData.student_id.isnot(None)
        )
        
        if class_name:
            detected_students = detected_students.filter(
                StudentData.class_name == class_name
            )
        
        detected_data = detected_students.all()
        
        # Filter out unknown students từ detected data
        valid_detected_data = []
        for record in detected_data:
            if not is_unknown_student(record.student_name, record.student_id):
                valid_detected_data.append(record)
        
        # Tính detection rate (tỷ lệ phát hiện)
        detected_student_ids = set([record.student_id for record in valid_detected_data if record.student_id])
        detection_rate = round((len(detected_student_ids) / total_students) * 100, 1) if total_students > 0 else 0
        
        # Tính điểm tập trung trung bình
        focus_scores = [record.focus_score for record in valid_detected_data if record.focus_score is not None]
        avg_focus_score = round(np.mean(focus_scores), 1) if focus_scores else 75.0
        
        # Tính điểm hành vi trung bình
        behavior_scores = [record.behavior_score for record in valid_detected_data if record.behavior_score is not None]
        avg_behavior_score = round(np.mean(behavior_scores), 1) if behavior_scores else 80.0
        
        # Tìm cảm xúc phổ biến nhất
        emotion_counts = {}
        for record in valid_detected_data:
            if record.emotion:
                emotion_counts[record.emotion] = emotion_counts.get(record.emotion, 0) + 1
        
        top_emotion = max(emotion_counts.items(), key=lambda x: x[1])[0] if emotion_counts else "neutral"
        
        # 6. Chuẩn bị danh sách học sinh chi tiết
        students_detail = []
        for class_student in class_students:
            # Tìm dữ liệu real-time gần nhất của học sinh này
            latest_data = db.query(StudentData).filter(
                StudentData.student_id == class_student.student_id,
                StudentData.recorded_at >= recent_threshold
            ).order_by(StudentData.recorded_at.desc()).first()
            
            # Tìm điểm danh hôm nay
            today_attendance = next(
                (record for record in attendance_data if record.student_id == class_student.student_id),
                None
            )
            
            students_detail.append({
                "student_id": class_student.student_id,
                "student_name": class_student.student_name,
                "student_code": class_student.student_code,
                "class_name": class_student.class_name,
                "is_active": class_student.is_active,
                "attendance_status": today_attendance.attendance_status if today_attendance else "absent",
                "check_in_time": today_attendance.check_in_time.isoformat() if today_attendance and today_attendance.check_in_time else None,
                "latest_emotion": latest_data.emotion if latest_data else None,
                "latest_emotion_confidence": latest_data.emotion_confidence if latest_data else None,
                "latest_focus_score": latest_data.focus_score if latest_data else None,
                "latest_behavior": latest_data.behavior_details if latest_data else None,
                "last_detected": latest_data.recorded_at.isoformat() if latest_data else None
            })
        
        # Sắp xếp: có mặt trước, vắng sau
        students_detail.sort(key=lambda x: (
            x["attendance_status"] == "absent",
            x["student_name"]
        ))
        
        return {
            "status": "success",
            "date": today,
            "class_name": class_name or "Tất cả các lớp",
            "summary": {
                "total_students": total_students,
                "present_count": present_count,
                "absent_count": absent_count,
                "attendance_rate": attendance_rate,
                "detection_rate": detection_rate
            },
            "realtime_stats": {
                "detected_students": len(detected_student_ids),
                "detection_rate": detection_rate,
                "avg_focus_score": avg_focus_score,
                "avg_behavior_score": avg_behavior_score,
                "top_emotion": top_emotion,
                "data_source": "realtime" if len(valid_detected_data) > 0 else "demo"
            },
            "students": students_detail
        }
        
    except Exception as e:
        logger.error(f"Error in get_class_dashboard_stats: {e}")
        
        # Fallback data cho dashboard
        today = datetime.now().strftime("%Y-%m-%d")
        return {
            "status": "success",
            "date": today,
            "class_name": class_name or "Lớp 10A1",
            "summary": {
                "total_students": 25,
                "present_count": 22,
                "absent_count": 3,
                "attendance_rate": 88.0,
                "detection_rate": 80.0
            },
            "realtime_stats": {
                "detected_students": 20,
                "detection_rate": 80.0,
                "avg_focus_score": 78.5,
                "avg_behavior_score": 82.3,
                "top_emotion": "happy",
                "data_source": "fallback"
            },
            "students": [
                {
                    "student_id": "SV001",
                    "student_name": "Nguyễn Văn A",
                    "student_code": "2024001",
                    "class_name": "Lớp 10A1",
                    "is_active": True,
                    "attendance_status": "present",
                    "check_in_time": "07:30",
                    "latest_emotion": "happy",
                    "latest_emotion_confidence": 0.85,
                    "latest_focus_score": 85.0,
                    "latest_behavior": "Tích cực tham gia",
                    "last_detected": datetime.now().isoformat()
                }
                # ... (có thể thêm nhiều học sinh mẫu)
            ]
        }
        
@app.get("/api/dashboard/stats")
async def get_dashboard_stats(
    token: str = Query(None),
    db: Session = Depends(get_db)
):
    """Lấy thống kê dashboard - CẬP NHẬT với student count - BỎ QUA UNKNOWN"""
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        
        # Tổng số bản ghi KHÔNG PHẢI UNKNOWN
        total_students = db.query(func.count(StudentData.id)).filter(
            StudentData.student_name.isnot(None)
        ).scalar() or 0
        
        # Điểm danh hôm nay - BỎ QUA UNKNOWN
        present_today = db.query(func.count(StudentData.id)).filter(
            StudentData.date == today,
            StudentData.attendance_status == "present",
            StudentData.student_name.isnot(None)
        ).scalar() or 0
        
        late_today = db.query(func.count(StudentData.id)).filter(
            StudentData.date == today,
            StudentData.attendance_status == "late",
            StudentData.student_name.isnot(None)
        ).scalar() or 0
        
        absent_today = max(0, total_students - present_today - late_today)
        
        attendance_rate = round(((present_today + late_today) / max(total_students, 1)) * 100, 1)
        
        # Thống kê cảm xúc hôm nay - BỎ QUA UNKNOWN
        today_emotions = db.query(StudentData).filter(
            StudentData.date == today,
            StudentData.emotion.isnot(None),
            StudentData.student_name.isnot(None)
        ).all()
        
        # Filter out unknown students
        valid_emotions = []
        for record in today_emotions:
            if not is_unknown_student(record.student_name, record.student_id):
                valid_emotions.append(record)
        
        emotion_distribution = {}
        for record in valid_emotions:
            emotion = record.emotion
            emotion_distribution[emotion] = emotion_distribution.get(emotion, 0) + 1
        
        top_emotion = max(emotion_distribution.items(), key=lambda x: x[1])[0] if emotion_distribution else "neutral"
        
        # Thống kê độ tập trung trung bình - BỎ QUA UNKNOWN
        recent_focus = db.query(StudentData).filter(
            StudentData.focus_score.isnot(None),
            StudentData.date >= (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d"),
            StudentData.student_name.isnot(None)
        ).all()
        
        # Filter out unknown students
        valid_focus = []
        for record in recent_focus:
            if not is_unknown_student(record.student_name, record.student_id):
                valid_focus.append(record)
        
        avg_focus = 75.0
        if valid_focus:
            focus_scores = [record.focus_score for record in valid_focus if record.focus_score is not None]
            if focus_scores:
                avg_focus = round(np.mean(focus_scores), 1)
        
        return {
            "status": "success",
            "user_type": "admin",
            "stats": {
                "total_students": total_students,  # Từ bảng students
                "attendance_today": {
                    "present": present_today,
                    "absent": absent_today,
                    "late": late_today,
                    "total": total_students,
                    "attendance_rate": attendance_rate
                },
                "avg_focus_score": avg_focus,
                "top_emotion": top_emotion,
                "system_status": "online",
                "last_update": datetime.now().isoformat()
            }
        }
        
    except Exception as e:
        logger.error(f"Error in get_dashboard_stats: {e}")
        # Fallback...

@app.get("/api/dashboard/attendance-chart")
async def get_attendance_chart(
    days: int = Query(7, ge=1, le=30),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu biểu đồ điểm danh - BỎ QUA UNKNOWN"""
    try:
        dates = []
        present_counts = []
        absent_counts = []
        late_counts = []
        
        for i in range(days):
            date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            dates.append(date)
            
            try:
                attendance = db.query(StudentData).filter(
                    StudentData.date == date,
                    StudentData.attendance_status.isnot(None),
                    StudentData.student_name.isnot(None)
                ).all()
                
                # Filter out unknown students
                valid_attendance = []
                for record in attendance:
                    if not is_unknown_student(record.student_name, record.student_id):
                        valid_attendance.append(record)
                
                present = sum(1 for record in valid_attendance if record.attendance_status == "present")
                absent = sum(1 for record in valid_attendance if record.attendance_status == "absent")
                late = sum(1 for record in valid_attendance if record.attendance_status == "late")
                
                present_counts.append(present)
                absent_counts.append(absent)
                late_counts.append(late)
            except:
                present_counts.append(0)
                absent_counts.append(0)
                late_counts.append(0)
        
        dates.reverse()
        present_counts.reverse()
        absent_counts.reverse()
        late_counts.reverse()
        
        return {
            "status": "success",
            "chart_data": {
                "labels": dates,
                "datasets": [
                    {
                        "label": "Có mặt",
                        "data": present_counts,
                        "backgroundColor": "rgba(34, 197, 94, 0.5)",
                        "borderColor": "rgb(34, 197, 94)"
                    },
                    {
                        "label": "Vắng mặt",
                        "data": absent_counts,
                        "backgroundColor": "rgba(239, 68, 68, 0.5)",
                        "borderColor": "rgb(239, 68, 68)"
                    },
                    {
                        "label": "Muộn",
                        "data": late_counts,
                        "backgroundColor": "rgba(245, 158, 11, 0.5)",
                        "borderColor": "rgb(245, 158, 11)"
                    }
                ]
            }
        }
        
    except Exception as e:
        logger.error(f"Error in get_attendance_chart: {e}")
        dates = [(datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
        dates.reverse()
        
        return {
            "status": "success",
            "chart_data": {
                "labels": dates,
                "datasets": [
                    {
                        "label": "Có mặt",
                        "data": [10, 12, 9, 11, 13, 10, 12],
                        "backgroundColor": "rgba(34, 197, 94, 0.5)",
                        "borderColor": "rgb(34, 197, 94)"
                    },
                    {
                        "label": "Vắng mặt",
                        "data": [2, 1, 3, 1, 0, 2, 1],
                        "backgroundColor": "rgba(239, 68, 68, 0.5)",
                        "borderColor": "rgb(239, 68, 68)"
                    },
                    {
                        "label": "Muộn",
                        "data": [1, 0, 2, 1, 0, 1, 0],
                        "backgroundColor": "rgba(245, 158, 11, 0.5)",
                        "borderColor": "rgb(245, 158, 11)"
                    }
                ]
            }
        }

@app.get("/api/dashboard/emotion-chart")
async def get_emotion_chart(
    days: int = Query(7, ge=1, le=30),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu biểu đồ cảm xúc - BỎ QUA UNKNOWN"""
    try:
        emotion_counts = {}
        total_emotions = 0
        
        # Lấy dữ liệu cảm xúc trong khoảng thời gian
        start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        
        emotion_records = db.query(StudentData).filter(
            StudentData.emotion.isnot(None),
            StudentData.date >= start_date,
            StudentData.student_name.isnot(None)
        ).all()
        
        # Filter out unknown students
        valid_records = []
        for record in emotion_records:
            if not is_unknown_student(record.student_name, record.student_id):
                valid_records.append(record)
        
        # Thống kê cảm xúc từ dữ liệu hợp lệ
        for record in valid_records:
            emotion = record.emotion
            if emotion:
                emotion_counts[emotion] = emotion_counts.get(emotion, 0) + 1
                total_emotions += 1
        
        # Chuẩn bị dữ liệu biểu đồ
        emotion_labels = list(emotion_counts.keys())
        emotion_data = list(emotion_counts.values())
        
        # Màu sắc cho từng cảm xúc
        emotion_colors = {
            "happy": "rgba(34, 197, 94, 0.7)",
            "neutral": "rgba(156, 163, 175, 0.7)",
            "sad": "rgba(59, 130, 246, 0.7)",
            "angry": "rgba(239, 68, 68, 0.7)",
            "surprised": "rgba(245, 158, 11, 0.7)",
            "fearful": "rgba(168, 85, 247, 0.7)",
            "disgusted": "rgba(20, 184, 166, 0.7)"
        }
        
        colors = [emotion_colors.get(emotion, "rgba(100, 100, 100, 0.7)") for emotion in emotion_labels]
        
        return {
            "status": "success",
            "chart_data": {
                "labels": emotion_labels,
                "datasets": [{
                    "data": emotion_data,
                    "backgroundColor": colors,
                    "borderColor": [color.replace('0.7', '1') for color in colors],
                    "borderWidth": 1
                }]
            }
        }
        
    except Exception as e:
        logger.error(f"Error in get_emotion_chart: {e}")
        return {
            "status": "success",
            "chart_data": {
                "labels": ["happy", "neutral", "sad", "surprised"],
                "datasets": [{
                    "data": [45, 30, 15, 10],
                    "backgroundColor": [
                        "rgba(34, 197, 94, 0.7)",
                        "rgba(156, 163, 175, 0.7)",
                        "rgba(59, 130, 246, 0.7)",
                        "rgba(245, 158, 11, 0.7)"
                    ],
                    "borderColor": [
                        "rgb(34, 197, 94)",
                        "rgb(156, 163, 175)",
                        "rgb(59, 130, 246)",
                        "rgb(245, 158, 11)"
                    ],
                    "borderWidth": 1
                }]
            }
        }

@app.get("/api/dashboard/focus-chart")
async def get_focus_chart(
    days: int = Query(7, ge=1, le=30),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu biểu đồ độ tập trung - BỎ QUA UNKNOWN"""
    try:
        dates = []
        avg_focus_scores = []
        
        for i in range(days):
            date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            dates.append(date)
            
            try:
                focus_records = db.query(StudentData).filter(
                    StudentData.date == date,
                    StudentData.focus_score.isnot(None),
                    StudentData.student_name.isnot(None)
                ).all()
                
                # Filter out unknown students
                valid_records = []
                for record in focus_records:
                    if not is_unknown_student(record.student_name, record.student_id):
                        valid_records.append(record)
                
                if valid_records:
                    focus_scores = [record.focus_score for record in valid_records if record.focus_score is not None]
                    avg_focus = round(np.mean(focus_scores), 1) if focus_scores else 0
                    avg_focus_scores.append(avg_focus)
                else:
                    avg_focus_scores.append(0)
            except:
                avg_focus_scores.append(0)
        
        dates.reverse()
        avg_focus_scores.reverse()
        
        return {
            "status": "success",
            "chart_data": {
                "labels": dates,
                "datasets": [{
                    "label": "Độ tập trung trung bình",
                    "data": avg_focus_scores,
                    "borderColor": "rgb(59, 130, 246)",
                    "backgroundColor": "rgba(59, 130, 246, 0.1)",
                    "tension": 0.4,
                    "fill": True
                }]
            }
        }
        
    except Exception as e:
        logger.error(f"Error in get_focus_chart: {e}")
        dates = [(datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
        dates.reverse()
        
        return {
            "status": "success",
            "chart_data": {
                "labels": dates,
                "datasets": [{
                    "label": "Độ tập trung trung bình",
                    "data": [75, 78, 72, 80, 76, 82, 79],
                    "borderColor": "rgb(59, 130, 246)",
                    "backgroundColor": "rgba(59, 130, 246, 0.1)",
                    "tension": 0.4,
                    "fill": True
                }]
            }
        }

# ==================== ANALYTICS ENDPOINTS ====================
@app.get("/api/analytics/engagement")
async def get_engagement_analytics(
    days: int = Query(7, ge=1, le=30),
    student_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu phân tích độ tập trung (engagement) - Alias cho /api/analytics/focus-analytics"""
    # Redirect đến endpoint focus-analytics vì engagement = focus trong hệ thống mới
    return await get_focus_analytics(days, student_id, db)
    
@app.get("/api/analytics/emotion-trend")
async def get_emotion_trend(
    days: int = Query(7, ge=1, le=30),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu xu hướng cảm xúc từ REAL DATA - BỎ QUA UNKNOWN"""
    try:
        emotion_data = []
        emotion_counts = {}
        total_emotions = 0
        
        # Lấy dữ liệu cảm xúc trong khoảng thời gian
        start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        
        # 🔴 THAY ĐỔI: Lấy cả dữ liệu gần đây nhất (last 30 phút)
        recent_emotions = db.query(StudentData).filter(
            StudentData.emotion.isnot(None),
            StudentData.recorded_at >= (datetime.now() - timedelta(minutes=30)),
            StudentData.student_name.isnot(None)
        ).all()
        
        # Filter out unknown students
        valid_emotions = []
        for record in recent_emotions:
            if not is_unknown_student(record.student_name, record.student_id):
                valid_emotions.append(record)
        
        # Thống kê cảm xúc từ dữ liệu REAL-TIME
        for record in valid_emotions:
            emotion = record.emotion
            if emotion:
                emotion_counts[emotion] = emotion_counts.get(emotion, 0) + 1
                total_emotions += 1
        
        # Nếu có dữ liệu real-time, ưu tiên dùng
        if emotion_counts:
            # Tính tỷ lệ phần trăm
            emotion_percentages = []
            for emotion, count in emotion_counts.items():
                percentage = round((count / max(total_emotions, 1)) * 100, 1)
                emotion_percentages.append({
                    "emotion": emotion,
                    "count": count,
                    "percentage": percentage
                })
            
            # Sắp xếp theo count giảm dần
            emotion_percentages.sort(key=lambda x: x["count"], reverse=True)
            
            return {
                "status": "success",
                "summary": {
                    "total_emotions": total_emotions,
                    "top_emotion": emotion_percentages[0]["emotion"] if emotion_percentages else "neutral",
                    "top_emotion_percentage": emotion_percentages[0]["percentage"] if emotion_percentages else 0,
                    "data_source": "realtime"  # 🔴 THÊM: Đánh dấu là real-time data
                },
                "emotion_distribution": emotion_percentages,
                "days": days
            }
        
        # Fallback: dùng dữ liệu demo
        return {
            "status": "success",
            "summary": {
                "total_emotions": 45,
                "top_emotion": "happy",
                "top_emotion_percentage": 45.3,
                "data_source": "demo"  # 🔴 THÊM: Đánh dấu là demo data
            },
            "emotion_distribution": [
                {"emotion": "happy", "count": 20, "percentage": 45.3},
                {"emotion": "neutral", "count": 12, "percentage": 28.0},
                {"emotion": "sad", "count": 8, "percentage": 16.7},
                {"emotion": "surprised", "count": 3, "percentage": 6.7},
                {"emotion": "angry", "count": 2, "percentage": 3.3}
            ],
            "days": days
        }
        
    except Exception as e:
        logger.error(f"Error in get_emotion_trend: {e}")
        return {
            "status": "success",
            "summary": {
                "total_emotions": 45,
                "top_emotion": "happy",
                "top_emotion_percentage": 45.3,
                "data_source": "fallback"
            },
            "emotion_distribution": [
                {"emotion": "happy", "count": 20, "percentage": 45.3},
                {"emotion": "neutral", "count": 12, "percentage": 28.0},
                {"emotion": "sad", "count": 8, "percentage": 16.7},
                {"emotion": "surprised", "count": 3, "percentage": 6.7}
            ],
            "days": 7
        }

@app.get("/api/analytics/focus-analytics")
async def get_focus_analytics(
    days: int = Query(7, ge=1, le=30),
    student_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu phân tích độ tập trung từ REAL DATA - BỎ QUA UNKNOWN"""
    try:
        # 🔴 ƯU TIÊN: Lấy dữ liệu từ 30 phút gần nhất
        recent_threshold = datetime.now() - timedelta(minutes=30)
        
        # Query REAL-TIME focus data
        query = db.query(StudentData).filter(
            StudentData.recorded_at >= recent_threshold,
            StudentData.focus_score.isnot(None),
            StudentData.student_name.isnot(None)
        )
        
        if student_id:
            query = query.filter(StudentData.student_id == student_id)
        
        focus_records = query.order_by(StudentData.recorded_at.desc()).all()
        
        # Filter out unknown students
        valid_records = []
        for record in focus_records:
            if not is_unknown_student(record.student_name, record.student_id):
                valid_records.append(record)
        
        if valid_records:
            # Tính toán thống kê từ REAL DATA
            focus_scores = [record.focus_score for record in valid_records if record.focus_score is not None]
            avg_focus = np.mean(focus_scores) if focus_scores else 75.0
            max_focus = max(focus_scores) if focus_scores else 90.0
            min_focus = min(focus_scores) if focus_scores else 60.0
            
            # Phân loại focus levels từ real data
            excellent = len([s for s in focus_scores if s >= 85])
            good = len([s for s in focus_scores if 70 <= s < 85])
            average = len([s for s in focus_scores if 55 <= s < 70])
            poor = len([s for s in focus_scores if s < 55])
            
            # Top students từ real data
            top_students_dict = {}
            for record in valid_records:
                if record.student_name and record.focus_score:
                    if record.student_name not in top_students_dict:
                        top_students_dict[record.student_name] = {
                            'scores': [],
                            'count': 0
                        }
                    top_students_dict[record.student_name]['scores'].append(record.focus_score)
                    top_students_dict[record.student_name]['count'] += 1
            
            # Tính average cho mỗi student
            top_students_list = []
            for name, data in top_students_dict.items():
                if data['scores']:
                    avg_score = np.mean(data['scores'])
                    top_students_list.append({
                        "name": name,
                        "avg_focus": round(avg_score, 1),
                        "records": data['count']
                    })
            
            # Sắp xếp theo avg_focus giảm dần
            top_students_list.sort(key=lambda x: x['avg_focus'], reverse=True)
            
            # Dữ liệu biểu đồ
            dates = []
            daily_avg_scores = []
            
            # Lấy dữ liệu 7 ngày gần nhất
            for i in range(min(7, days)):
                date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
                dates.append(date)
                
                # Tính avg focus cho ngày đó
                day_records = [r for r in valid_records if r.date == date]
                if day_records:
                    day_scores = [r.focus_score for r in day_records if r.focus_score is not None]
                    if day_scores:
                        daily_avg_scores.append(round(np.mean(day_scores), 1))
                    else:
                        daily_avg_scores.append(0)
                else:
                    daily_avg_scores.append(0)
            
            dates.reverse()
            daily_avg_scores.reverse()
            
            return {
                "status": "success",
                "summary": {
                    "total_records": len(valid_records),
                    "avg_focus": round(avg_focus, 1),
                    "max_focus": round(max_focus, 1),
                    "min_focus": round(min_focus, 1),
                    "trend": "improving" if len(daily_avg_scores) > 1 and daily_avg_scores[-1] > daily_avg_scores[0] else "stable",
                    "data_source": "realtime"
                },
                "focus_levels": {
                    "excellent": excellent,
                    "good": good,
                    "average": average,
                    "poor": poor
                },
                "chart_data": {
                    "labels": dates,
                    "datasets": [
                        {
                            "label": "Độ tập trung trung bình",
                            "data": daily_avg_scores,
                            "borderColor": "rgb(59, 130, 246)",
                            "backgroundColor": "rgba(59, 130, 246, 0.1)",
                            "tension": 0.4,
                            "fill": True
                        }
                    ]
                },
                "top_students": top_students_list[:5],  # Top 5 students
                "days": days
            }
        
        # Fallback: dùng demo data
        return get_fallback_focus_data(days)
        
    except Exception as e:
        logger.error(f"Error in get_focus_analytics: {e}")
        return get_fallback_focus_data(days)

def get_fallback_focus_data(days: int):
    """Fallback data cho focus analytics"""
    dates = [(datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(days)][::-1]
    
    # Tạo dữ liệu mẫu có xu hướng tăng
    base_scores = [65, 68, 70, 72, 75, 78, 80, 82, 85, 88]
    if days > len(base_scores):
        base_scores = base_scores * (days // len(base_scores) + 1)
    daily_scores = base_scores[:days]
    
    return {
        "status": "success",
        "summary": {
            "total_records": days * 8,
            "avg_focus": 75.5,
            "max_focus": 92.0,
            "min_focus": 58.0,
            "trend": "improving"
        },
        "focus_levels": {
            "excellent": days * 2,
            "good": days * 3,
            "average": days * 2,
            "poor": days * 1
        },
        "chart_data": {
            "labels": dates,
            "datasets": [
                {
                    "label": "Độ tập trung trung bình",
                    "data": daily_scores,
                    "borderColor": "rgb(59, 130, 246)",
                    "backgroundColor": "rgba(59, 130, 246, 0.1)",
                    "tension": 0.4,
                    "fill": True
                }
            ]
        },
        "top_students": [
            {"name": "Nguyễn Văn A", "avg_focus": 92.5, "records": 15},
            {"name": "Trần Thị B", "avg_focus": 88.3, "records": 12},
            {"name": "Lê Văn C", "avg_focus": 85.7, "records": 10},
            {"name": "Phạm Thị D", "avg_focus": 82.1, "records": 8},
            {"name": "Hoàng Văn E", "avg_focus": 79.8, "records": 7}
        ],
        "days": days
    }

# ==================== REPORT ENDPOINTS ====================
@app.get("/api/reports")
async def get_reports():
    """Lấy danh sách báo cáo"""
    return {
        "reports": [
            {
                "id": 1,
                "name": "Báo cáo điểm danh hàng ngày",
                "date": datetime.now().date().isoformat(),
                "type": "PDF",
                "size": "245 KB",
                "icon": "📊"
            },
            {
                "id": 2,
                "name": "Báo cáo cảm xúc tuần",
                "date": (datetime.now() - timedelta(days=7)).date().isoformat(),
                "type": "CSV",
                "size": "128 KB",
                "icon": "😊"
            },
            {
                "id": 3,
                "name": "Báo cáo độ tập trung",
                "date": datetime.now().date().isoformat(),
                "type": "PDF",
                "size": "312 KB",
                "icon": "🎯"
            }
        ]
    }

@app.get("/api/reports/export")
async def export_reports(
    report_type: str = "attendance",
    date: Optional[str] = None,
    include_analytics: bool = Query(True, description="Bao gồm sheet phân tích chi tiết"),
    db: Session = Depends(get_db)
):
    """Xuất báo cáo chi tiết - Sheet 1: Điểm danh, Sheet 2: Phân tích cảm xúc, hành vi, độ tập trung"""
    try:
        logger.info(f"🔄 Starting report export: type={report_type}, date={date}")
        
        if not date:
            date = datetime.now().strftime("%Y-%m-%d")
            logger.info(f"📅 Using current date: {date}")
        
        if report_type == "attendance":
            logger.info(f"🔍 Querying attendance data for date: {date}")
            
            try:
                # Lấy danh sách học sinh DUY NHẤT đã điểm danh trong ngày
                subquery = db.query(
                    StudentData.student_id,
                    StudentData.student_name,
                    func.max(StudentData.recorded_at).label('latest_record')
                ).filter(
                    StudentData.date == date,
                    StudentData.attendance_status.isnot(None),
                    StudentData.student_name.isnot(None)
                ).group_by(
                    StudentData.student_id,
                    StudentData.student_name
                ).subquery()
                
                logger.info(f"📊 Subquery created successfully")
                
                # Lấy dữ liệu đầy đủ từ bản ghi mới nhất của mỗi học sinh
                attendance_records = db.query(StudentData).join(
                    subquery,
                    (StudentData.student_id == subquery.c.student_id) &
                    (StudentData.recorded_at == subquery.c.latest_record)
                ).filter(
                    StudentData.date == date,
                    StudentData.attendance_status.isnot(None),
                    StudentData.student_name.isnot(None)
                ).order_by(
                    StudentData.student_name.asc()
                ).all()
                
                logger.info(f"📋 Found {len(attendance_records)} attendance records")
                
                # Filter out unknown students
                valid_records = []
                unknown_count = 0
                for record in attendance_records:
                    if not is_unknown_student(record.student_name, record.student_id):
                        valid_records.append(record)
                    else:
                        unknown_count += 1
                
                logger.info(f"✅ After filtering: {len(valid_records)} valid records, {unknown_count} unknown filtered")
                
                # Chuẩn bị dữ liệu cho sheet 1 (điểm danh)
                attendance_data = []
                
                # Tạo dict để lưu chi tiết từng học sinh cho sheet 2
                student_analytics = {}
                
                for idx, record in enumerate(valid_records):
                    try:
                        # Lấy thông tin học sinh từ bảng ClassStudent nếu có
                        class_student = None
                        if record.student_id:
                            class_student = db.query(ClassStudent).filter(
                                ClassStudent.student_id == record.student_id
                            ).first()
                        
                        # LẤY TẤT CẢ FEEDBACK CỦA HỌC SINH
                        student_feedbacks = []
                        if record.student_id:
                            feedbacks = db.query(StudentFeedback).filter(
                                StudentFeedback.student_id == record.student_id,
                                StudentFeedback.created_at >= datetime.strptime(date, "%Y-%m-%d")
                            ).order_by(StudentFeedback.created_at.desc()).limit(10).all()
                            
                            for fb in feedbacks:
                                feedback_text = ""
                                if fb.feedback_type == "text":
                                    feedback_text = fb.feedback_text or ""
                                elif fb.feedback_type == "voice":
                                    feedback_text = fb.transcribed_text or fb.feedback_text or ""
                                
                                if feedback_text:
                                    fb_time = fb.created_at.strftime("%H:%M") if fb.created_at else ""
                                    fb_rating = f"{fb.rating}⭐" if fb.rating else ""
                                    fb_emotion = fb.emotion or ""
                                    
                                    feedback_display = f"[{fb_time}]"
                                    if fb_rating:
                                        feedback_display += f" {fb_rating}"
                                    if fb_emotion:
                                        feedback_display += f" ({fb_emotion})"
                                    feedback_display += f": {feedback_text}"
                                    
                                    student_feedbacks.append(feedback_display)
                        
                        # Gộp tất cả feedback
                        all_feedbacks = " | ".join(student_feedbacks) if student_feedbacks else "Không có feedback"
                        
                        # Xác định trạng thái điểm danh với biểu tượng
                        status_display = record.attendance_status or "unknown"
                        if record.attendance_status == "present":
                            status_display = "✅ Có mặt"
                        elif record.attendance_status == "absent":
                            status_display = "❌ Vắng mặt"
                        elif record.attendance_status == "late":
                            status_display = "⚠️ Muộn"
                        
                        # Xác định mức độ tập trung
                        focus_display = "Không có dữ liệu"
                        if record.focus_score is not None:
                            try:
                                focus_score = float(record.focus_score)
                                if focus_score >= 85:
                                    focus_display = f"Rất cao ({focus_score:.1f})"
                                elif focus_score >= 70:
                                    focus_display = f"Cao ({focus_score:.1f})"
                                elif focus_score >= 55:
                                    focus_display = f"Trung bình ({focus_score:.1f})"
                                else:
                                    focus_display = f"Thấp ({focus_score:.1f})"
                            except:
                                focus_display = f"{record.focus_score}"
                        
                        # Xác định thời gian điểm danh
                        check_in_time_display = "Không có"
                        if record.check_in_time:
                            try:
                                if isinstance(record.check_in_time, datetime):
                                    check_in_time_display = record.check_in_time.strftime("%H:%M:%S")
                                else:
                                    check_in_time_display = str(record.check_in_time)
                            except:
                                check_in_time_display = str(record.check_in_time)
                        
                        # Xác định lớp học
                        class_name_display = "Chưa xác định"
                        if class_student and class_student.class_name:
                            class_name_display = class_student.class_name
                        elif record.class_name:
                            class_name_display = record.class_name
                        
                        # Thêm vào attendance data
                        attendance_data.append({
                            "STT": idx + 1,
                            "Mã học sinh": record.student_id or "Chưa có",
                            "Họ và tên": record.student_name or "Chưa có tên",
                            "Lớp": class_name_display,
                            "Điểm danh": status_display,
                            "Thời gian vào lớp": check_in_time_display,
                            "Cảm xúc": record.emotion or "Không có",
                            "Hành vi": record.behavior_details or "Không có",
                            "Độ tập trung": focus_display,
                            "Mức tập trung": record.concentration_level or "Không có",
                            "Ghi chú": record.attendance_notes or "",
                            "Feedback": all_feedbacks,
                        })
                        
                        # ==================== CHUẨN BỊ DỮ LIỆU CHO SHEET 2 ====================
                        if record.student_id:
                            # Lấy TẤT CẢ dữ liệu của học sinh này trong ngày
                            daily_records = db.query(StudentData).filter(
                                StudentData.student_id == record.student_id,
                                StudentData.date == date,
                                StudentData.recorded_at.isnot(None)
                            ).order_by(StudentData.recorded_at.asc()).all()
                            
                            if daily_records:
                                # Phân tích cảm xúc
                                emotions = []
                                emotion_timestamps = []
                                emotion_confidence_scores = []
                                
                                # Phân tích hành vi
                                behaviors = []
                                behavior_timestamps = []
                                behavior_scores = []
                                behavior_types = []
                                
                                # Phân tích độ tập trung
                                focus_scores = []
                                focus_timestamps = []
                                concentration_levels = []
                                
                                # Lịch sử thay đổi
                                timeline = []
                                
                                for dr in daily_records:
                                    timestamp = dr.recorded_at.strftime("%H:%M:%S") if dr.recorded_at else "N/A"
                                    
                                    # Thu thập dữ liệu cảm xúc
                                    if dr.emotion:
                                        emotions.append(dr.emotion)
                                        emotion_timestamps.append(timestamp)
                                        emotion_confidence_scores.append(float(dr.emotion_confidence or 0.5))
                                    
                                    # Thu thập dữ liệu hành vi
                                    if dr.behavior_details or dr.behavior_type:
                                        behavior_text = f"{dr.behavior_type or 'N/A'}: {dr.behavior_details or ''}"
                                        behaviors.append(behavior_text)
                                        behavior_timestamps.append(timestamp)
                                        behavior_scores.append(float(dr.behavior_score or 0))
                                        behavior_types.append(dr.behavior_type or 'N/A')
                                    
                                    # Thu thập dữ liệu độ tập trung
                                    if dr.focus_score is not None:
                                        focus_scores.append(float(dr.focus_score))
                                        focus_timestamps.append(timestamp)
                                        concentration_levels.append(dr.concentration_level or 'N/A')
                                    
                                    # Ghi nhận timeline
                                    timeline_entry = f"{timestamp}: "
                                    if dr.emotion:
                                        timeline_entry += f"Cảm xúc: {dr.emotion} ({dr.emotion_confidence or 0:.0%})"
                                    if dr.behavior_details:
                                        timeline_entry += f", Hành vi: {dr.behavior_details}"
                                    if dr.focus_score:
                                        timeline_entry += f", Tập trung: {dr.focus_score:.1f}"
                                    timeline.append(timeline_entry)
                                
                                # Tính toán thống kê
                                emotion_summary = {}
                                for emotion in emotions:
                                    emotion_summary[emotion] = emotion_summary.get(emotion, 0) + 1
                                
                                behavior_summary = {}
                                for behavior in behavior_types:
                                    if behavior:
                                        behavior_summary[behavior] = behavior_summary.get(behavior, 0) + 1
                                
                                # Tính điểm trung bình
                                avg_emotion_confidence = np.mean(emotion_confidence_scores) if emotion_confidence_scores else 0
                                avg_behavior_score = np.mean(behavior_scores) if behavior_scores else 0
                                avg_focus_score = np.mean(focus_scores) if focus_scores else 0
                                
                                # Tìm cảm xúc, hành vi phổ biến nhất
                                dominant_emotion = max(emotion_summary.items(), key=lambda x: x[1])[0] if emotion_summary else "Không có"
                                dominant_behavior = max(behavior_summary.items(), key=lambda x: x[1])[0] if behavior_summary else "Không có"
                                
                                # Tính % thời gian tập trung cao/trung bình/thấp
                                concentration_counts = {}
                                for level in concentration_levels:
                                    concentration_counts[level] = concentration_counts.get(level, 0) + 1
                                
                                # Độ dao động tập trung
                                focus_variation = np.std(focus_scores) if len(focus_scores) > 1 else 0
                                
                                # Lưu vào student_analytics
                                student_analytics[record.student_id] = {
                                    "student_name": record.student_name,
                                    "student_id": record.student_id,
                                    "class_name": class_name_display,
                                    "total_records": len(daily_records),
                                    "emotion_summary": emotion_summary,
                                    "behavior_summary": behavior_summary,
                                    "concentration_summary": concentration_counts,
                                    "avg_emotion_confidence": round(avg_emotion_confidence, 2),
                                    "avg_behavior_score": round(avg_behavior_score, 1),
                                    "avg_focus_score": round(avg_focus_score, 1),
                                    "dominant_emotion": dominant_emotion,
                                    "dominant_behavior": dominant_behavior,
                                    "focus_variation": round(focus_variation, 2),
                                    "timeline": timeline,
                                    "raw_data": {
                                        "emotions": emotions,
                                        "emotion_timestamps": emotion_timestamps,
                                        "behavior_types": behavior_types,
                                        "behavior_timestamps": behavior_timestamps,
                                        "focus_scores": focus_scores,
                                        "focus_timestamps": focus_timestamps
                                    }
                                }
                                
                    except Exception as record_error:
                        logger.error(f"❌ Error processing record {idx}: {record_error}")
                        logger.error(f"Record data: student_id={record.student_id}, student_name={record.student_name}")
                        continue
                
                logger.info(f"📝 Prepared {len(attendance_data)} rows for attendance sheet")
                logger.info(f"📊 Prepared analytics for {len(student_analytics)} students")
                
                # Nếu không có dữ liệu
                if not attendance_data:
                    logger.warning("⚠️ No valid data found for export")
                    attendance_data.append({
                        "STT": 1,
                        "Mã học sinh": "",
                        "Họ và tên": "Không có dữ liệu điểm danh",
                        "Lớp": "",
                        "Điểm danh": "",
                        "Thời gian vào lớp": "",
                        "Cảm xúc": "",
                        "Hành vi": "",
                        "Độ tập trung": "",
                        "Mức tập trung": "",
                        "Ghi chú": f"Không có dữ liệu điểm danh ngày {date}",
                        "Feedback": "",
                    })
                
                filename = f"bao_cao_diem_danh_{date}.xlsx"
                logger.info(f"💾 Creating Excel file: {filename}")
                
                try:
                    # Tạo Excel file trong memory
                    stream = io.BytesIO()
                    
                    # Sử dụng engine mặc định
                    with pd.ExcelWriter(stream, engine='openpyxl') as writer:
                        workbook = writer.book
                        
                        # ==================== SHEET 1: ĐIỂM DANH ====================
                        df_attendance = pd.DataFrame(attendance_data)
                        df_attendance.to_excel(writer, index=False, sheet_name='ĐiểmDanh')
                        
                        # Định dạng sheet 1
                        ws_attendance = writer.sheets['ĐiểmDanh']
                        
                        # Đặt độ rộng cột
                        column_widths_sheet1 = {
                            "A": 6,    # STT
                            "B": 12,   # Mã học sinh
                            "C": 25,   # Họ và tên
                            "D": 10,   # Lớp
                            "E": 15,   # Điểm danh
                            "F": 15,   # Thời gian vào lớp
                            "G": 15,   # Cảm xúc
                            "H": 25,   # Hành vi
                            "I": 15,   # Độ tập trung
                            "J": 15,   # Mức tập trung
                            "K": 30,   # Ghi chú
                            "L": 50,   # Feedback
                        }
                        
                        for col, width in column_widths_sheet1.items():
                            ws_attendance.column_dimensions[col].width = width
                        
                        # Style cho header sheet 1
                        from openpyxl.styles import (Alignment, Border, Font,
                                                     PatternFill, Side)
                        
                        header_font = Font(bold=True, color="FFFFFF", size=11)
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        border = Border(left=Side(style='thin'), 
                                       right=Side(style='thin'), 
                                       top=Side(style='thin'), 
                                       bottom=Side(style='thin'))
                        
                        # Áp dụng style cho header
                        for col_num in range(1, len(df_attendance.columns) + 1):
                            cell = ws_attendance.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = border
                        
                        # Đặt wrap text cho các cột dài
                        for row in range(2, len(df_attendance) + 2):
                            for col in ['H', 'K', 'L']:  # Hành vi, Ghi chú, Feedback
                                col_idx = ord(col) - 64
                                cell = ws_attendance.cell(row=row, column=col_idx)
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                        
                        # Thêm màu cho các trạng thái
                        for row in range(2, len(df_attendance) + 2):
                            status_cell = ws_attendance.cell(row=row, column=5)  # Cột E - Điểm danh
                            if "✅" in str(status_cell.value):
                                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif "❌" in str(status_cell.value):
                                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            elif "⚠️" in str(status_cell.value):
                                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        
                        logger.info("✅ Sheet 1 (ĐiểmDanh) created and formatted")
                        
                        # ==================== SHEET 2: PHÂN TÍCH CHI TIẾT ====================
                        if include_analytics and student_analytics:
                            # Chuẩn bị dữ liệu cho sheet 2
                            analytics_data = []
                            
                            for student_id, analytics in student_analytics.items():
                                # Chuyển dict thành string dễ đọc
                                emotion_summary_str = ", ".join([f"{k}: {v} lần" for k, v in analytics['emotion_summary'].items()]) if analytics['emotion_summary'] else "Không có"
                                behavior_summary_str = ", ".join([f"{k}: {v} lần" for k, v in analytics['behavior_summary'].items()]) if analytics['behavior_summary'] else "Không có"
                                concentration_summary_str = ", ".join([f"{k}: {v} lần" for k, v in analytics['concentration_summary'].items()]) if analytics['concentration_summary'] else "Không có"
                                
                                # Lấy 5 dòng timeline đầu tiên
                                timeline_preview = "\n".join(analytics['timeline'][:5]) if analytics['timeline'] else "Không có"
                                if len(analytics['timeline']) > 5:
                                    timeline_preview += f"\n... và {len(analytics['timeline']) - 5} sự kiện khác"
                                
                                # Đánh giá tổng thể
                                overall_assessment = ""
                                if analytics['avg_focus_score'] >= 80:
                                    overall_assessment = "Rất tập trung"
                                elif analytics['avg_focus_score'] >= 60:
                                    overall_assessment = "Tập trung tốt"
                                elif analytics['avg_focus_score'] >= 40:
                                    overall_assessment = "Tập trung trung bình"
                                else:
                                    overall_assessment = "Cần cải thiện sự tập trung"
                                
                                if analytics['dominant_emotion'] == 'happy':
                                    overall_assessment += ", Tâm trạng tích cực"
                                elif analytics['dominant_emotion'] in ['sad', 'angry']:
                                    overall_assessment += ", Cần quan tâm đến cảm xúc"
                                
                                analytics_data.append({
                                    "STT": len(analytics_data) + 1,
                                    "Mã học sinh": student_id,
                                    "Họ và tên": analytics['student_name'],
                                    "Lớp": analytics['class_name'],
                                    "Số lần phát hiện": analytics['total_records'],
                                    "Cảm xúc chủ đạo": analytics['dominant_emotion'],
                                    "Thống kê cảm xúc": emotion_summary_str,
                                    "Độ tin cậy cảm xúc TB": f"{analytics['avg_emotion_confidence']:.0%}",
                                    "Hành vi chủ đạo": analytics['dominant_behavior'],
                                    "Thống kê hành vi": behavior_summary_str,
                                    "Điểm hành vi TB": analytics['avg_behavior_score'],
                                    "Điểm tập trung TB": analytics['avg_focus_score'],
                                    "Thống kê mức tập trung": concentration_summary_str,
                                    "Độ dao động tập trung": analytics['focus_variation'],
                                    "Đánh giá tổng thể": overall_assessment,
                                    "Lịch sử hoạt động": timeline_preview
                                })
                            
                            # Tạo sheet 2
                            df_analytics = pd.DataFrame(analytics_data)
                            df_analytics.to_excel(writer, index=False, sheet_name='PhânTíchChiTiết')
                            
                            ws_analytics = writer.sheets['PhânTíchChiTiết']
                            
                            # Đặt độ rộng cột cho sheet 2
                            column_widths_sheet2 = {
                                "A": 6,    # STT
                                "B": 12,   # Mã học sinh
                                "C": 25,   # Họ và tên
                                "D": 10,   # Lớp
                                "E": 12,   # Số lần phát hiện
                                "F": 15,   # Cảm xúc chủ đạo
                                "G": 30,   # Thống kê cảm xúc
                                "H": 18,   # Độ tin cậy cảm xúc TB
                                "I": 20,   # Hành vi chủ đạo
                                "J": 30,   # Thống kê hành vi
                                "K": 15,   # Điểm hành vi TB
                                "L": 15,   # Điểm tập trung TB
                                "M": 25,   # Thống kê mức tập trung
                                "N": 18,   # Độ dao động tập trung
                                "O": 40,   # Đánh giá tổng thể
                                "P": 60,   # Lịch sử hoạt động
                            }
                            
                            for col, width in column_widths_sheet2.items():
                                ws_analytics.column_dimensions[col].width = width
                            
                            # Style cho header sheet 2
                            for col_num in range(1, len(df_analytics.columns) + 1):
                                cell = ws_analytics.cell(row=1, column=col_num)
                                cell.font = Font(bold=True, color="FFFFFF", size=11)
                                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                cell.border = border
                            
                            # Đặt wrap text cho các cột dài
                            for row in range(2, len(df_analytics) + 2):
                                for col in ['G', 'J', 'O', 'P']:  # Các cột có text dài
                                    col_idx = ord(col) - 64
                                    cell = ws_analytics.cell(row=row, column=col_idx)
                                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                            
                            # Thêm màu nền cho điểm số
                            for row in range(2, len(df_analytics) + 2):
                                # Điểm tập trung
                                focus_cell = ws_analytics.cell(row=row, column=12)  # Cột L
                                try:
                                    focus_score = float(focus_cell.value) if focus_cell.value else 0
                                    if focus_score >= 80:
                                        focus_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                    elif focus_score >= 60:
                                        focus_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                    else:
                                        focus_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                except:
                                    pass
                                
                                # Điểm hành vi
                                behavior_cell = ws_analytics.cell(row=row, column=11)  # Cột K
                                try:
                                    behavior_score = float(behavior_cell.value) if behavior_cell.value else 0
                                    if behavior_score >= 80:
                                        behavior_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                    elif behavior_score >= 60:
                                        behavior_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                    else:
                                        behavior_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                except:
                                    pass
                            
                            logger.info("✅ Sheet 2 (PhânTíchChiTiết) created and formatted")
                            
                            # ==================== SHEET 3: BIỂU ĐỒ TỔNG HỢP ====================
                            try:
                                from openpyxl.chart import (BarChart,
                                                            LineChart,
                                                            PieChart,
                                                            Reference, Series)

                                # Tạo sheet mới cho biểu đồ
                                ws_charts = workbook.create_sheet(title="BiểuĐồTổngHợp")
                                
                                # Tiêu đề
                                ws_charts.merge_cells('A1:H1')
                                title_cell = ws_charts.cell(row=1, column=1)
                                title_cell.value = f"BIỂU ĐỒ PHÂN TÍCH LỚP HỌC - NGÀY {date}"
                                title_cell.font = Font(bold=True, size=16, color="366092")
                                title_cell.alignment = Alignment(horizontal="center")
                                
                                # 1. Biểu đồ cảm xúc tổng hợp
                                emotion_chart_data = []
                                for student_id, analytics in student_analytics.items():
                                    for emotion, count in analytics['emotion_summary'].items():
                                        emotion_chart_data.append({
                                            "Học sinh": analytics['student_name'],
                                            "Cảm xúc": emotion,
                                            "Số lần": count
                                        })
                                
                                if emotion_chart_data:
                                    # Ghi dữ liệu cho biểu đồ cảm xúc
                                    ws_charts['A3'] = "Học sinh"
                                    ws_charts['B3'] = "Cảm xúc"
                                    ws_charts['C3'] = "Số lần"
                                    
                                    row = 4
                                    for item in emotion_chart_data[:20]:  # Giới hạn 20 dòng
                                        ws_charts.cell(row=row, column=1, value=item['Học sinh'])
                                        ws_charts.cell(row=row, column=2, value=item['Cảm xúc'])
                                        ws_charts.cell(row=row, column=3, value=item['Số lần'])
                                        row += 1
                                    
                                    # Tạo biểu đồ cột
                                    chart1 = BarChart()
                                    chart1.type = "col"
                                    chart1.style = 10
                                    chart1.title = "Phân bố cảm xúc học sinh"
                                    chart1.y_axis.title = 'Số lần'
                                    chart1.x_axis.title = 'Học sinh'
                                    
                                    data = Reference(ws_charts, min_col=3, min_row=3, max_row=row-1, max_col=3)
                                    cats = Reference(ws_charts, min_col=1, min_row=4, max_row=row-1)
                                    chart1.add_data(data, titles_from_data=True)
                                    chart1.set_categories(cats)
                                    chart1.shape = 4
                                    
                                    ws_charts.add_chart(chart1, "E3")
                                
                                # 2. Biểu đồ điểm tập trung trung bình
                                focus_chart_data = []
                                for student_id, analytics in student_analytics.items():
                                    focus_chart_data.append({
                                        "Học sinh": analytics['student_name'],
                                        "Điểm tập trung TB": analytics['avg_focus_score']
                                    })
                                
                                if focus_chart_data:
                                    # Ghi dữ liệu cho biểu đồ điểm tập trung
                                    start_row = row + 5
                                    ws_charts.cell(row=start_row, column=1, value="Học sinh")
                                    ws_charts.cell(row=start_row, column=2, value="Điểm tập trung TB")
                                    
                                    chart_row = start_row + 1
                                    for item in focus_chart_data:
                                        ws_charts.cell(row=chart_row, column=1, value=item['Học sinh'])
                                        ws_charts.cell(row=chart_row, column=2, value=item['Điểm tập trung TB'])
                                        chart_row += 1
                                    
                                    # Tạo biểu đồ đường
                                    chart2 = LineChart()
                                    chart2.title = "Điểm tập trung trung bình"
                                    chart2.style = 12
                                    chart2.y_axis.title = "Điểm"
                                    chart2.x_axis.title = "Học sinh"
                                    
                                    data = Reference(ws_charts, min_col=2, min_row=start_row, max_row=chart_row-1)
                                    cats = Reference(ws_charts, min_col=1, min_row=start_row+1, max_row=chart_row-1)
                                    chart2.add_data(data, titles_from_data=True)
                                    chart2.set_categories(cats)
                                    chart2.shape = 4
                                    
                                    ws_charts.add_chart(chart2, "E20")
                                
                                # 3. Thống kê tổng hợp
                                summary_row = chart_row + 5
                                ws_charts.cell(row=summary_row, column=1, value="THỐNG KÊ TỔNG HỢP").font = Font(bold=True, size=12)
                                
                                summary_row += 2
                                metrics = [
                                    ("Tổng số học sinh", len(student_analytics)),
                                    ("Điểm tập trung TB lớp", round(np.mean([a['avg_focus_score'] for a in student_analytics.values()]), 1)),
                                    ("Điểm hành vi TB lớp", round(np.mean([a['avg_behavior_score'] for a in student_analytics.values()]), 1)),
                                    ("Cảm xúc phổ biến nhất", max(
                                        [(emotion, sum(a['emotion_summary'].get(emotion, 0) for a in student_analytics.values())) 
                                         for emotion in set().union(*[a['emotion_summary'].keys() for a in student_analytics.values()])],
                                        key=lambda x: x[1]
                                    )[0] if student_analytics else "N/A"),
                                    ("Học sinh tập trung nhất", max(student_analytics.items(), key=lambda x: x[1]['avg_focus_score'])[1]['student_name'] if student_analytics else "N/A"),
                                    ("Học sinh cần quan tâm", min(student_analytics.items(), key=lambda x: x[1]['avg_focus_score'])[1]['student_name'] if student_analytics else "N/A")
                                ]
                                
                                for i, (label, value) in enumerate(metrics):
                                    ws_charts.cell(row=summary_row + i, column=1, value=label)
                                    ws_charts.cell(row=summary_row + i, column=2, value=value)
                                    if "điểm" in label.lower() or "tb" in label.lower():
                                        try:
                                            num_value = float(value)
                                            if num_value >= 80:
                                                ws_charts.cell(row=summary_row + i, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                            elif num_value >= 60:
                                                ws_charts.cell(row=summary_row + i, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                        except:
                                            pass
                                
                                logger.info("✅ Sheet 3 (BiểuĐồTổngHợp) created with charts")
                                
                            except Exception as chart_error:
                                logger.warning(f"⚠️ Could not create charts sheet: {chart_error}")
                                # Vẫn tiếp tục nếu không tạo được biểu đồ
                        
                        # ==================== SHEET 4: THỐNG KÊ ĐƠN GIẢN ====================
                        try:
                            ws_stats = workbook.create_sheet(title="ThốngKê")
                            
                            # Tiêu đề
                            ws_stats.merge_cells('A1:D1')
                            title_cell = ws_stats.cell(row=1, column=1)
                            title_cell.value = f"THỐNG KÊ ĐIỂM DANH - NGÀY {date}"
                            title_cell.font = Font(bold=True, size=14, color="366092")
                            title_cell.alignment = Alignment(horizontal="center")
                            
                            # Thống kê cơ bản
                            stats_data = []
                            
                            # Đếm trạng thái
                            present_count = len([d for d in attendance_data if "✅" in str(d.get("Điểm danh", ""))])
                            absent_count = len([d for d in attendance_data if "❌" in str(d.get("Điểm danh", ""))])
                            late_count = len([d for d in attendance_data if "⚠️" in str(d.get("Điểm danh", ""))])
                            
                            total_students = len(attendance_data)
                            if attendance_data[0].get("Họ và tên") == "Không có dữ liệu điểm danh":
                                total_students = 0
                            
                            stats_data.append({
                                "Chỉ số": "Tổng số học sinh",
                                "Giá trị": total_students,
                                "Tỷ lệ": "100%"
                            })
                            
                            if present_count > 0:
                                stats_data.append({
                                    "Chỉ số": "Học sinh có mặt",
                                    "Giá trị": present_count,
                                    "Tỷ lệ": f"{(present_count/total_students*100):.1f}%" if total_students > 0 else "0%"
                                })
                            
                            if absent_count > 0:
                                stats_data.append({
                                    "Chỉ số": "Học sinh vắng mặt",
                                    "Giá trị": absent_count,
                                    "Tỷ lệ": f"{(absent_count/total_students*100):.1f}%" if total_students > 0 else "0%"
                                })
                            
                            if late_count > 0:
                                stats_data.append({
                                    "Chỉ số": "Học sinh đến muộn",
                                    "Giá trị": late_count,
                                    "Tỷ lệ": f"{(late_count/total_students*100):.1f}%" if total_students > 0 else "0%"
                                })
                            
                            # Thống kê từ analytics nếu có
                            if student_analytics:
                                avg_class_focus = np.mean([a['avg_focus_score'] for a in student_analytics.values()])
                                avg_class_behavior = np.mean([a['avg_behavior_score'] for a in student_analytics.values()])
                                
                                stats_data.append({
                                    "Chỉ số": "Điểm tập trung TB lớp",
                                    "Giá trị": round(avg_class_focus, 1),
                                    "Tỷ lệ": f"{'✅ Tốt' if avg_class_focus >= 70 else '⚠️ Cần cải thiện'}"
                                })
                                
                                stats_data.append({
                                    "Chỉ số": "Điểm hành vi TB lớp",
                                    "Giá trị": round(avg_class_behavior, 1),
                                    "Tỷ lệ": f"{'✅ Tốt' if avg_class_behavior >= 70 else '⚠️ Cần cải thiện'}"
                                })
                                
                                # Đếm cảm xúc tổng
                                all_emotions = {}
                                for analytics in student_analytics.values():
                                    for emotion, count in analytics['emotion_summary'].items():
                                        all_emotions[emotion] = all_emotions.get(emotion, 0) + count
                                
                                if all_emotions:
                                    top_emotion = max(all_emotions.items(), key=lambda x: x[1])
                                    stats_data.append({
                                        "Chỉ số": "Cảm xúc phổ biến nhất",
                                        "Giá trị": top_emotion[0],
                                        "Tỷ lệ": f"{top_emotion[1]} lần"
                                    })
                            
                            # Ghi dữ liệu
                            for i, stat in enumerate(stats_data):
                                ws_stats.cell(row=3 + i, column=1, value=stat["Chỉ số"])
                                ws_stats.cell(row=3 + i, column=2, value=stat["Giá trị"])
                                ws_stats.cell(row=3 + i, column=3, value=stat["Tỷ lệ"])
                            
                            # Định dạng
                            for col in ['A', 'B', 'C']:
                                ws_stats.column_dimensions[col].width = 25
                            
                            # Style cho header thống kê
                            for col_num in range(1, 4):
                                cell = ws_stats.cell(row=2, column=col_num)
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center")
                            
                            logger.info("✅ Sheet 4 (ThốngKê) created")
                            
                        except Exception as stats_error:
                            logger.warning(f"⚠️ Could not create statistics sheet: {stats_error}")
                    
                    stream.seek(0)
                    file_size = len(stream.getvalue())
                    logger.info(f"✅ Excel file created successfully: {file_size} bytes")
                    logger.info(f"📊 File contains {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
                    
                    return StreamingResponse(
                        iter([stream.getvalue()]),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={
                            "Content-Disposition": f"attachment; filename={filename}",
                            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        }
                    )
                    
                except Exception as excel_error:
                    logger.error(f"❌ Excel creation error: {excel_error}", exc_info=True)
                    raise HTTPException(
                        status_code=500, 
                        detail=f"Lỗi tạo file Excel: {str(excel_error)}"
                    )
                    
            except Exception as query_error:
                logger.error(f"❌ Database query error: {query_error}", exc_info=True)
                raise HTTPException(
                    status_code=500, 
                    detail=f"Lỗi truy vấn database: {str(query_error)}"
                )
        
        elif report_type == "engagement":
            raise HTTPException(
                status_code=400, 
                detail="Loại báo cáo độ tập trung đang được phát triển"
            )
        elif report_type == "feedback":
            # Báo cáo chuyên về feedback
            raise HTTPException(
                status_code=400, 
                detail="Loại báo cáo feedback đang được phát triển"
            )
        else:
            raise HTTPException(
                status_code=400, 
                detail=f"Loại báo cáo không hợp lệ: {report_type}. Chỉ hỗ trợ 'attendance'"
            )
            
    except HTTPException as http_err:
        logger.error(f"❌ HTTP Exception in export_reports: {http_err.detail}")
        raise
    except Exception as e:
        logger.error(f"❌ Critical error in export_reports: {e}", exc_info=True)
        raise HTTPException(
            status_code=500, 
            detail=f"Lỗi xuất báo cáo: {str(e)}"
        )

# ==================== REAL-TIME ENDPOINTS ====================
@app.post("/api/realtime/update")
async def realtime_update(data: Dict[str, Any], db: Session = Depends(get_db)):
    """API nhận real-time data từ Python AI model - BỎ QUA UNKNOWN"""
    try:
        logger.info(f"📡 Received real-time update: {data.get('type', 'unknown')}")
        
        # Broadcast qua WebSocket
        await manager_ws.broadcast(data)
        
        # Lưu dữ liệu vào database nếu cần
        if data.get('type') == 'attendance':
            attendance_data = data.get('data', {})
            if attendance_data:
                student_name = attendance_data.get('student_name', '')
                student_id = attendance_data.get('student_id', '')
                
                # KIỂM TRA NẾU LÀ UNKNOWN STUDENT
                if is_unknown_student(student_name, student_id):
                    logger.info(f"⏭️ Skipping unknown student in realtime update: {student_name}")
                    return {"status": "success", "message": "Skipped unknown student"}
                
                student_data = StudentData(
                    student_id=student_id,
                    student_name=student_name,
                    date=datetime.now().strftime("%Y-%m-%d"),
                    attendance_status=attendance_data.get('status', 'present'),
                    check_in_time=datetime.now(),
                    recorded_at=datetime.now()
                )
                db.add(student_data)
                db.commit()
        
        return {"status": "success", "message": "Real-time data received"}
        
    except Exception as e:
        logger.error(f"❌ Error in realtime_update: {e}")
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

# ==================== AI BATCH PROCESS ENDPOINT ====================
@app.api_route("/api/ai/batch-process", methods=["GET", "POST"])
async def ai_batch_process(
    request: Request,
    batch_data: Optional[Dict[str, Any]] = None,
    db: Session = Depends(get_db)
):
    """AI Batch Process endpoint - ĐÃ FIX HOÀN TOÀN - BỎ QUA UNKNOWN"""
    
    # Xử lý GET request (cho testing)
    if request.method == "GET":
        try:
            # Lấy dữ liệu mẫu từ database để demo
            recent_data = db.query(StudentData).filter(
                StudentData.recorded_by.like("%AI%"),
                StudentData.student_name.isnot(None)
            ).order_by(StudentData.recorded_at.desc()).limit(5).all()
            
            # Filter out unknown students
            valid_data = []
            for record in recent_data:
                if not is_unknown_student(record.student_name, record.student_id):
                    valid_data.append(record)
            
            recent_examples = []
            for record in valid_data:
                recent_examples.append({
                    "student_id": record.student_id,
                    "student_name": record.student_name,
                    "emotion": record.emotion,
                    "behavior_type": record.behavior_type,
                    "behavior_details": record.behavior_details,
                    "focus_score": record.focus_score,
                    "recorded_at": record.recorded_at.isoformat() if record.recorded_at else None
                })
            
            # Thống kê dữ liệu AI
            ai_data_count = db.query(StudentData).filter(
                StudentData.recorded_by.like("%AI%"),
                StudentData.student_name.isnot(None)
            ).count()
            
            today_ai_count = db.query(StudentData).filter(
                StudentData.recorded_by.like("%AI%"),
                StudentData.date == datetime.now().strftime("%Y-%m-%d"),
                StudentData.student_name.isnot(None)
            ).count()
            
            return {
                "status": "success",
                "method": "GET",
                "endpoint_info": {
                    "name": "AI Batch Process Endpoint",
                    "primary_method": "POST",
                    "description": "Nhận batch data từ AI model và lưu vào database",
                    "supported_ai_systems": [
                        "Face Recognition System",
                        "Emotion Detection AI",
                        "Behavior Analysis AI",
                        "Focus/Engagement Tracking AI"
                    ],
                    "unknown_filtering": "ENABLED - All unknown students will be skipped"
                },
                "stats": {
                    "total_ai_records": ai_data_count,
                    "today_ai_records": today_ai_count,
                    "active_sessions": len([ws for ws in manager_ws.active_connections])
                },
                "example_payload": {
                    "type": "batch_detection",
                    "session_id": "ai_session_20241215_001",
                    "timestamp": datetime.now().isoformat(),
                    "data": [
                        {
                            "student_id": "AI_STUDENT_001",
                            "student_name": "Nam",
                            "student_code": "AI001",
                            "name": "Nam",
                            "date": datetime.now().strftime("%Y-%m-%d"),
                            "attendance_status": "present",
                            "status": "present",
                            "check_in_time": datetime.now().isoformat(),  # ISO format string
                            "emotion": "happy",
                            "emotion_confidence": 0.85,
                            "confidence": 0.85,
                            "behavior_type": "engagement",
                            "behavior_score": 88.5,
                            "score": 88.5,
                            "behavior_details": "writing notes",
                            "details": "writing notes",
                            "behavior": "writing",
                            "focus_score": 90.0,
                            "engagement": 90.0,
                            "concentration_level": "high",
                            "focus_duration": 45.0,
                            "class_name": "AI Class",
                            "session_id": "ai_session_001",
                            "recorded_by": "AI System"
                        }
                    ]
                },
                "recent_ai_data": recent_examples,
                "testing_instructions": {
                    "curl_example": 'curl -X POST "http://localhost:8000/api/ai/batch-process" -H "Content-Type: application/json" -d \'{"type":"test","data":[{"student_name":"Test","emotion":"happy"}]}\'',
                    "python_example": "import requests\nrequests.post('http://localhost:8000/api/ai/batch-process', json={'type':'test','data':[]})"
                }
            }
            
        except Exception as e:
            logger.error(f"GET handler error: {e}")
            return JSONResponse(
                status_code=500,
                content={
                    "status": "error",
                    "method": "GET",
                    "message": f"Error in GET handler: {str(e)}",
                    "basic_info": {
                        "endpoint": "/api/ai/batch-process",
                        "supported_methods": ["POST"],
                        "purpose": "Process batch data from AI systems",
                        "unknown_filtering": "ENABLED"
                    }
                }
            )
    
    # ==================== XỬ LÝ POST REQUEST ====================
    try:
        logger.info(f"📦 Received AI batch data via POST: type={batch_data.get('type', 'unknown') if batch_data else 'no data'}")
        
        if not batch_data:
            return {
                "status": "error",
                "message": "No data provided in POST request",
                "timestamp": datetime.now().isoformat()
            }
        
        # Lấy thông tin từ batch data
        data_type = batch_data.get('type', 'batch_detection')
        session_id = batch_data.get('session_id', f"session_{int(time.time())}")
        timestamp = batch_data.get('timestamp', datetime.now().isoformat())
        data_items = batch_data.get('data', [])
        
        logger.info(f"📊 Processing {len(data_items)} items from session: {session_id}")
        
        # ==================== FILTER OUT UNKNOWN STUDENTS ====================
        filtered_items = []
        unknown_count = 0
        
        for item in data_items:
            # Kiểm tra nếu là "unknown" student
            student_name = item.get('student_name') or item.get('name') or ""
            student_id = item.get('student_id') or item.get('student_code') or ""
            
            # Kiểm tra xem có phải unknown student không
            if is_unknown_student(student_name, student_id):
                unknown_count += 1
                logger.info(f"⏭️ Filtering out unknown student: {student_name}")
                continue
            
            filtered_items.append(item)
        
        logger.info(f"🔄 After filtering: {len(filtered_items)} valid items, {unknown_count} unknown items filtered out")
        
        # Nếu tất cả đều là unknown, trả về message
        if unknown_count == len(data_items):
            return {
                "status": "success",
                "message": f"All {unknown_count} items are unknown students, nothing to save",
                "unknown_filtered": unknown_count,
                "processed_count": 0,
                "timestamp": datetime.now().isoformat()
            }
        
        # Sử dụng danh sách đã filter
        data_items = filtered_items
        
        # ==================== TIẾP TỤC XỬ LÝ NHƯ BÌNH THƯỜNG ====================
        success_count = 0
        failed_count = 0
        
        # ==================== HÀM PARSE DATETIME NỘI BỘ ====================
        def parse_datetime_safe(value):
            """Chuyển đổi string thành datetime object - SAFE VERSION"""
            if value is None:
                return None
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                try:
                    # Thử parse ISO format
                    if 'T' in value:
                        # Format: 2025-12-22T16:09:12.675711
                        try:
                            return datetime.fromisoformat(value.replace('Z', '+00:00'))
                        except ValueError:
                            # Thử parse với milliseconds
                            if '.' in value:
                                try:
                                    return datetime.strptime(value, "%Y-%m-%dT%H:%M:%S.%f")
                                except:
                                    return datetime.strptime(value, "%Y-%m-%dT%H:%M:%S")
                            else:
                                return datetime.strptime(value, "%Y-%m-%dT%H:%M:%S")
                    else:
                        # Other formats
                        formats = [
                            "%Y-%m-%d %H:%M:%S.%f",
                            "%Y-%m-%d %H:%M:%S",
                            "%Y-%m-%d %H:%M",
                            "%Y-%m-%d"
                        ]
                        for fmt in formats:
                            try:
                                return datetime.strptime(value, fmt)
                            except ValueError:
                                continue
                except Exception as e:
                    logger.warning(f"Cannot parse datetime: {value}, error: {e}")
                    return None
            return None
        
        # ==================== XỬ LÝ TỪNG ITEM ====================
        for i, item in enumerate(data_items):
            try:
                # Extract data từ item với các alias
                student_id = item.get('student_id') or item.get('student_code') or f"AI_{int(time.time()) % 10000:04d}"
                student_name = item.get('student_name') or item.get('name') or ""
                
                if not student_id or not student_name:
                    logger.warning(f"⚠️ Item {i+1} missing student_id or student_name")
                    failed_count += 1
                    continue
                
                # 🔴 DEBUG LOG
                logger.debug(f"🔄 Processing item {i+1}: {student_name}, "
                           f"check_in_time: {item.get('check_in_time')}, "
                           f"emotion: {item.get('emotion')}, "
                           f"behavior: {item.get('behavior')}")
                
                # ==================== PARSE DATETIME FIELDS ====================
                # Parse check_in_time (BẮT BUỘC có giá trị)
                check_in_time_raw = item.get('check_in_time')
                if check_in_time_raw:
                    check_in_time = parse_datetime_safe(check_in_time_raw)
                    if check_in_time is None:
                        # Nếu không parse được, dùng thời gian hiện tại
                        logger.warning(f"⚠️ Cannot parse check_in_time: {check_in_time_raw}, using current time")
                        check_in_time = datetime.now()
                else:
                    # Nếu không có check_in_time, dùng thời gian hiện tại
                    check_in_time = datetime.now()
                
                # Parse check_out_time (có thể None)
                check_out_time_raw = item.get('check_out_time')
                check_out_time = parse_datetime_safe(check_out_time_raw) if check_out_time_raw else None
                
                # ==================== MAP FIELDS VỚI ALIAS ====================
                attendance_status = item.get('attendance_status') or item.get('status') or 'present'
                
                # Emotion fields (đảm bảo không null)
                emotion = item.get('emotion', 'neutral')
                emotion_confidence = item.get('emotion_confidence') or item.get('confidence') or 0.5
                
                # Behavior fields (đảm bảo không null)
                behavior_type = item.get('behavior_type') or 'engagement'
                behavior_score = item.get('behavior_score') or item.get('score') or 75.0
                
                # Behavior details (lấy từ nhiều alias)
                behavior_details = (
                    item.get('behavior_details') or 
                    item.get('details') or 
                    item.get('behavior') or 
                    item.get('behavior_text') or 
                    'Auto-detected by AI'
                )
                
                # Focus/Engagement fields
                focus_score = item.get('focus_score') or item.get('engagement') or 75.0
                concentration_level = item.get('concentration_level') or 'medium'
                attendance_notes = item.get('attendance_notes') or item.get('notes') or f"AI Auto: {emotion} emotion, {behavior_details}"
                
                # ==================== VALIDATE VÀ NORMALIZE DỮ LIỆU ====================
                # Validate và giới hạn focus_score (0-100)
                try:
                    focus_score_float = float(focus_score)
                    if focus_score_float > 100:
                        logger.warning(f"Focus score {focus_score_float} > 100, capping to 100")
                        focus_score_float = 100.0
                    if focus_score_float < 0:
                        logger.warning(f"Focus score {focus_score_float} < 0, setting to 0")
                        focus_score_float = 0.0
                except (ValueError, TypeError):
                    logger.warning(f"Invalid focus score: {focus_score}, using default 75.0")
                    focus_score_float = 75.0
                
                # Validate behavior_score (0-100)
                try:
                    behavior_score_float = float(behavior_score)
                    if behavior_score_float > 100:
                        behavior_score_float = 100.0
                    if behavior_score_float < 0:
                        behavior_score_float = 0.0
                except (ValueError, TypeError):
                    behavior_score_float = 75.0
                
                # Validate emotion_confidence (0-1)
                try:
                    emotion_confidence_float = float(emotion_confidence)
                    if emotion_confidence_float > 1:
                        emotion_confidence_float = 1.0
                    if emotion_confidence_float < 0:
                        emotion_confidence_float = 0.0
                except (ValueError, TypeError):
                    emotion_confidence_float = 0.5
                
                # ==================== TẠO BẢN GHI DATABASE ====================
                student_record = StudentData(
                    # Student info
                    student_id=student_id,
                    student_name=student_name,
                    
                    # Date and time
                    date=item.get('date') or datetime.now().strftime("%Y-%m-%d"),
                    check_in_time=check_in_time,  # 🔴 ĐẢM BẢO LÀ DATETIME OBJECT
                    check_out_time=check_out_time,
                    
                    # Attendance data
                    attendance_status=attendance_status,
                    attendance_notes=attendance_notes,
                    
                    # Emotion data (đảm bảo không null)
                    emotion=emotion,
                    emotion_confidence=emotion_confidence_float,
                    
                    # Behavior data (đảm bảo không null)
                    behavior_type=behavior_type,
                    behavior_score=behavior_score_float,
                    behavior_details=behavior_details,
                    
                    # Focus/Engagement data
                    focus_score=focus_score_float,
                    concentration_level=concentration_level,
                    focus_duration=item.get('focus_duration') or 45.0,
                    
                    # Metadata
                    class_name=item.get('class_name') or 'AI Class',
                    session_id=session_id,
                    recorded_by=item.get('recorded_by') or "AI System",
                    recorded_at=datetime.now()
                )
                
                # Thêm vào database
                db.add(student_record)
                success_count += 1
                
                # Log thành công
                if i < 3:  # Chỉ log 3 cái đầu để không spam
                    logger.info(f"  ✅ Item {i+1}: {student_name} - {emotion} - {behavior_details} - Focus: {focus_score_float}")
                
            except Exception as e:
                logger.error(f"❌ Error processing item {i+1}: {e}")
                import traceback
                traceback.print_exc()
                failed_count += 1
                continue
        
        # ==================== COMMIT DATABASE ====================
        try:
            db.commit()
            logger.info(f"✅ Database committed successfully")
        except Exception as e:
            logger.error(f"❌ Database commit failed: {e}")
            db.rollback()
            failed_count = len(data_items)  # Đánh dấu tất cả thất bại
            success_count = 0
        
        # ==================== BROADCAST WEBSOCKET ====================
        try:
            await manager_ws.broadcast({
                "type": "batch_processed",
                "timestamp": datetime.now().isoformat(),
                "session_id": session_id,
                "processed_count": success_count,
                "total_count": len(data_items),
                "unknown_filtered": unknown_count,
                "message": f"Processed {success_count}/{len(data_items)} items successfully, filtered {unknown_count} unknown students"
            })
        except Exception as e:
            logger.warning(f"⚠️ WebSocket broadcast failed: {e}")
        
        # ==================== LOG SUMMARY ====================
        if len(data_items) > 3:
            logger.info(f"  ... and {len(data_items) - 3} more items")
        
        logger.info(f"✅ Batch processed via POST: {success_count} success, {failed_count} failed, {unknown_count} unknown filtered")
        
        # ==================== RETURN RESPONSE ====================
        return {
            "status": "success" if success_count > 0 else "partial_success",
            "method": "POST",
            "message": f"Batch processed: {success_count} success, {failed_count} failed, {unknown_count} unknown filtered",
            "session_id": session_id,
            "processed_count": success_count,
            "failed_count": failed_count,
            "unknown_filtered": unknown_count,
            "total_count": len(data_items) + unknown_count,
            "success_rate": f"{(success_count/len(data_items)*100 if data_items else 0):.1f}%" if data_items else "0%",
            "timestamp": datetime.now().isoformat(),
            "websocket_broadcast": True,
            "next_steps": [
                "Data saved to student_data table",
                f"View at: GET /api/student-data?session_id={session_id}",
                f"Total records: {db.query(StudentData).filter(StudentData.session_id == session_id).count()}"
            ]
        }
        
    except Exception as e:
        logger.error(f"❌ Critical error in ai_batch_process: {e}")
        import traceback
        traceback.print_exc()
        
        try:
            db.rollback()
        except:
            pass
        
        return JSONResponse(
            status_code=200,  # Luôn trả về 200 để AI không dừng
            content={
                "status": "error",
                "method": "POST",
                "message": f"Batch processing failed: {str(e)}",
                "timestamp": datetime.now().isoformat(),
                "suggestion": "Check datetime format and field names"
            }
        )

@app.post("/api/websocket/broadcast")
async def websocket_broadcast(data: Dict[str, Any]):
    """API để broadcast message qua WebSocket"""
    try:
        logger.info(f"📢 WebSocket broadcast: {data.get('type', 'unknown')}")
        
        await manager_ws.broadcast(data)
        
        return {"status": "success", "message": "Message broadcasted"}
        
    except Exception as e:
        logger.error(f"❌ Error in websocket_broadcast: {e}")
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

# ==================== WEBSOCKET ENDPOINT ====================
@app.websocket("/ws/live")
async def websocket_endpoint(websocket: WebSocket):
    await manager_ws.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            # Xử lý dữ liệu từ client nếu cần
            await websocket.send_text(f"Message received: {data}")
    except WebSocketDisconnect:
        manager_ws.disconnect(websocket)
    except Exception as e:
        logger.error(f"WebSocket error: {e}")
        manager_ws.disconnect(websocket)

# ==================== SYSTEM ENDPOINTS ====================
@app.get("/api/system/health")
async def system_health_check():
    """Health check endpoint (detailed)"""
    try:
        # Kiểm tra database connection
        db = SessionLocal()
        db.execute("SELECT 1")
        
        # Kiểm tra các bảng
        tables = ["users", "student_data"]
        table_status = {}
        
        for table in tables:
            try:
                db.execute(f"SELECT COUNT(*) FROM {table}")
                table_status[table] = "OK"
            except Exception as e:
                table_status[table] = f"Error: {str(e)}"
        
        db.close()
        
        return {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "database": "connected",
            "websocket_connections": len(manager_ws.active_connections),
            "tables": table_status,
            "service": "Classroom Management System API",
            "version": "1.0.0",
            "unknown_filtering": "ENABLED"
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy",
                "timestamp": datetime.now().isoformat(),
                "database": "disconnected",
                "error": str(e),
                "service": "Classroom Management System API"
            }
        )

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    try:
        # Kiểm tra database connection
        db = SessionLocal()
        db.execute("SELECT 1")
        db.close()
        
        return {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "database": "connected",
            "websocket_connections": len(manager_ws.active_connections),
            "service": "Classroom Management System API",
            "unknown_filtering": "ENABLED"
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return {
            "status": "unhealthy",
            "timestamp": datetime.now().isoformat(),
            "database": "disconnected",
            "error": str(e)
        }

# ==================== OTHER ENDPOINTS ====================
@app.get("/")
async def root():
    return JSONResponse({
        "message": "Classroom Management System API",
        "version": "1.0.0",
        "description": "Hệ thống quản lý lớp học với điểm danh và theo dõi hành vi, cảm xúc, độ tập trung",
        "authentication_required": "Hầu hết endpoints cần token",
        "user_types": {
            "admin": "Truy cập đầy đủ tính năng",
            "user": "Truy cập giới hạn, chỉ xem thông tin lớp"
        },
        "unknown_filtering": "ENABLED - All unknown students are automatically filtered out",
        "check_permission": "GET /api/auth/check?token=YOUR_TOKEN",
        "endpoints": {
            "authentication": {
                "login": "POST /api/auth/login",
                "register": "POST /api/auth/register",
                "demo_login": "POST /api/auth/demo-login",
                "check_auth": "GET /api/auth/check",
                "user_dashboard": "GET /api/user/dashboard"
            },
            "student_data": {
                "get_student_data": "GET /api/student-data",
                "create_student_data": "POST /api/student-data",
                "update_student_data": "PUT /api/student-data/{id}",
                "delete_student_data": "DELETE /api/student-data/{id}"
            },
            "attendance": {
                "get_attendance": "GET /api/attendance",
                "create_attendance": "POST /api/attendance",
                "check_out": "PUT /api/attendance/{id}/checkout"
            },
            "emotion": {
                "get_emotion": "GET /api/emotion",
                "create_emotion": "POST /api/emotion"
            },
            "behavior": {
                "get_behavior": "GET /api/behavior",
                "create_behavior": "POST /api/behavior"
            },
            "focus": {
                "get_focus": "GET /api/focus",
                "create_focus": "POST /api/focus"
            },
            "ai_endpoints": {
                "ai_attendance": "POST /api/ai/attendance",
                "ai_emotion": "POST /api/ai/emotion",
                "ai_behavior": "POST /api/ai/behavior",
                "ai_focus": "POST /api/ai/focus",
                "ai_student_data": "POST /api/ai/student-data",
                "ai_batch_process": "POST /api/ai/batch-process"
            },
            "dashboard": {
                "stats": "GET /api/dashboard/stats",
                "attendance_chart": "GET /api/dashboard/attendance-chart",
                "emotion_chart": "GET /api/dashboard/emotion-chart",
                "focus_chart": "GET /api/dashboard/focus-chart"
            },
            "analytics": {
                "emotion_trend": "GET /api/analytics/emotion-trend",
                "focus_analytics": "GET /api/analytics/focus-analytics"
            },
            "reports": {
                "get_reports": "GET /api/reports",
                "export_reports": "GET /api/reports/export"
            },
            "realtime": {
                "websocket": "ws://localhost:8000/ws/live",
                "update": "POST /api/realtime/update",
                "broadcast": "POST /api/websocket/broadcast"
            },
            "system": {
                "health": "GET /api/system/health"
            },
            "student_management": {
                "get_students_list": "GET /api/students/list",
                "create_student": "POST /api/students",
                "update_student": "PUT /api/students/{student_id}",
                "import_students": "POST /api/students/import"
            },
            "attendance_summary": {
                "get_attendance_summary": "GET /api/attendance/summary",
                "get_daily_attendance": "GET /api/attendance/daily"
            },
        },
        "demo_account": {
            "username": "demo",
            "password": "demo123"
        }
    })

# ==================== INITIALIZATION ====================
def create_sample_data():
    """Tạo dữ liệu mẫu"""
    db = SessionLocal()
    try:
        print("=" * 60)
        print("🔄 ĐANG TẠO DỮ LIỆU MẪU CHO DATABASE MỚI")
        print("=" * 60)
        
        # Xóa tất cả dữ liệu cũ nếu có (trong trường hợp reset)
        print("🧹 Đang làm sạch dữ liệu cũ...")
        db.query(StudentData).delete()
        db.query(ClassStudent).delete()
        db.query(User).delete()
        db.commit()
        
        print("👤 Đang tạo tài khoản người dùng...")
        
        # Tạo users mẫu
        users = [
            User(
                username="admin",
                email="admin@school.edu.vn",
                hashed_password=hash_password("admin123"),
                full_name="Administrator",
                is_active=True,
                is_admin=True,
                created_at=datetime.now()
            ),
            User(
                username="teacher1",
                email="teacher1@school.edu.vn",
                hashed_password=hash_password("teacher123"),
                full_name="Nguyễn Văn Giáo",
                is_active=True,
                is_admin=False,
                created_at=datetime.now()
            ),
            User(
                username="demo",
                email="demo@school.edu.vn",
                hashed_password=hash_password("demo123"),
                full_name="Demo Teacher",
                is_active=True,
                is_admin=False,
                created_at=datetime.now()
            ),
        ]
        
        for user in users:
            db.add(user)
        db.commit()
        
        # Tạo student data mẫu
        today = datetime.now().strftime("%Y-%m-%d")
        
        student_data_samples = [
            # Dữ liệu điểm danh
            StudentData(
                student_id="SV001",
                student_name="Dino",
                date=today,
                emotion="happy",
                emotion_confidence=0.85,
                attendance_status="absent",
                check_in_time=datetime.now().replace(hour=7, minute=30, second=0),
                class_name="STEM 1",
                behavior_type="look_straight",
                recorded_by="teacher1",
                recorded_at=datetime.now(),
                focus_score=34.56
            ),
            StudentData(
                student_id="SV003",
                student_name="Thinh",
                date=today,
                emotion="sad",
                emotion_confidence=0.85,
                attendance_status="absent",
                check_in_time=datetime.now().replace(hour=7, minute=35, second=0),
                class_name="STEM 1",
                behavior_type="writing",
                recorded_by="teacher1",
                recorded_at=datetime.now(),
                focus_score=54.76
            ),
            StudentData(
                student_id="SV002",
                student_name="Minh",
                date=today,
                emotion="angry",
                emotion_confidence=0.85,
                attendance_status="absent",
                check_in_time=datetime.now().replace(hour=8, minute=0, second=0),
                attendance_notes="Đến muộn 30 phút",
                class_name="STEM 1",
                behavior_type="look_around",
                recorded_by="teacher1",
                recorded_at=datetime.now(),
                focus_score=62.57
            ),
            # Dữ liệu cảm xúc
            StudentData(
                student_id="SV004",
                student_name="Mini",
                date=today,
                emotion="happy",
                attendance_status="absent",
                emotion_confidence=0.95,
                class_name="STEM 1",
                behavior_type="writing",
                session_id="SESS001",
                recorded_by="teacher1",
                recorded_at=datetime.now()
            ),
            StudentData(
                student_id="SV005",
                student_name="Khoa",
                date=today,
                emotion="neutral",
                attendance_status="absent",
                emotion_confidence=0.72,
                class_name="STEM 1",
                behavior_type="look_straight",
                session_id="SESS001",
                recorded_by="teacher1",
                recorded_at=datetime.now(),
                focus_score=30.16
            ),
            StudentData(
                student_id="SV006",
                student_name="Nam",
                date=today,
                emotion="neutral",
                attendance_status="absent",
                emotion_confidence=0.72,
                class_name="STEM 1",
                behavior_type="writing",
                session_id="SESS001",
                recorded_by="teacher1",
                recorded_at=datetime.now(),
                focus_score=84.96
            ),
            StudentData(
                student_id="SV007",
                student_name="Thanh",
                date=today,
                emotion="neutral",
                attendance_status="absent",
                emotion_confidence=0.82,
                class_name="STEM 1",
                behavior_type="look_around",
                session_id="SESS001",
                recorded_by="teacher1",
                recorded_at=datetime.now(),
                focus_score=50.46
            ),
        ]
        
        for data in student_data_samples:
            db.add(data)
        db.commit()
        
        # Tạo danh sách lớp học cố định
        class_students_samples = [
            ClassStudent(
                student_id="SV001",
                student_name="Dino",
                student_code="2024001",
                class_name="Lớp 10A1",
                gender="Nam",
                date_of_birth="2008-05-15",
                is_active=True,
                enrollment_date="2024-09-01"
            ),
            ClassStudent(
                student_id="SV002",
                student_name="Minh",
                student_code="2024002",
                class_name="Lớp 10A1",
                gender="Nam",
                date_of_birth="2008-07-22",
                is_active=True,
                enrollment_date="2024-09-01"
            ),
            ClassStudent(
                student_id="SV003",
                student_name="Thinh",
                student_code="2024003",
                class_name="Lớp 10A1",
                gender="Nam",
                date_of_birth="2008-03-10",
                is_active=True,
                enrollment_date="2024-09-01"
            ),
            ClassStudent(
                student_id="SV004",
                student_name="Mini",
                student_code="2024025",
                class_name="Lớp 10A1",
                gender="Nữ",
                date_of_birth="2008-11-30",
                is_active=True,
                enrollment_date="2024-09-01"
            ),
            ClassStudent(
                student_id="SV005",
                student_name="Khoa",
                student_code="2024025",
                class_name="Lớp 10A1",
                gender="Nam",
                date_of_birth="2008-11-30",
                is_active=True,
                enrollment_date="2024-09-01"
            ),
             ClassStudent(
                student_id="SV006",
                student_name="Nam",
                student_code="2024025",
                class_name="Lớp 10A1",
                gender="Nam",
                date_of_birth="2008-11-30",
                is_active=True,
                enrollment_date="2024-09-01"
            ),
             ClassStudent(
                student_id="SV007",
                student_name="Thanh",
                student_code="2024025",
                class_name="Lớp 10A1",
                gender="Nam",
                date_of_birth="2008-11-30",
                is_active=True,
                enrollment_date="2024-09-01"
            ),
        ]
        
        # Kiểm tra xem đã có dữ liệu chưa
        existing_class_students = db.query(ClassStudent).count()
        if existing_class_students == 0:
            for student in class_students_samples:
                db.add(student)
            print(f"✅ Đã tạo {len(class_students_samples)} học sinh mẫu")
        
        db.commit()
        
        print("✅ Dữ liệu mẫu đã được tạo thành công!")
        
    except Exception as e:
        print(f"❌ Lỗi khi tạo dữ liệu mẫu: {e}")
        db.rollback()
    finally:
        db.close()

# Khởi tạo database với dữ liệu mẫu
create_sample_data()

# Thay thế endpoint cũ bằng cái này
@app.post("/api/system/reset-database")
async def reset_database_simple(
    request: ResetDatabaseRequest
):
    """
    Reset database đơn giản - không yêu cầu authentication
    """
    try:
        logger.info("🔄 SIMPLE DATABASE RESET REQUEST RECEIVED")
        
        if not request.confirm:
            raise HTTPException(status_code=400, detail="Vui lòng xác nhận reset database")
        
        print("="*60)
        print("🔄 SIMPLE DATABASE RESET - NO AUTH REQUIRED")
        print("="*60)
        
        db_file_path = "classroom_ai.db"
        backup_created = False
        backup_path = None
        
        # 1. Đóng tất cả kết nối database trước
        logger.info("1. Đang đóng kết nối database...")
        try:
            # Đóng session hiện tại nếu có
            db = SessionLocal()
            db.close()
            
            # Dispose engine
            engine.dispose()
            time.sleep(1)  # Chờ 1 giây
            logger.info("✅ Đã đóng kết nối database")
        except Exception as e:
            logger.warning(f"⚠️ Lỗi khi đóng kết nối: {e}")
            # Vẫn tiếp tục dù có lỗi
        
        # 2. Tạo backup nếu có file cũ
        if os.path.exists(db_file_path):
            try:
                backup_dir = "database_backups"
                os.makedirs(backup_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f"classroom_ai_backup_{timestamp}.db"
                backup_path = os.path.join(backup_dir, backup_filename)
                
                logger.info(f"2. Đang tạo backup: {backup_path}")
                shutil.copy2(db_file_path, backup_path)
                backup_created = True
                logger.info(f"✅ Đã tạo backup thành công")
            except Exception as e:
                logger.error(f"❌ Không thể tạo backup: {e}")
                # Vẫn tiếp tục dù không tạo được backup
        
        # 3. Xóa file database cũ
        try:
            if os.path.exists(db_file_path):
                logger.info("3. Đang xóa database file cũ...")
                
                # Thử nhiều lần nếu file bị lock
                max_retries = 5
                for attempt in range(max_retries):
                    try:
                        os.remove(db_file_path)
                        logger.info(f"✅ Đã xóa database file: {db_file_path}")
                        break
                    except PermissionError:
                        if attempt < max_retries - 1:
                            logger.warning(f"⚠️ File đang bị lock, thử lại lần {attempt + 1}...")
                            time.sleep(1)
                            continue
                        else:
                            raise Exception("Không thể xóa file database, có thể đang bị sử dụng")
                    except FileNotFoundError:
                        logger.info("✅ Database file không tồn tại (đã bị xóa trước đó)")
                        break
            else:
                logger.info("ℹ️ Database file không tồn tại, bỏ qua bước xóa")
                
        except Exception as e:
            logger.error(f"❌ Lỗi xóa database file: {e}")
            return JSONResponse(
                status_code=500,
                content={
                    "status": "error",
                    "message": f"Không thể xóa database file: {str(e)}",
                    "suggestion": "Đảm bảo không có chương trình nào đang sử dụng file classroom_ai.db"
                }
            )
        
        # 4. Tạo database mới và schema
        try:
            logger.info("4. Đang tạo database schema mới...")
            
            # Tạo lại engine mới
            new_engine = create_engine(
                SQLALCHEMY_DATABASE_URL,
                connect_args={"check_same_thread": False},
                pool_pre_ping=True,
                pool_recycle=3600,
            )
            
            # Tạo tất cả bảng
            Base.metadata.create_all(bind=new_engine)
            logger.info("✅ Đã tạo database schema mới")
            
        except Exception as e:
            logger.error(f"❌ Lỗi tạo database schema: {e}")
            return JSONResponse(
                status_code=500,
                content={
                    "status": "error",
                    "message": f"Lỗi tạo database schema: {str(e)}",
                    "suggestion": "Kiểm tra quyền ghi trong thư mục hiện tại"
                }
            )
        
        # 5. Tạo dữ liệu mẫu
        sample_data_created = False
        if request.create_sample_data:
            try:
                logger.info("5. Đang tạo dữ liệu mẫu...")
                
                # Tạo session mới
                db = SessionLocal()
                
                # Gọi hàm create_sample_data
                create_sample_data()
                
                sample_data_created = True
                logger.info("✅ Đã tạo dữ liệu mẫu thành công")
                
            except Exception as e:
                logger.error(f"❌ Lỗi tạo dữ liệu mẫu: {e}")
                import traceback
                traceback.print_exc()
                sample_data_created = False
                # Vẫn trả về success nhưng cảnh báo
        
        # 6. Broadcast qua WebSocket
        try:
            await manager_ws.broadcast({
                "type": "database_reset",
                "timestamp": datetime.now().isoformat(),
                "message": "Database đã được reset thành công",
                "sample_data_created": sample_data_created
            })
        except:
            pass
        
        # 7. Trả về kết quả thành công
        logger.info("✅ DATABASE RESET COMPLETED SUCCESSFULLY")
        
        return {
            "status": "success",
            "message": "Database đã được reset và tạo mới thành công!",
            "timestamp": datetime.now().isoformat(),
            "actions": {
                "database_deleted": True,
                "database_created": True,
                "schema_created": True,
                "sample_data_created": sample_data_created,
                "backup_created": backup_created,
                "backup_path": backup_path
            },
            "demo_accounts": {
                "demo": {"username": "demo", "password": "demo123", "role": "teacher"},
                "admin": {"username": "admin", "password": "admin123", "role": "admin"}
            },
            "next_steps": [
                "Reload trang dashboard",
                f"Đăng nhập với: demo / demo123",
                f"Admin: admin / admin123" if sample_data_created else "Không có dữ liệu mẫu"
            ]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ CRITICAL ERROR IN DATABASE RESET: {e}")
        import traceback
        traceback.print_exc()
        
        return JSONResponse(
            status_code=500,
            content={
                "status": "error",
                "message": f"Lỗi nghiêm trọng khi reset database: {str(e)}",
                "timestamp": datetime.now().isoformat(),
                "debug_info": {
                    "db_file_exists": os.path.exists("classroom_ai.db") if 'os' in globals() else "unknown",
                    "python_version": sys.version if 'sys' in globals() else "unknown"
                }
            }
        )
        
print("\n" + "="*80)
print("🚀 CLASSROOM MANAGEMENT SYSTEM API (SIMPLIFIED) - AI COMPATIBLE")
print("="*80)
print("📊 Database: SQLite (classroom_ai.db)")
print("📁 Tables: 2 (users, student_data)")
print("🌐 API Documentation: http://localhost:8000/api/docs")
print("🔗 WebSocket: ws://localhost:8000/ws/live")
print("🔐 Demo Account: demo / demo123")
print("👑 Admin Account: admin / admin123")
print("🚫 UNKNOWN FILTERING: ENABLED - All unknown students are automatically skipped")
print("📋 New Authentication Endpoints:")
print("   • GET /api/auth/check?token=TOKEN  - Check user permissions")
print("   • GET /api/user/dashboard?token=TOKEN - User-specific dashboard")
print("🤖 AI-Compatible Endpoints:")
print("   • POST /api/ai/attendance      - Flexible attendance from AI")
print("   • POST /api/ai/emotion         - Flexible emotion from AI")
print("   • POST /api/ai/behavior        - Flexible behavior from AI")
print("   • POST /api/ai/focus           - Flexible focus from AI")
print("   • POST /api/ai/student-data    - All-in-one AI endpoint")
print("   • POST /api/ai/batch-process   - Batch processing for AI (RECOMMENDED)")
print("="*80)
print("📋 Available Endpoints:")
print("   • GET    /api/health             - Health check")
print("   • GET    /api/dashboard/stats    - Dashboard statistics (auto-detect user type)")
print("   • GET    /api/student-data       - All student data")
print("   • GET    /api/attendance         - Attendance records")
print("   • GET    /api/emotion            - Emotion data")
print("   • GET    /api/focus              - Focus/concentration data")
print("   • GET    /api/analytics/focus-analytics - Focus analytics")
print("   • POST   /api/ai/batch-process   - AI batch data (RECOMMENDED)")
print("="*80)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("database_server:app", host="0.0.0.0", port=8000, reload=True)    uvicorn.run("database_server:app", host="0.0.0.0", port=8000, reload=True)